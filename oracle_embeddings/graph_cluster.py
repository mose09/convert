import logging
from collections import defaultdict

logger = logging.getLogger(__name__)

MAX_GROUP_SIZE = 30  # 그룹당 최대 테이블 수


def find_groups(schema: dict, joins: list[dict],
                max_group_size: int = MAX_GROUP_SIZE,
                query_tables: set = None,
                common_threshold: int = None,
                common_tables_manual: set = None,
                table_usage: dict = None) -> tuple:
    """Find table groups based on JOIN relationships using connected components.

    query_tables: XML 쿼리에서 사용된 전체 테이블 목록 (JOIN 여부 무관)
    common_threshold: 이 수 이상의 다른 테이블과 JOIN되면 공통 테이블로 간주
    common_tables_manual: 수동 지정 공통 테이블 목록
    table_usage: mybatis_parser.extract_table_usage 결과 (as_main_count, as_join_count)
    """

    # Build adjacency list
    adj = defaultdict(set)
    for j in joins:
        adj[j["table1"]].add(j["table2"])
        adj[j["table2"]].add(j["table1"])

    all_tables = {t["name"] for t in schema.get("tables", [])}
    tables_in_joins = set(adj.keys())

    # Detect common/reference tables
    common_tables = set()
    if common_tables_manual:
        common_tables = common_tables_manual & tables_in_joins

    # Auto-detect using main vs join ratio from SQL analysis
    if not common_tables_manual and table_usage:
        for table in tables_in_joins:
            u = table_usage.get(table, {})
            main_count = u.get("as_main_count", 0)
            join_count = u.get("as_join_count", 0)
            total = main_count + join_count
            if total == 0:
                continue
            # JOIN으로만 사용되는 비율이 80% 이상이면 공통/참조 테이블
            join_ratio = join_count / total
            if join_ratio >= 0.8 and join_count >= 3:
                common_tables.add(table)

    # Fallback: connectivity threshold (table_usage가 없을 때)
    if not common_tables_manual and not table_usage:
        if common_threshold is None:
            common_threshold = max(5, int(len(tables_in_joins) * 0.3))
        for table, neighbors in adj.items():
            if len(neighbors) >= common_threshold:
                common_tables.add(table)

    if common_tables:
        logger.info("Common tables detected (%d): %s", len(common_tables), sorted(common_tables))

    # Build adjacency list WITHOUT common tables for clustering
    adj_filtered = defaultdict(set)
    for j in joins:
        t1, t2 = j["table1"], j["table2"]
        if t1 in common_tables or t2 in common_tables:
            continue
        adj_filtered[t1].add(t2)
        adj_filtered[t2].add(t1)

    tables_in_filtered = set(adj_filtered.keys())

    # Find connected components (BFS) without common tables
    visited = set()
    components = []

    for table in sorted(tables_in_filtered):
        if table in visited:
            continue
        component = _bfs(table, adj_filtered, visited)
        components.append(component)

    # Sort components by size (largest first)
    components.sort(key=len, reverse=True)

    # Split large components
    final_groups = []
    for comp in components:
        if len(comp) <= max_group_size:
            final_groups.append(comp)
        else:
            sub_groups = _split_large_group(comp, adj_filtered, max_group_size)
            final_groups.extend(sub_groups)

    # For each group, find which common tables are referenced and add them
    group_with_common = []
    for group_tables in final_groups:
        referenced_common = set()
        for j in joins:
            if j["table1"] in group_tables and j["table2"] in common_tables:
                referenced_common.add(j["table2"])
            elif j["table2"] in group_tables and j["table1"] in common_tables:
                referenced_common.add(j["table1"])
        group_with_common.append((group_tables, referenced_common))

    # Isolated tables: XML에서 사용됐지만 JOIN이 없는 테이블만 포함
    if query_tables:
        isolated = (query_tables & all_tables) - tables_in_joins - common_tables
    else:
        isolated = all_tables - tables_in_joins - common_tables
    if isolated:
        isolated_list = sorted(isolated)
        for i in range(0, len(isolated_list), max_group_size):
            chunk = set(isolated_list[i:i + max_group_size])
            group_with_common.append((chunk, set()))

    # Build group info with schema data
    schema_map = {t["name"]: t for t in schema.get("tables", [])}
    groups = []

    for idx, (group_tables, ref_common) in enumerate(group_with_common):
        # All tables in this group = business tables + referenced common tables
        all_group_tables = group_tables | ref_common

        # Filter joins for this group (including joins to common tables)
        group_joins = [
            j for j in joins
            if j["table1"] in all_group_tables and j["table2"] in all_group_tables
        ]

        # Filter schema for this group
        group_schema_tables = [
            schema_map[t] for t in sorted(all_group_tables) if t in schema_map
        ]

        # Generate group name from top BUSINESS tables (exclude common)
        connection_count = defaultdict(int)
        for j in group_joins:
            if j["table1"] not in common_tables:
                connection_count[j["table1"]] += 1
            if j["table2"] not in common_tables:
                connection_count[j["table2"]] += 1

        if connection_count:
            top_tables = sorted(connection_count, key=connection_count.get, reverse=True)[:3]
        else:
            top_tables = sorted(group_tables)[:3]

        is_isolated = len(group_joins) == 0

        # Skip groups where no tables exist in schema
        if not is_isolated and not group_schema_tables:
            continue

        groups.append({
            "index": 0,
            "tables": sorted(group_tables),
            "common_tables": sorted(ref_common),
            "all_tables": sorted(all_group_tables),
            "table_count": len(all_group_tables),
            "business_table_count": len(group_tables),
            "joins": group_joins,
            "join_count": len(group_joins),
            "top_tables": top_tables,
            "schema_tables": group_schema_tables,
            "is_isolated": is_isolated,
        })

    # Re-number groups (only valid ones)
    rel_idx = 0
    iso_idx = 0
    for g in groups:
        if g["is_isolated"]:
            iso_idx += 1
            g["index"] = iso_idx
        else:
            rel_idx += 1
            g["index"] = rel_idx

    # Classify tables
    if query_tables is None:
        query_tables = set()

    classification = {
        "common_tables": sorted(common_tables & all_tables),
        "tables_with_joins": sorted((tables_in_joins - common_tables) & all_tables),
        "tables_in_xml_no_join": sorted((query_tables - tables_in_joins - common_tables) & all_tables),
        "tables_not_in_xml": sorted(all_tables - query_tables - tables_in_joins - common_tables),
        "tables_in_xml_not_in_schema": sorted(query_tables - all_tables),
    }

    logger.info("Found %d groups (%d with relationships, %d isolated)",
                len(groups),
                sum(1 for g in groups if not g["is_isolated"]),
                sum(1 for g in groups if g["is_isolated"]))
    logger.info("Classification: %d with JOINs, %d in XML no JOIN, %d not in XML, %d in XML not in schema",
                len(classification["tables_with_joins"]),
                len(classification["tables_in_xml_no_join"]),
                len(classification["tables_not_in_xml"]),
                len(classification["tables_in_xml_not_in_schema"]))

    return groups, classification


def _bfs(start: str, adj: dict, visited: set) -> set:
    """BFS to find connected component."""
    queue = [start]
    component = set()
    while queue:
        node = queue.pop(0)
        if node in visited:
            continue
        visited.add(node)
        component.add(node)
        for neighbor in adj.get(node, set()):
            if neighbor not in visited:
                queue.append(neighbor)
    return component


def _split_large_group(tables: set, adj: dict, max_size: int) -> list[set]:
    """Split a large connected component into smaller sub-groups.

    Strategy: pick the most connected node, BFS up to max_size,
    then repeat for remaining nodes.
    """
    remaining = set(tables)
    sub_groups = []

    while remaining:
        # Find most connected node in remaining
        best = max(remaining, key=lambda t: len(adj.get(t, set()) & remaining))

        # BFS from best, limited to max_size
        sub = set()
        queue = [best]
        while queue and len(sub) < max_size:
            node = queue.pop(0)
            if node not in remaining or node in sub:
                continue
            sub.add(node)
            # Prioritize neighbors that are in remaining
            neighbors = sorted(adj.get(node, set()) & remaining - sub)
            queue.extend(neighbors)

        remaining -= sub
        sub_groups.append(sub)

    return sub_groups


def build_summary_markdown(groups: list[dict], classification: dict = None) -> str:
    """Build a summary markdown showing all groups and table classification."""
    lines = []
    lines.append("# ERD Groups Summary\n")

    rel_groups = [g for g in groups if not g["is_isolated"]]
    iso_groups = [g for g in groups if g["is_isolated"]]

    lines.append(f"- Total groups: {len(groups)}")
    lines.append(f"- Groups with relationships: {len(rel_groups)}")
    lines.append(f"- Isolated table groups: {len(iso_groups)}")
    total_tables = sum(g["table_count"] for g in groups)
    total_joins = sum(g["join_count"] for g in groups)
    lines.append(f"- Total tables: {total_tables}")
    lines.append(f"- Total relationships: {total_joins}")
    lines.append("")

    # Table classification
    if classification:
        lines.append("## Table Classification\n")
        c = classification
        lines.append(f"| Category | Count |")
        lines.append(f"|----------|-------|")
        lines.append(f"| 공통 테이블 (다수 도메인에서 참조) | {len(c.get('common_tables', []))} |")
        lines.append(f"| JOIN 관계가 있는 테이블 (ERD에 포함) | {len(c['tables_with_joins'])} |")
        lines.append(f"| XML에 존재하지만 JOIN 없음 | {len(c['tables_in_xml_no_join'])} |")
        lines.append(f"| XML에 존재하지 않는 테이블 | {len(c['tables_not_in_xml'])} |")
        lines.append(f"| XML에 있지만 스키마에 없는 테이블 | {len(c['tables_in_xml_not_in_schema'])} |")
        lines.append("")

    # Common tables
    if classification and classification.get("common_tables"):
        tables = classification["common_tables"]
        lines.append(f"## 공통 테이블 ({len(tables)}개)\n")
        lines.append("다수의 업무 도메인에서 참조하는 테이블입니다. 각 ERD 그룹에 참조용으로 포함됩니다.\n")
        for t in tables:
            lines.append(f"- {t}")
        lines.append("")

    # Groups with relationships
    if rel_groups:
        lines.append("## ERD Groups (with Relationships)\n")
        lines.append("| Group | Business Tables | Common Tables | Relationships | Key Tables |")
        lines.append("|-------|-----------------|---------------|---------------|------------|")
        for g in rel_groups:
            top = ", ".join(g["top_tables"])
            common_count = len(g.get("common_tables", []))
            lines.append(f"| {g['index']:02d} | {g['business_table_count']} | {common_count} | {g['join_count']} | {top} |")
        lines.append("")

    # Tables in XML but no JOIN
    if classification and classification["tables_in_xml_no_join"]:
        tables = classification["tables_in_xml_no_join"]
        lines.append(f"## XML에 존재하지만 JOIN 없는 테이블 ({len(tables)}개)\n")
        lines.append("쿼리에서 단독 SELECT/INSERT/UPDATE/DELETE로 사용되지만 다른 테이블과 JOIN이 없는 테이블입니다.\n")
        for t in tables:
            lines.append(f"- {t}")
        lines.append("")

    # Tables not in XML
    if classification and classification["tables_not_in_xml"]:
        tables = classification["tables_not_in_xml"]
        lines.append(f"## XML에 존재하지 않는 테이블 ({len(tables)}개)\n")
        lines.append("스키마에는 있지만 MyBatis XML 쿼리에서 사용되지 않는 테이블입니다.\n")
        for t in tables:
            lines.append(f"- {t}")
        lines.append("")

    # Tables in XML but not in schema
    if classification and classification["tables_in_xml_not_in_schema"]:
        tables = classification["tables_in_xml_not_in_schema"]
        lines.append(f"## XML에 있지만 스키마에 없는 테이블 ({len(tables)}개)\n")
        lines.append("쿼리에서 참조하지만 스키마 .md에 정의가 없는 테이블입니다. (다른 스키마 소유이거나 뷰일 수 있음)\n")
        for t in tables:
            lines.append(f"- {t}")
        lines.append("")

    return "\n".join(lines)


def build_summary_excel(groups: list[dict], classification: dict,
                        schema: dict, output_path: str) -> str:
    """Build an Excel summary with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Build table comment lookup
    table_comments = {}
    for t in schema.get("tables", []):
        table_comments[t["name"]] = t.get("comment") or ""

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def _write_row(ws, row_num, values):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_header(ws_summary, ["Category", "Count"])

    rel_groups = [g for g in groups if not g["is_isolated"]]
    summary_data = [
        ("ERD 그룹 수", len(rel_groups)),
        ("JOIN 관계 테이블", len(classification.get("tables_with_joins", []))),
        ("XML에 있지만 JOIN 없음 (단독)", len(classification.get("tables_in_xml_no_join", []))),
        ("XML에 없는 테이블 (미사용)", len(classification.get("tables_not_in_xml", []))),
        ("XML에만 있고 스키마에 없음", len(classification.get("tables_in_xml_not_in_schema", []))),
    ]
    for i, (cat, cnt) in enumerate(summary_data, 2):
        _write_row(ws_summary, i, [cat, cnt])

    # Group detail table
    ws_summary.cell(row=len(summary_data) + 3, column=1, value="ERD Groups").font = Font(bold=True, size=12)
    group_start = len(summary_data) + 4
    _write_header_at = lambda ws, row, headers: [
        setattr(ws.cell(row=row, column=c, value=h), 'font', header_font) or
        setattr(ws.cell(row=row, column=c, value=h), 'fill', header_fill) or
        setattr(ws.cell(row=row, column=c, value=h), 'border', thin_border)
        for c, h in enumerate(headers, 1)
    ]
    headers = ["Group", "Tables", "Relationships", "Key Tables"]
    for c, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=group_start, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for i, g in enumerate(rel_groups, group_start + 1):
        _write_row(ws_summary, i, [
            f"Group {g['index']:02d}",
            g["table_count"],
            g["join_count"],
            ", ".join(g["top_tables"][:3]),
        ])

    _auto_width(ws_summary)

    # === Sheet 2: JOIN 관계 테이블 ===
    ws_join = wb.create_sheet("JOIN관계테이블")
    _write_header(ws_join, ["No", "테이블 물리명", "테이블명(Comment)", "ERD Group", "관계 수"])

    join_tables = classification.get("tables_with_joins", [])
    for i, t in enumerate(join_tables, 2):
        # Find which group this table belongs to
        group_num = ""
        rel_count = 0
        for g in rel_groups:
            if t in g["tables"]:
                group_num = f"Group {g['index']:02d}"
                rel_count = sum(1 for j in g["joins"] if j["table1"] == t or j["table2"] == t)
                break
        _write_row(ws_join, i, [i - 1, t, table_comments.get(t, ""), group_num, rel_count])

    _auto_width(ws_join)

    # === Sheet 3: 단독 테이블 ===
    ws_solo = wb.create_sheet("단독테이블")
    _write_header(ws_solo, ["No", "테이블 물리명", "테이블명(Comment)"])

    solo_tables = classification.get("tables_in_xml_no_join", [])
    for i, t in enumerate(solo_tables, 2):
        _write_row(ws_solo, i, [i - 1, t, table_comments.get(t, "")])

    _auto_width(ws_solo)

    # === Sheet 4: 공통 테이블 ===
    ws_common = wb.create_sheet("공통테이블")
    _write_header(ws_common, ["No", "테이블 물리명", "테이블명(Comment)", "참조 그룹 수"])

    common_tables_list = classification.get("common_tables", [])
    for i, t in enumerate(common_tables_list, 2):
        ref_count = sum(1 for g in rel_groups if t in g.get("common_tables", []))
        _write_row(ws_common, i, [i - 1, t, table_comments.get(t, ""), ref_count])

    _auto_width(ws_common)

    # === Sheet 5: XML에 없는 테이블 (미사용) ===
    ws_unused = wb.create_sheet("미사용테이블")
    _write_header(ws_unused, ["No", "테이블 물리명", "테이블명(Comment)"])

    unused_tables = classification.get("tables_not_in_xml", [])
    for i, t in enumerate(unused_tables, 2):
        _write_row(ws_unused, i, [i - 1, t, table_comments.get(t, "")])

    _auto_width(ws_unused)

    # Save
    wb.save(output_path)
    logger.info("Summary Excel saved: %s", output_path)
    return output_path
