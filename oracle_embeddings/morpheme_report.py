"""Morpheme analysis report — .md summary + single-sheet .xlsx.

Excel 시트 컬럼:
  속성명 | 컨피던스 | 단어1 | 단어2 | ... | 단어12 | 비고
"""

from __future__ import annotations

import logging
import os
from datetime import datetime

from oracle_embeddings.morpheme_analyzer import (
    AnalysisStats,
    LOW_CONFIDENCE_THRESHOLD,
    MAX_TOKENS_PER_ATTR,
    MorphemeResult,
)

logger = logging.getLogger(__name__)


def _md_escape(text: str) -> str:
    if not text:
        return ""
    return str(text).replace("|", "\\|").replace("\n", "<br>").replace("\r", "")


def _format_elapsed(sec: float) -> str:
    if sec < 60:
        return f"{sec:.1f}s"
    m, s = divmod(int(sec), 60)
    if m < 60:
        return f"{m}m {s}s"
    h, m = divmod(m, 60)
    return f"{h}h {m}m {s}s"


def save_morpheme_markdown(
    results: list[MorphemeResult],
    stats: AnalysisStats,
    output_dir: str,
    guide_path: str,
) -> str:
    """md 요약 리포트 작성."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"morpheme_{timestamp}.md")

    low_conf_rows = [
        r for r in results if 0.0 < r.confidence < LOW_CONFIDENCE_THRESHOLD
    ]
    failed_rows = [r for r in results if r.note.startswith("파싱 실패")]
    truncated_rows = [r for r in results if "생략" in r.note]

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# 형태소분석 리포트 (Morpheme Analysis)\n\n")
        f.write(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"- Guide file: `{guide_path}`\n\n")

        f.write("## Summary\n\n")
        f.write("| 항목 | 값 |\n|------|----|\n")
        f.write(f"| 총 속성 수 | {stats.total:,} |\n")
        f.write(f"| 분해 성공 | {stats.success:,} |\n")
        f.write(f"| 저신뢰도 (< {LOW_CONFIDENCE_THRESHOLD}) | {stats.low_confidence:,} |\n")
        f.write(f"| 파싱 실패 | {stats.parse_failed:,} |\n")
        f.write(f"| 단어 잘림 (> {MAX_TOKENS_PER_ATTR}개) | {stats.truncated:,} |\n")
        f.write(f"| 적용 배치 크기 | {stats.batch_size_effective} |\n")
        f.write(f"| 총 배치 수 | {stats.batches_total:,} |\n")
        f.write(f"| 재시도 횟수 | {stats.retries:,} |\n")
        f.write(f"| 병렬도 | {stats.parallel} |\n")
        f.write(f"| 소요시간 | {_format_elapsed(stats.elapsed_sec)} |\n\n")

        token_sep = " \\| "  # Markdown 테이블 셀 안의 `|` 이스케이프
        if low_conf_rows:
            f.write(f"## 저신뢰도 상위 20 샘플 ({len(low_conf_rows)}개 중)\n\n")
            f.write("| 속성명 | 컨피던스 | 분해결과 | 비고 |\n")
            f.write("|--------|----------|----------|------|\n")
            for r in low_conf_rows[:20]:
                joined = _md_escape(token_sep.join(r.tokens))
                f.write(
                    f"| {_md_escape(r.attr)} | {r.confidence:.2f} "
                    f"| {joined} | {_md_escape(r.note)} |\n"
                )
            f.write("\n")

        if failed_rows:
            f.write(f"## 파싱 실패 상위 20 샘플 ({len(failed_rows)}개 중)\n\n")
            f.write("| 속성명 | 비고 |\n|--------|------|\n")
            for r in failed_rows[:20]:
                f.write(f"| {_md_escape(r.attr)} | {_md_escape(r.note)} |\n")
            f.write("\n")

        if truncated_rows:
            f.write(f"## 단어 잘림 상위 20 샘플 ({len(truncated_rows)}개 중)\n\n")
            f.write("| 속성명 | 유지된 12개 | 비고 |\n")
            f.write("|--------|------------|------|\n")
            for r in truncated_rows[:20]:
                joined = _md_escape(token_sep.join(r.tokens))
                f.write(
                    f"| {_md_escape(r.attr)} | {joined} "
                    f"| {_md_escape(r.note)} |\n"
                )
            f.write("\n")

    logger.info("Morpheme markdown saved: %s", filepath)
    return filepath


def save_morpheme_excel(
    results: list[MorphemeResult],
    output_dir: str,
) -> str:
    """단일 시트 xlsx 리포트 작성."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"morpheme_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "형태소분석"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    low_conf_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fail_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    truncated_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    headers = (
        ["속성명", "컨피던스"]
        + [f"단어{i+1}" for i in range(MAX_TOKENS_PER_ATTR)]
        + ["비고"]
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, r in enumerate(results, 2):
        row_fill = None
        if r.note.startswith("파싱 실패"):
            row_fill = fail_fill
        elif 0.0 < r.confidence < LOW_CONFIDENCE_THRESHOLD:
            row_fill = low_conf_fill
        elif "생략" in r.note:
            row_fill = truncated_fill

        values: list[object] = [r.attr, round(r.confidence, 2)]
        for i in range(MAX_TOKENS_PER_ATTR):
            values.append(r.tokens[i] if i < len(r.tokens) else "")
        values.append(r.note)

        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.border = thin_border
            if row_fill is not None:
                cell.fill = row_fill

    # column widths
    ws.column_dimensions["A"].width = 32  # 속성명
    ws.column_dimensions["B"].width = 10  # 컨피던스
    for i in range(MAX_TOKENS_PER_ATTR):
        col_letter = ws.cell(row=1, column=3 + i).column_letter
        ws.column_dimensions[col_letter].width = 14
    note_col = ws.cell(row=1, column=3 + MAX_TOKENS_PER_ATTR).column_letter
    ws.column_dimensions[note_col].width = 40

    ws.freeze_panes = "C2"

    wb.save(filepath)
    logger.info("Morpheme excel saved: %s", filepath)
    return filepath
