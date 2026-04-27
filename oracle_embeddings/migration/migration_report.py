"""Excel report for ``migrate-sql`` (docs/migration/spec.md §12.1).

Takes a flat ``List[RewriteResult]`` (across every XML file processed) plus
the loaded :class:`Mapping` and emits a 5-sheet workbook:

1. Summary — totals, status breakdown, Stage A/B pass rate, automation %
2. Conversions — main sheet, 18 columns per spec
3. Validation Errors — PARSE_FAIL + Stage A/B failures
4. Unresolved Queue — rows awaiting human review (NEEDS_LLM / UNRESOLVED)
5. Mapping Coverage — per-mapping hit counts, surfaces unused entries

The writer is intentionally side-effect only — no aggregation state is
returned, no file I/O happens outside ``output_path``. Callers (``migrate-sql``
in Step 14) prepare the result list and invoke ``write_migration_report``.
"""
from __future__ import annotations

import logging
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .mapping_model import ChangeItem, Mapping, RewriteResult

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------


_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)

# 노랑 — AUTO_WARN, NEEDS_LLM
_WARN_FILL = PatternFill("solid", fgColor="FFF5CC")
# 빨강 — UNRESOLVED, PARSE_FAIL, Stage B 실패
_ERROR_FILL = PatternFill("solid", fgColor="FFE4E1")
# 회색 — 변환 대상 없음 (AUTO w/o changes)
_GRAY_FILL = PatternFill("solid", fgColor="E8E8E8")

_WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def write_migration_report(
    results: List[RewriteResult],
    mapping: Mapping,
    output_path: Path,
    *,
    mybatis_dir: Optional[Path] = None,
    mapping_path: Optional[Path] = None,
) -> None:
    """Render the 5-sheet migration workbook and save to ``output_path``."""

    wb = Workbook()
    _write_summary(wb, results, mapping, mybatis_dir, mapping_path)
    _write_conversions(wb, results)
    _write_validation_errors(wb, results)
    _write_unresolved(wb, results)
    _write_coverage(wb, results, mapping)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Wrote migration report: %s", output_path)


# ---------------------------------------------------------------------------
# Sheet: Summary
# ---------------------------------------------------------------------------


def _write_summary(
    wb: Workbook,
    results: List[RewriteResult],
    mapping: Mapping,
    mybatis_dir: Optional[Path],
    mapping_path: Optional[Path],
) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws.append(["SQL Migration Report"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    ws.append(["Mapping", str(mapping_path) if mapping_path else "-"])
    ws.append(["MyBatis dir", str(mybatis_dir) if mybatis_dir else "-"])
    ws.append(["Mapping version", mapping.version])
    ws.append([])

    total = len(results)
    status_count = Counter(r.status for r in results)
    auto = status_count.get("AUTO", 0)
    auto_warn = status_count.get("AUTO_WARN", 0)
    needs_llm = status_count.get("NEEDS_LLM", 0)
    unresolved = status_count.get("UNRESOLVED", 0)
    parse_fail = status_count.get("PARSE_FAIL", 0)

    ws.append(["Total statements", total])
    ws.append(["  AUTO", auto])
    ws.append(["  AUTO_WARN", auto_warn])
    ws.append(["  NEEDS_LLM", needs_llm])
    ws.append(["  UNRESOLVED", unresolved])
    ws.append(["  PARSE_FAIL", parse_fail])
    ws.append([])

    if total:
        automation_pct = (auto + auto_warn) / total * 100
        ws.append(["Automation rate", f"{automation_pct:.1f}% (AUTO + AUTO_WARN)"])
    else:
        ws.append(["Automation rate", "-"])

    stage_a_ran = [r for r in results if r.stage_a_pass is not None]
    if stage_a_ran:
        pa = sum(1 for r in stage_a_ran if r.stage_a_pass) / len(stage_a_ran) * 100
        ws.append(["Stage A pass rate", f"{pa:.1f}% ({sum(1 for r in stage_a_ran if r.stage_a_pass)}/{len(stage_a_ran)})"])
    else:
        ws.append(["Stage A pass rate", "not run"])

    stage_b_ran = [r for r in results if r.stage_b_pass is not None]
    if stage_b_ran:
        pb = sum(1 for r in stage_b_ran if r.stage_b_pass) / len(stage_b_ran) * 100
        ws.append(["Stage B pass rate", f"{pb:.1f}% ({sum(1 for r in stage_b_ran if r.stage_b_pass)}/{len(stage_b_ran)})"])
    else:
        ws.append(["Stage B pass rate", "not run"])

    ws.append([])
    ws.append(["Statements by type"])
    type_count = Counter(r.sql_type for r in results)
    for t, c in sorted(type_count.items()):
        ws.append([f"  {t}", c])

    ws.append([])
    ws.append(["Method breakdown"])
    method_count = Counter(r.conversion_method for r in results)
    for m, c in sorted(method_count.items()):
        ws.append([f"  {m}", c])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


# ---------------------------------------------------------------------------
# Sheet: Conversions (main, 18 columns)
# ---------------------------------------------------------------------------


_CONVERSION_COLUMNS = [
    ("No", 6),
    ("XML File", 40),
    ("Namespace", 28),
    ("SQL ID", 28),
    ("SQL Type", 10),
    ("AS-IS SQL", 60),
    ("TO-BE SQL", 60),
    ("Status", 14),
    ("Applied Transformers", 28),
    ("Conversion Method", 16),
    ("Changed Items", 40),
    ("Dynamic Paths", 12),
    ("Stage A Pass", 12),
    ("Stage B Pass", 12),
    ("ORA Error", 40),
    ("LLM Confidence", 14),
    ("Notes", 30),
    ("Last Modified", 18),
]


def _write_conversions(wb: Workbook, results: List[RewriteResult]) -> None:
    ws = wb.create_sheet("Conversions")
    headers = [h for h, _w in _CONVERSION_COLUMNS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for i, (_h, width) in enumerate(_CONVERSION_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for idx, r in enumerate(results, start=1):
        row = [
            idx,
            str(r.xml_file),
            r.namespace,
            r.sql_id,
            r.sql_type,
            r.as_is_sql,
            r.to_be_sql or "",
            r.status,
            ", ".join(r.applied_transformers),
            r.conversion_method,
            _format_changes(r.changed_items),
            r.dynamic_paths_expanded,
            _tri(r.stage_a_pass),
            _tri(r.stage_b_pass),
            r.parse_error or "",
            f"{r.llm_confidence:.2f}" if r.llm_confidence is not None else "",
            "; ".join(r.notes + r.warnings),
            r.last_modified.strftime("%Y-%m-%d %H:%M") if r.last_modified else "",
        ]
        ws.append(row)
        fill = _status_fill(r)
        if fill is not None:
            for cell in ws[ws.max_row]:
                cell.fill = fill
        # Wrap long SQL
        for col_idx in (6, 7, 11, 17):
            ws.cell(row=ws.max_row, column=col_idx).alignment = _WRAP


# ---------------------------------------------------------------------------
# Sheet: Validation Errors
# ---------------------------------------------------------------------------


def _write_validation_errors(wb: Workbook, results: List[RewriteResult]) -> None:
    ws = wb.create_sheet("Validation Errors")
    headers = [
        "No", "XML File", "Namespace", "SQL ID", "Status",
        "Stage A", "Stage B", "ORA Error", "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    filtered = [
        r for r in results
        if r.status == "PARSE_FAIL"
        or r.stage_a_pass is False
        or r.stage_b_pass is False
    ]
    for idx, r in enumerate(filtered, start=1):
        ws.append([
            idx,
            str(r.xml_file),
            r.namespace,
            r.sql_id,
            r.status,
            _tri(r.stage_a_pass),
            _tri(r.stage_b_pass),
            r.parse_error or "",
            "; ".join(r.warnings + r.notes),
        ])
        for cell in ws[ws.max_row]:
            cell.fill = _ERROR_FILL

    widths = [6, 40, 28, 28, 14, 10, 10, 50, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet: Unresolved Queue
# ---------------------------------------------------------------------------


def _write_unresolved(wb: Workbook, results: List[RewriteResult]) -> None:
    ws = wb.create_sheet("Unresolved Queue")
    headers = [
        "No", "XML File", "Namespace", "SQL ID", "Status",
        "AS-IS SQL", "TO-BE SQL (suggested)",
        "Reason / Warnings", "LLM Confidence",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    filtered = [r for r in results if r.status in ("NEEDS_LLM", "UNRESOLVED")]
    for idx, r in enumerate(filtered, start=1):
        ws.append([
            idx,
            str(r.xml_file),
            r.namespace,
            r.sql_id,
            r.status,
            r.as_is_sql,
            r.to_be_sql or "",
            "; ".join(r.warnings + r.notes) or "-",
            f"{r.llm_confidence:.2f}" if r.llm_confidence is not None else "",
        ])
        for cell in ws[ws.max_row]:
            cell.fill = _WARN_FILL
        for col_idx in (6, 7, 8):
            ws.cell(row=ws.max_row, column=col_idx).alignment = _WRAP

    widths = [6, 40, 28, 28, 14, 60, 60, 40, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Sheet: Mapping Coverage
# ---------------------------------------------------------------------------


def _write_coverage(
    wb: Workbook,
    results: List[RewriteResult],
    mapping: Mapping,
) -> None:
    ws = wb.create_sheet("Mapping Coverage")
    ws.append([
        "Kind", "AS-IS", "TO-BE", "Applied count",
        "XML file count", "Status",
    ])
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    # Tally hits per (kind, as_is) — pre-grouped over to_be so coverage
    # lookups below are O(1). Earlier impl carried the to_be in the key and
    # the lookup scanned every entry → O(n×m) for a mapping with n hits and
    # m mapping rows; pre-grouping here is O(n+m) total.
    grouped_count: Counter = Counter()
    grouped_files: Dict[Tuple[str, str], set] = {}
    for r in results:
        for c in r.changed_items:
            key = (c.kind, c.as_is)
            grouped_count[key] += c.count
            grouped_files.setdefault(key, set()).add(str(r.xml_file))

    # Emit tables first, then columns — grouped by mapping entry
    for tm in mapping.tables:
        to_be_str = ",".join(tm.to_be_tables()) or "<dropped>"
        for as_is_name in tm.as_is_tables():
            count, files = _coverage_lookup(
                grouped_count, grouped_files, "table", as_is_name.upper()
            )
            _append_coverage_row(ws, tm.type, as_is_name.upper(), to_be_str, count, len(files))

    for cm in mapping.columns:
        # Build to_be display
        if cm.to_be is None:
            to_be_disp = "<dropped>"
        elif isinstance(cm.to_be, list):
            to_be_disp = ",".join(f"{x.table}.{x.column}" for x in cm.to_be)
        else:
            to_be_disp = f"{cm.to_be.table}.{cm.to_be.column}"
        for ref in cm.as_is_refs():
            as_is_disp = f"{ref.table}.{ref.column}"
            count, files = _coverage_lookup(
                grouped_count, grouped_files, "column", as_is_disp.upper()
            )
            _append_coverage_row(ws, cm.kind, as_is_disp.upper(), to_be_disp, count, len(files))

    widths = [14, 30, 40, 14, 16, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _coverage_lookup(
    grouped_count: Counter,
    grouped_files: Dict[Tuple[str, str], set],
    kind: str,
    as_is_upper: str,
) -> Tuple[int, set]:
    """O(1) lookup against pre-grouped hit tallies — see caller for the
    grouping pass."""
    key = (kind, as_is_upper)
    return grouped_count.get(key, 0), grouped_files.get(key, set())


def _append_coverage_row(
    ws, kind: str, as_is: str, to_be: str, count: int, file_count: int
) -> None:
    ws.append([kind, as_is, to_be, count, file_count, "UNUSED" if count == 0 else ""])
    if count == 0:
        for cell in ws[ws.max_row]:
            cell.fill = _GRAY_FILL


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _status_fill(r: RewriteResult) -> Optional[PatternFill]:
    if r.status in ("UNRESOLVED", "PARSE_FAIL"):
        return _ERROR_FILL
    # Validation failures override AUTO/AUTO_WARN — without this, a row that
    # passed DSL rewrite (status=AUTO) but failed Stage A schema lookup would
    # show no fill, even though Validation Errors sheet flags it red.
    if r.stage_a_pass is False or r.stage_b_pass is False:
        return _ERROR_FILL
    if r.status in ("AUTO_WARN", "NEEDS_LLM"):
        return _WARN_FILL
    if not r.changed_items and r.status == "AUTO":
        return _GRAY_FILL
    return None


def _tri(v: Optional[bool]) -> str:
    if v is True:
        return "Y"
    if v is False:
        return "N"
    return "-"


def _format_changes(items: List[ChangeItem]) -> str:
    if not items:
        return ""
    parts = [f"{c.as_is}→{c.to_be}({c.count})" for c in items]
    return ", ".join(parts)
