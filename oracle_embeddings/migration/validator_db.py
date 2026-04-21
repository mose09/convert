"""Stage B — actual Oracle parse via ``cursor.parse()``.

``cursor.parse`` performs the same syntactic + semantic analysis as
``DBMS_SQL.PARSE`` without executing the statement. No transactions, no
locks, no row fetches — exactly the pre-flight check we want.

This module keeps ``oracledb`` as an optional runtime dep (``import`` inside
functions) so the rest of the migration toolkit keeps working in environments
where Oracle client libs aren't available (CI, code review, offline dev).
"""
from __future__ import annotations

import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from .bind_dummifier import dummify
from .validator_static import ValidationIssue, ValidationResult

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Single-statement validator
# ---------------------------------------------------------------------------


def validate_db(sql: str, conn: Any) -> ValidationResult:
    """Hand ``sql`` (already containing OGNL placeholders or not) to an Oracle
    cursor's ``.parse()`` method via ``conn``. Returns a
    :class:`ValidationResult` with a single issue on failure.

    ``conn`` must be an ``oracledb.Connection`` (thin or thick mode). The
    statement is first dummified so Oracle sees native ``:name`` binds.
    """
    dummified = dummify(sql)
    try:
        cur = conn.cursor()
    except Exception as exc:  # pragma: no cover - defensive
        return ValidationResult(
            ok=False,
            parse_error=str(exc),
            issues=[ValidationIssue(
                level="error", code="CURSOR_FAIL",
                message=f"failed to open cursor: {exc}",
            )],
        )

    try:
        cur.parse(dummified)
        return ValidationResult(ok=True)
    except Exception as exc:
        code, msg = _extract_oracle_error(exc)
        return ValidationResult(
            ok=False,
            parse_error=msg,
            issues=[ValidationIssue(
                level="error",
                code=code or "DB_PARSE_FAIL",
                message=msg,
            )],
        )
    finally:
        try:
            cur.close()
        except Exception:  # pragma: no cover - defensive
            pass


# ---------------------------------------------------------------------------
# Batch validation
# ---------------------------------------------------------------------------


@dataclass
class BatchItem:
    """One statement queued for Stage B. Key is an arbitrary caller-provided
    tuple (e.g. ``(xml_file, sql_id)``) used to correlate back to the report."""

    key: Tuple[str, ...]
    sql: str


@dataclass
class BatchResult:
    key: Tuple[str, ...]
    result: ValidationResult


def validate_db_batch(
    items: List[BatchItem],
    *,
    dsn: str,
    user: str,
    password: str,
    parallel: int = 10,
    thick_mode: bool = True,
    oracle_client_dir: Optional[str] = None,
) -> List[BatchResult]:
    """Validate many statements concurrently.

    Opens one connection per worker (they don't share cursors), runs
    ``cursor.parse`` on each statement, collects results. Errors during
    connection setup propagate to the caller (if the DSN is wrong, we want
    to fail fast rather than silently marking every row as failed).
    """
    import oracledb  # optional dep, imported here

    if thick_mode:
        try:
            oracledb.init_oracle_client(lib_dir=oracle_client_dir)
        except Exception as exc:
            logger.warning("thick-mode init failed, falling back to thin: %s", exc)

    def _worker(batch: List[BatchItem]) -> List[BatchResult]:
        conn = oracledb.connect(user=user, password=password, dsn=dsn)
        try:
            return [
                BatchResult(key=it.key, result=validate_db(it.sql, conn))
                for it in batch
            ]
        finally:
            try:
                conn.close()
            except Exception:  # pragma: no cover - defensive
                pass

    if parallel <= 1 or len(items) <= 1:
        return _worker(items)

    # Chunk to ``parallel`` workers, each owns one connection for its slice.
    chunk_size = max(1, (len(items) + parallel - 1) // parallel)
    chunks = [items[i:i + chunk_size] for i in range(0, len(items), chunk_size)]

    out: List[BatchResult] = []
    with ThreadPoolExecutor(max_workers=min(parallel, len(chunks))) as ex:
        futures = [ex.submit(_worker, chunk) for chunk in chunks]
        for f in as_completed(futures):
            out.extend(f.result())
    # Keep caller-supplied order for stable report output.
    order = {it.key: i for i, it in enumerate(items)}
    out.sort(key=lambda br: order.get(br.key, 0))
    return out


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def write_validation_report(
    results: List[BatchResult],
    stmt_meta: Dict[Tuple[str, ...], Dict[str, Any]],
    output_path: Path,
) -> None:
    """Write a 2-sheet xlsx (Summary + per-statement grid) for Stage B results.

    ``stmt_meta`` carries the original SQL / file path for each key so the
    Excel row can show context without a second pass over the XMLs.
    """
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Stage B Validation Report"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append([])

    total = len(results)
    passed = sum(1 for r in results if r.result.ok)
    failed = total - passed
    summary.append(["Total statements", total])
    summary.append(["Pass", passed])
    summary.append(["Fail", failed])
    if total:
        summary.append(["Pass rate", f"{passed / total * 100:.1f}%"])
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 40

    ws = wb.create_sheet("Statements")
    headers = ["No", "XML File", "Namespace", "SQL ID", "Pass", "ORA Error", "SQL"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    widths = [6, 40, 28, 28, 8, 50, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    err_fill = PatternFill("solid", fgColor="FFE4E1")
    wrap = Alignment(wrap_text=True, vertical="top")

    for idx, br in enumerate(results, start=1):
        meta = stmt_meta.get(br.key, {}) or {}
        ok = br.result.ok
        err_msg = ""
        if not ok:
            errs = br.result.errors
            if errs:
                err_msg = f"[{errs[0].code}] {errs[0].message}"
            elif br.result.parse_error:
                err_msg = br.result.parse_error
        ws.append([
            idx,
            br.key[0] if len(br.key) > 0 else "",
            br.key[1] if len(br.key) > 1 else "",
            br.key[2] if len(br.key) > 2 else "",
            "Y" if ok else "N",
            err_msg,
            (meta.get("sql") or "")[:2000],
        ])
        if not ok:
            for cell in ws[ws.max_row]:
                cell.fill = err_fill
        for col_idx in (6, 7):
            ws.cell(row=ws.max_row, column=col_idx).alignment = wrap

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Wrote validation report: %s", output_path)


def _extract_oracle_error(exc: Exception) -> Tuple[Optional[str], str]:
    """Best-effort ORA-XXXXX code extraction. oracledb 2.x exposes
    ``.full_code`` and ``.code`` via ``e.args[0]`` (an ``_Error`` object)."""
    try:
        arg0 = exc.args[0] if exc.args else None
        full = getattr(arg0, "full_code", None)
        msg = getattr(arg0, "message", None) or str(exc)
        if full:
            return full, msg
        code = getattr(arg0, "code", None)
        if isinstance(code, int):
            return f"ORA-{code:05d}", msg
    except Exception:
        pass
    return None, str(exc)
