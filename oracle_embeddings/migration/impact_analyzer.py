"""Impact analysis for ``column_mapping.yaml`` against MyBatis mappers.

No rewriting — just counts. Used by the ``migration-impact`` command to give
the user a pre-flight check: how many statements / XML files will each
mapping touch, which mappings are unused, where the mapping file disagrees
with the schema. See docs/migration/spec.md §3.1.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..md_parser import parse_schema_md
from ..mybatis_parser import extract_table_usage, parse_all_mappers
from .mapping_loader import load_mapping_collect
from .mapping_model import ColumnMapping, LoaderError, Mapping

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclass
class StatementImpact:
    xml_file: str
    namespace: str
    sql_id: str
    sql_type: str
    tables_hit: List[str]
    columns_hit: List[str]
    kinds_hit: Set[str] = field(default_factory=set)


@dataclass
class TableImpact:
    table: str           # AS-IS upper
    mapping_kind: str    # rename / split / merge / drop
    to_be: str
    statement_count: int
    xml_file_count: int
    xml_files: List[str]


@dataclass
class ColumnImpact:
    table: str           # AS-IS upper
    column: str          # AS-IS upper
    mapping_kind: str    # rename / type_convert / value_map / split / merge / drop
    to_be: str
    statement_count: int
    xml_file_count: int
    xml_files: List[str]


@dataclass
class ImpactReport:
    mybatis_dir: Path
    mapping_path: Path
    as_is_schema_path: Optional[Path]
    to_be_schema_path: Optional[Path]

    # Statement-level
    statement_count: int
    affected_statement_count: int
    statements_per_type: Dict[str, int]

    # Per-mapping aggregates
    table_impacts: List[TableImpact]
    column_impacts: List[ColumnImpact]
    statement_impacts: List[StatementImpact]

    # Validation
    loader_errors: List[LoaderError]
    schema_mismatches: List[str]

    # Schema stats
    as_is_table_count: Optional[int] = None
    to_be_table_count: Optional[int] = None

    @property
    def unused_tables(self) -> List[TableImpact]:
        return [t for t in self.table_impacts if t.statement_count == 0]

    @property
    def unused_columns(self) -> List[ColumnImpact]:
        return [c for c in self.column_impacts if c.statement_count == 0]


# ---------------------------------------------------------------------------
# Schema loading helper (shared with future migrate-sql / validate-migration)
# ---------------------------------------------------------------------------


def load_schema_tables(md_path: Path) -> Dict[str, Set[str]]:
    """Return ``{TABLE_UPPER: {COLUMN_UPPER, ...}}`` from a schema .md file
    (as produced by the ``schema`` command)."""
    schema = parse_schema_md(str(md_path))
    out: Dict[str, Set[str]] = {}
    for tbl in schema["tables"]:
        name = tbl["name"].upper()
        cols = {c["column_name"].upper() for c in tbl["columns"]}
        out[name] = cols
    return out


# ---------------------------------------------------------------------------
# Core analyzer
# ---------------------------------------------------------------------------


def analyze_impact(
    mybatis_dir: Path,
    mapping_path: Path,
    *,
    as_is_schema_path: Optional[Path] = None,
    to_be_schema_path: Optional[Path] = None,
) -> ImpactReport:
    """Load mapping + scan mappers + build ``ImpactReport``.

    Soft-fails: if the mapping file itself is unparseable we still return a
    report (with ``loader_errors`` populated and zero statement counts) so the
    CLI can surface every problem at once.
    """

    as_is_schema = (
        load_schema_tables(as_is_schema_path) if as_is_schema_path else None
    )
    to_be_schema = (
        load_schema_tables(to_be_schema_path) if to_be_schema_path else None
    )

    mapping, loader_errors = load_mapping_collect(
        mapping_path,
        as_is_schema=as_is_schema,
        to_be_schema=to_be_schema,
    )

    if mapping is None:
        return _empty_report(
            mybatis_dir, mapping_path, as_is_schema_path, to_be_schema_path,
            loader_errors=loader_errors,
            as_is_schema=as_is_schema, to_be_schema=to_be_schema,
        )

    schema_mismatches = _collect_schema_mismatches(
        mapping, as_is_schema, to_be_schema
    )

    logger.info("Scanning MyBatis dir: %s", mybatis_dir)
    parsed = parse_all_mappers(str(mybatis_dir))
    statements: List[dict] = parsed["statements"]

    stmt_impacts, table_stmt, table_files, column_stmt, column_files, per_type = (
        _scan_statements(statements, mapping, mybatis_dir)
    )

    table_impacts = _build_table_impacts(mapping, table_stmt, table_files)
    column_impacts = _build_column_impacts(mapping, column_stmt, column_files)

    affected = sum(
        1 for s in stmt_impacts if s.tables_hit or s.columns_hit
    )

    return ImpactReport(
        mybatis_dir=mybatis_dir,
        mapping_path=mapping_path,
        as_is_schema_path=as_is_schema_path,
        to_be_schema_path=to_be_schema_path,
        statement_count=len(statements),
        affected_statement_count=affected,
        statements_per_type=per_type,
        table_impacts=table_impacts,
        column_impacts=column_impacts,
        statement_impacts=stmt_impacts,
        loader_errors=loader_errors,
        schema_mismatches=schema_mismatches,
        as_is_table_count=len(as_is_schema) if as_is_schema else None,
        to_be_table_count=len(to_be_schema) if to_be_schema else None,
    )


def _empty_report(
    mybatis_dir: Path,
    mapping_path: Path,
    as_is_schema_path: Optional[Path],
    to_be_schema_path: Optional[Path],
    *,
    loader_errors: List[LoaderError],
    as_is_schema: Optional[Dict[str, Set[str]]],
    to_be_schema: Optional[Dict[str, Set[str]]],
) -> ImpactReport:
    return ImpactReport(
        mybatis_dir=mybatis_dir,
        mapping_path=mapping_path,
        as_is_schema_path=as_is_schema_path,
        to_be_schema_path=to_be_schema_path,
        statement_count=0,
        affected_statement_count=0,
        statements_per_type={},
        table_impacts=[],
        column_impacts=[],
        statement_impacts=[],
        loader_errors=loader_errors,
        schema_mismatches=[],
        as_is_table_count=len(as_is_schema) if as_is_schema else None,
        to_be_table_count=len(to_be_schema) if to_be_schema else None,
    )


def _collect_schema_mismatches(
    mapping: Mapping,
    as_is_schema: Optional[Dict[str, Set[str]]],
    to_be_schema: Optional[Dict[str, Set[str]]],
) -> List[str]:
    out: List[str] = []
    if as_is_schema is not None:
        for tm in mapping.tables:
            for n in tm.as_is_tables():
                if n.upper() not in as_is_schema:
                    out.append(f"[tables.as_is] '{n}' not in AS-IS schema")
        for (t, c), _cm in mapping.column_by_as_is.items():
            cols = as_is_schema.get(t)
            if cols is not None and c not in cols:
                out.append(f"[columns.as_is] '{t}.{c}' not in AS-IS schema")
    if to_be_schema is not None:
        for tm in mapping.tables:
            for n in tm.to_be_tables():
                if n.upper() not in to_be_schema:
                    out.append(f"[tables.to_be] '{n}' not in TO-BE schema")
        for _key, cm in mapping.column_by_as_is.items():
            for ref in cm.to_be_refs():
                table = ref.table.upper()
                column = ref.column.upper()
                cols = to_be_schema.get(table)
                if cols is not None and column not in cols:
                    out.append(
                        f"[columns.to_be] '{table}.{column}' not in TO-BE schema"
                    )
    return out


def _scan_statements(
    statements: List[dict],
    mapping: Mapping,
    mybatis_dir: Path,
) -> Tuple[
    List[StatementImpact],
    Dict[str, int],
    Dict[str, Set[str]],
    Dict[Tuple[str, str], int],
    Dict[Tuple[str, str], Set[str]],
    Dict[str, int],
]:
    table_stmt: Dict[str, int] = {}
    table_files: Dict[str, Set[str]] = {}
    column_stmt: Dict[Tuple[str, str], int] = {}
    column_files: Dict[Tuple[str, str], Set[str]] = {}
    per_type: Dict[str, int] = {}
    stmt_impacts: List[StatementImpact] = []

    # Pre-bucket columns by AS-IS table for O(1) lookup per stmt
    cols_by_table: Dict[str, List[Tuple[str, ColumnMapping]]] = {}
    for (t, c), cm in mapping.column_by_as_is.items():
        cols_by_table.setdefault(t, []).append((c, cm))

    # Cache compiled column-name regexes (word boundary, upper)
    col_regex: Dict[str, re.Pattern] = {}

    def _get_col_re(col: str) -> re.Pattern:
        if col not in col_regex:
            col_regex[col] = re.compile(rf"\b{re.escape(col)}\b")
        return col_regex[col]

    for stmt in statements:
        sql_upper = stmt["sql"].upper()
        stype = stmt.get("type", "?")
        per_type[stype] = per_type.get(stype, 0) + 1

        per_usage = extract_table_usage([stmt])
        used_tables_upper = {k.upper() for k in per_usage.keys()}

        tables_hit: List[str] = []
        cols_hit: List[str] = []
        kinds_hit: Set[str] = set()

        for tu in sorted(used_tables_upper):
            tm = mapping.find_table(tu)
            if tm is not None:
                tables_hit.append(tu)
                kinds_hit.add(tm.type)

            for col, cm in cols_by_table.get(tu, []):
                if _get_col_re(col).search(sql_upper):
                    cols_hit.append(f"{tu}.{col}")
                    kinds_hit.add(cm.kind)

        try:
            rel = str(Path(stmt["mapper_path"]).relative_to(mybatis_dir))
        except ValueError:
            rel = stmt.get("mapper_path", stmt.get("mapper", ""))

        stmt_impacts.append(StatementImpact(
            xml_file=rel,
            namespace=stmt.get("namespace", ""),
            sql_id=stmt.get("id", ""),
            sql_type=stype,
            tables_hit=tables_hit,
            columns_hit=cols_hit,
            kinds_hit=kinds_hit,
        ))

        for th in tables_hit:
            table_stmt[th] = table_stmt.get(th, 0) + 1
            table_files.setdefault(th, set()).add(rel)
        for ch in cols_hit:
            key_t, key_c = ch.split(".", 1)
            k = (key_t, key_c)
            column_stmt[k] = column_stmt.get(k, 0) + 1
            column_files.setdefault(k, set()).add(rel)

    return stmt_impacts, table_stmt, table_files, column_stmt, column_files, per_type


def _build_table_impacts(
    mapping: Mapping,
    table_stmt: Dict[str, int],
    table_files: Dict[str, Set[str]],
) -> List[TableImpact]:
    out: List[TableImpact] = []
    seen: Set[str] = set()
    for tm in mapping.tables:
        for n in tm.as_is_tables():
            key = n.upper()
            if key in seen:
                continue
            seen.add(key)
            files = sorted(table_files.get(key, set()))
            to_be_str = ",".join(tm.to_be_tables()) or "<dropped>"
            out.append(TableImpact(
                table=key,
                mapping_kind=tm.type,
                to_be=to_be_str,
                statement_count=table_stmt.get(key, 0),
                xml_file_count=len(files),
                xml_files=files,
            ))
    return out


def _build_column_impacts(
    mapping: Mapping,
    column_stmt: Dict[Tuple[str, str], int],
    column_files: Dict[Tuple[str, str], Set[str]],
) -> List[ColumnImpact]:
    out: List[ColumnImpact] = []
    seen: Set[Tuple[str, str]] = set()
    for (t, c), cm in mapping.column_by_as_is.items():
        if (t, c) in seen:
            continue
        seen.add((t, c))
        if cm.to_be is None:
            to_be_str = "<dropped>"
        elif isinstance(cm.to_be, list):
            to_be_str = ",".join(f"{x.table}.{x.column}" for x in cm.to_be)
        else:
            to_be_str = f"{cm.to_be.table}.{cm.to_be.column}"
        files = sorted(column_files.get((t, c), set()))
        out.append(ColumnImpact(
            table=t,
            column=c,
            mapping_kind=cm.kind,
            to_be=to_be_str,
            statement_count=column_stmt.get((t, c), 0),
            xml_file_count=len(files),
            xml_files=files,
        ))
    return out


# ---------------------------------------------------------------------------
# Reporting (stdout + Excel)
# ---------------------------------------------------------------------------


def print_impact_summary(report: ImpactReport) -> None:
    """Count-based stdout summary, mirroring the style of existing commands."""
    print()
    print("=== Migration Impact Summary ===")
    print(f"Mapping           : {report.mapping_path}")
    print(f"MyBatis dir       : {report.mybatis_dir}")
    if report.as_is_schema_path:
        print(f"AS-IS schema      : {report.as_is_schema_path} "
              f"({report.as_is_table_count} tables)")
    if report.to_be_schema_path:
        print(f"TO-BE schema      : {report.to_be_schema_path} "
              f"({report.to_be_table_count} tables)")
    print()
    print(f"Statements        : {report.statement_count}")
    if report.statement_count:
        pct = report.affected_statement_count / report.statement_count * 100
        print(f"Affected          : {report.affected_statement_count} "
              f"({pct:.1f}%)")
    else:
        print(f"Affected          : 0")
    if report.statements_per_type:
        types_str = ", ".join(
            f"{t}:{c}" for t, c in sorted(report.statements_per_type.items())
        )
        print(f"By type           : {types_str}")
    print()
    print(f"Table mappings    : {len(report.table_impacts)} "
          f"(unused: {len(report.unused_tables)})")
    print(f"Column mappings   : {len(report.column_impacts)} "
          f"(unused: {len(report.unused_columns)})")

    if report.loader_errors:
        print()
        print(f"Loader errors     : {len(report.loader_errors)}")
        for e in report.loader_errors:
            print(f"  - {e}")
    if report.schema_mismatches:
        print()
        print(f"Schema mismatches : {len(report.schema_mismatches)}")
        for sm in report.schema_mismatches:
            print(f"  - {sm}")


def write_impact_excel(report: ImpactReport, output_path: Path) -> None:
    """Render a 5-sheet xlsx report."""
    wb = Workbook()

    _write_summary_sheet(wb, report)
    _write_table_impact_sheet(wb, report)
    _write_column_impact_sheet(wb, report)
    _write_affected_statements_sheet(wb, report)
    _write_validation_sheet(wb, report)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Wrote impact report: %s", output_path)


# --- Sheet writers ---------------------------------------------------------


_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)
_UNUSED_FILL = PatternFill("solid", fgColor="E0E0E0")
_WARN_FILL = PatternFill("solid", fgColor="FFF5CC")
_ERROR_FILL = PatternFill("solid", fgColor="FFE4E1")


def _bold_header(ws) -> None:
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _write_summary_sheet(wb: Workbook, report: ImpactReport) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.append(["SQL Migration Impact Report"])
    ws["A1"].font = _TITLE_FONT
    ws.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["Mapping", str(report.mapping_path)])
    ws.append(["MyBatis dir", str(report.mybatis_dir)])
    ws.append(["AS-IS schema", str(report.as_is_schema_path or "-")])
    ws.append(["TO-BE schema", str(report.to_be_schema_path or "-")])
    ws.append([])
    ws.append(["Total statements", report.statement_count])
    ws.append(["Affected statements", report.affected_statement_count])
    if report.statement_count:
        pct = report.affected_statement_count / report.statement_count * 100
        ws.append(["Affected %", f"{pct:.1f}%"])
    ws.append([])
    ws.append(["Statements by type"])
    for t, c in sorted(report.statements_per_type.items()):
        ws.append([f"  {t}", c])
    ws.append([])
    ws.append(["Mapped tables", len(report.table_impacts)])
    ws.append(["Mapped columns", len(report.column_impacts)])
    ws.append(["Unused table mappings", len(report.unused_tables)])
    ws.append(["Unused column mappings", len(report.unused_columns)])
    ws.append([])
    ws.append(["Loader errors", len(report.loader_errors)])
    ws.append(["Schema mismatches", len(report.schema_mismatches)])

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60


def _write_table_impact_sheet(wb: Workbook, report: ImpactReport) -> None:
    ws = wb.create_sheet("Table Impact")
    ws.append(
        ["AS-IS Table", "Kind", "TO-BE", "Statement count",
         "XML file count", "XML files (first 5)", "Unused"]
    )
    _bold_header(ws)
    rows = sorted(
        report.table_impacts, key=lambda x: (-x.statement_count, x.table)
    )
    for ti in rows:
        files_preview = ", ".join(ti.xml_files[:5])
        if len(ti.xml_files) > 5:
            files_preview += f", ... (+{len(ti.xml_files) - 5} more)"
        ws.append([
            ti.table,
            ti.mapping_kind,
            ti.to_be,
            ti.statement_count,
            ti.xml_file_count,
            files_preview,
            "YES" if ti.statement_count == 0 else "",
        ])
        if ti.statement_count == 0:
            for cell in ws[ws.max_row]:
                cell.fill = _UNUSED_FILL
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["C"].width = 36
    ws.column_dimensions["F"].width = 60


def _write_column_impact_sheet(wb: Workbook, report: ImpactReport) -> None:
    ws = wb.create_sheet("Column Impact")
    ws.append(
        ["AS-IS Column", "Kind", "TO-BE", "Statement count",
         "XML file count", "XML files (first 5)", "Unused"]
    )
    _bold_header(ws)
    rows = sorted(
        report.column_impacts,
        key=lambda x: (-x.statement_count, x.table, x.column),
    )
    for ci in rows:
        files_preview = ", ".join(ci.xml_files[:5])
        if len(ci.xml_files) > 5:
            files_preview += f", ... (+{len(ci.xml_files) - 5} more)"
        ws.append([
            f"{ci.table}.{ci.column}",
            ci.mapping_kind,
            ci.to_be,
            ci.statement_count,
            ci.xml_file_count,
            files_preview,
            "YES" if ci.statement_count == 0 else "",
        ])
        if ci.statement_count == 0:
            for cell in ws[ws.max_row]:
                cell.fill = _UNUSED_FILL
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["F"].width = 60


def _write_affected_statements_sheet(wb: Workbook, report: ImpactReport) -> None:
    ws = wb.create_sheet("Affected Statements")
    ws.append(
        ["No", "XML File", "Namespace", "SQL ID", "Type",
         "Tables hit", "Columns hit", "Kinds"]
    )
    _bold_header(ws)
    n = 0
    for si in report.statement_impacts:
        if not (si.tables_hit or si.columns_hit):
            continue
        n += 1
        ws.append([
            n,
            si.xml_file,
            si.namespace,
            si.sql_id,
            si.sql_type,
            ", ".join(si.tables_hit),
            ", ".join(si.columns_hit),
            ", ".join(sorted(si.kinds_hit)),
        ])
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["F"].width = 36
    ws.column_dimensions["G"].width = 40


def _write_validation_sheet(wb: Workbook, report: ImpactReport) -> None:
    ws = wb.create_sheet("Validation")
    ws.append(["Category", "Location", "Message"])
    _bold_header(ws)
    for le in report.loader_errors:
        ws.append(["LoaderError", le.location or "-", le.message])
        for cell in ws[ws.max_row]:
            cell.fill = _ERROR_FILL
    for sm in report.schema_mismatches:
        ws.append(["SchemaMismatch", "-", sm])
        for cell in ws[ws.max_row]:
            cell.fill = _WARN_FILL
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70
