"""Load the DB menu table and flatten it into program-level entries.

Each program (= leaf menu node) is returned with its ancestry resolved into
``main_menu`` / ``sub_menu`` / ``tab`` / ``program_name``, matching the
output columns expected by the legacy analyzer.

The menu table is highly project-specific so column names are configurable
via ``config.yaml``:

```
legacy:
  menu:
    table: "TB_MENU"
    columns:
      program_id: "PROGRAM_ID"
      program_nm: "PROGRAM_NM"
      url:        "URL"
      parent_id:  "PARENT_ID"
      level:      "LEVEL"
```
"""

import logging

from .legacy_util import normalize_url

logger = logging.getLogger(__name__)


DEFAULT_COLUMNS = {
    "program_id": "PROGRAM_ID",
    "program_nm": "PROGRAM_NM",
    "url": "URL",
    "parent_id": "PARENT_ID",
    "level": "LEVEL",
}


def _menu_config(config: dict) -> tuple[str, dict]:
    """Resolve ``table`` and ``columns`` from config with sane defaults."""
    legacy_cfg = (config or {}).get("legacy", {}) or {}
    menu_cfg = legacy_cfg.get("menu", {}) or {}
    table = menu_cfg.get("table", "TB_MENU")
    columns = {**DEFAULT_COLUMNS, **(menu_cfg.get("columns") or {})}
    return table, columns


def load_menu_rows(config: dict, table_override: str | None = None) -> list[dict]:
    """Query the configured menu table and return a list of raw rows.

    Each row is normalized to the internal key set
    ``program_id/program_nm/url/parent_id/level`` regardless of the
    project's column names. Rows with ``url IS NULL`` are kept because
    parent nodes often have no URL — the tree walk needs them.
    """
    from .db import execute_query, get_connection  # lazy: oracledb is optional

    table, columns = _menu_config(config)
    if table_override:
        table = table_override

    sql = (
        f'SELECT {columns["program_id"]} AS program_id, '
        f'{columns["program_nm"]} AS program_nm, '
        f'{columns["url"]} AS url, '
        f'{columns["parent_id"]} AS parent_id, '
        f'{columns["level"]} AS "LEVEL" '
        f'FROM {table}'
    )
    connection = get_connection(config)
    try:
        col_names, raw = execute_query(connection, sql)
    finally:
        connection.close()

    idx = {c.upper(): i for i, c in enumerate(col_names)}
    rows = []
    for r in raw:
        rows.append({
            "program_id": r[idx["PROGRAM_ID"]],
            "program_nm": r[idx["PROGRAM_NM"]],
            "url": r[idx["URL"]],
            "parent_id": r[idx["PARENT_ID"]],
            "level": r[idx["LEVEL"]],
        })
    logger.info("Loaded %d menu rows from %s", len(rows), table)
    return rows


def build_menu_tree(rows: list[dict]) -> list[dict]:
    """Turn the raw menu rows into a list of program entries with ancestry.

    The convention we apply to the resolved hierarchy:

    * level 1 ancestor → ``main_menu``
    * level 2 ancestor → ``sub_menu``
    * level 3 ancestor → ``tab``
    * leaf itself      → ``program_name``

    Works for any menu depth: the deepest three ancestors map onto
    main/sub/tab, the leaf name becomes ``program_name``, and anything
    shallower than level 3 leaves the unused slots empty.

    Only leaves with a non-empty URL produce program entries; parent-only
    rows are skipped but still used for the ancestry walk.
    """
    by_id = {r["program_id"]: r for r in rows}
    children = {}
    for r in rows:
        children.setdefault(r.get("parent_id"), []).append(r)

    def _ancestry(row: dict) -> list[dict]:
        chain = []
        cur = row
        safety = 0
        while cur and safety < 20:
            chain.append(cur)
            cur = by_id.get(cur.get("parent_id"))
            safety += 1
        chain.reverse()  # root-first
        return chain

    programs = []
    for r in rows:
        if not r.get("url"):
            continue
        chain = _ancestry(r)
        # Skip if not a leaf (has children with URLs pointing elsewhere)
        # We still emit nodes that happen to be branches with a URL — those
        # are navigation shortcuts, not pure containers.
        main_menu = chain[0]["program_nm"] if len(chain) >= 1 else ""
        sub_menu = chain[1]["program_nm"] if len(chain) >= 2 else ""
        tab = chain[2]["program_nm"] if len(chain) >= 3 else ""
        # ``program_name`` is always the leaf row name for visibility
        program_name = r["program_nm"]
        # If chain is only 1 deep, main_menu IS the program; blank sub/tab.
        if len(chain) == 1:
            sub_menu = ""
            tab = ""
        # If chain is 2 deep, the leaf is itself the sub_menu level. Keep
        # main_menu as the root, leave tab blank, and set program_name.
        if len(chain) == 2 and tab:
            tab = ""
        programs.append({
            "program_id": r["program_id"],
            "program_name": program_name,
            "main_menu": main_menu,
            "sub_menu": sub_menu,
            "tab": tab,
            "url": r["url"],
        })
    logger.info("Built menu tree: %d programs with URL", len(programs))
    return programs


def build_url_index(programs: list[dict]) -> dict:
    """Return ``{normalized_url: program_entry}``.

    Duplicate URLs (same normalized key) resolve to the first entry — those
    are typically data bugs in the menu table and are surfaced implicitly by
    the analyzer's "orphan menu" section.
    """
    idx = {}
    for p in programs:
        key = normalize_url(p.get("url", ""))
        if not key:
            continue
        idx.setdefault(key, p)
    return idx


def load_menu_hierarchy(config: dict, table_override: str | None = None) -> list[dict]:
    """Convenience: load + flatten into program entries in one call."""
    rows = load_menu_rows(config, table_override=table_override)
    return build_menu_tree(rows)


# ---------------------------------------------------------------------------
# Excel-based menu loader
# ---------------------------------------------------------------------------

# Header keywords we accept for each level. We try Korean first since most
# legacy projects ship Korean-labelled menus, then fall back to English.
_LEVEL_KEYWORDS = {
    1: ("1레벨", "level1", "lv1", "lvl1", "대분류"),
    2: ("2레벨", "level2", "lv2", "lvl2", "중분류"),
    3: ("3레벨", "level3", "lv3", "lvl3", "소분류"),
    4: ("4레벨", "level4", "lv4", "lvl4", "세분류"),
    5: ("5레벨", "level5", "lv5", "lvl5", "최하위"),
}
_URL_KEYWORDS = ("url", "uri", "경로", "path", "link", "endpoint")


def _norm_header(value) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace(" ", "")


def _find_header_indexes(header_row) -> dict:
    """Return ``{slot: column_index}`` for level1..5 and url, by header text.

    ``slot`` is one of ``"l1"`` .. ``"l5"`` and ``"url"``. Headers are
    matched case-insensitively against the synonym lists above. If a
    header is not present, the slot is omitted.
    """
    norm = [_norm_header(c) for c in header_row]
    result = {}
    for level, keywords in _LEVEL_KEYWORDS.items():
        for i, h in enumerate(norm):
            if not h:
                continue
            if any(k in h for k in keywords):
                result[f"l{level}"] = i
                break
    for i, h in enumerate(norm):
        if not h:
            continue
        if any(k == h or h.endswith(k) for k in _URL_KEYWORDS):
            result["url"] = i
            break
    return result


def _row_to_entry(row, header_idx: dict, row_num: int) -> dict | None:
    """Convert one Excel row into a program entry.

    Empty rows (no level / no url) are skipped. The deepest non-empty
    level becomes ``program_name``; the first three non-empty levels are
    placed in ``main_menu`` / ``sub_menu`` / ``tab`` (the legacy slots).
    Every non-empty level is preserved in ``menu_path`` joined with
    `` > `` so 4th and 5th-level information is not lost.
    """
    levels = []
    for n in range(1, 6):
        idx = header_idx.get(f"l{n}")
        if idx is None or idx >= len(row):
            continue
        cell = row[idx]
        if cell is None:
            continue
        text = str(cell).strip()
        if text:
            levels.append(text)

    url_idx = header_idx.get("url")
    url = ""
    if url_idx is not None and url_idx < len(row) and row[url_idx] is not None:
        url = str(row[url_idx]).strip()

    if not levels and not url:
        return None
    if not url:
        # Pure container row — keep the ancestry but no callable URL,
        # so it never matches a controller. Return None to skip; the
        # leaf rows above already hold the full ancestry path.
        return None
    if not levels:
        # URL present but no level labels — synthesise a placeholder.
        levels = [f"(menu row {row_num})"]

    program_name = levels[-1]
    main_menu = levels[0] if len(levels) >= 1 else ""
    sub_menu = levels[1] if len(levels) >= 2 else ""
    tab = levels[2] if len(levels) >= 3 else ""
    if len(levels) == 1:
        # Single-level menu — leaf is itself the main_menu; clear sub/tab.
        main_menu = levels[0]
        sub_menu = ""
        tab = ""

    return {
        "program_id": "",
        "program_name": program_name,
        "main_menu": main_menu,
        "sub_menu": sub_menu,
        "tab": tab,
        "menu_path": " > ".join(levels),
        "url": url,
    }


def load_menu_from_excel(xlsx_path: str, sheet_name: str | None = None) -> list[dict]:
    """Load a project-specific menu Excel and return program entries.

    Expected sheet layout::

        | 1레벨 | 2레벨 | 3레벨 | 4레벨 | 5레벨 | URL |
        | 주문 | 주문조회 |        |        |        | /api/order/list |
        | 주문 | 주문등록 |        |        |        | /api/order/save |
        | 설비 | 설비관리 | 모델링 | SVID 코드 |   | /api/svid/list |

    * Header text matches a small set of Korean / English synonyms (see
      ``_LEVEL_KEYWORDS`` and ``_URL_KEYWORDS``).
    * Rows whose ``URL`` is empty are treated as pure container nodes
      and skipped — only callable pages (rows with a URL) become
      program entries.
    * The deepest non-empty level becomes ``program_name``. The first
      three are mapped onto the legacy ``main_menu`` / ``sub_menu`` /
      ``tab`` slots; **all** non-empty levels are also preserved in a
      new ``menu_path`` field (e.g. ``설비 > 설비관리 > 모델링 >
      SVID 코드``) so 4th / 5th-level information is not lost.

    Raises FileNotFoundError if the path is missing and ImportError if
    openpyxl is unavailable.
    """
    if not xlsx_path:
        return []
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise ImportError(
            "openpyxl is required to read --menu-xlsx. Install with `pip install openpyxl`."
        ) from e

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        logger.warning("Menu Excel %s is empty", xlsx_path)
        return []

    header_idx = _find_header_indexes(header_row)
    if "url" not in header_idx:
        logger.warning(
            "Menu Excel %s has no URL column (looked for: %s) — every row will be skipped",
            xlsx_path, ", ".join(_URL_KEYWORDS),
        )
        return []
    if not any(k.startswith("l") for k in header_idx):
        logger.warning(
            "Menu Excel %s has no level columns (looked for 1레벨..5레벨)",
            xlsx_path,
        )

    programs = []
    for line_no, row in enumerate(rows_iter, start=2):
        entry = _row_to_entry(row, header_idx, line_no)
        if entry is not None:
            programs.append(entry)

    logger.info(
        "Loaded %d menu programs from Excel: %s (sheet=%s)",
        len(programs), xlsx_path, ws.title,
    )
    return programs


# ---------------------------------------------------------------------------
# Markdown-based menu loader
# ---------------------------------------------------------------------------

def load_menu_from_markdown(md_path: str) -> list[dict]:
    """Load a project-specific menu Markdown table and return program entries.

    Expected format — a standard Markdown pipe table with the same header
    synonyms accepted by the Excel loader::

        | 1레벨 | 2레벨 | 3레벨 | 4레벨 | 5레벨 | URL |
        |-------|-------|-------|-------|-------|-----|
        | 주문관리 | 주문조회 | | | | /api/order/list |

    * Separator rows (``|---|``) are skipped automatically.
    * Cells are stripped of leading/trailing whitespace.
    * Rows whose URL cell is empty are skipped (container nodes).
    * DRM-free alternative to ``--menu-xlsx`` for restricted environments.

    Raises ``FileNotFoundError`` if the path is missing.
    """
    if not md_path:
        return []

    with open(md_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    # Find the first pipe-table header row
    header_row = None
    header_line_idx = -1
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith("|") and "|" in stripped[1:]:
            # Skip separator rows like |---|---|
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if all(c.replace("-", "").replace(":", "") == "" for c in cells):
                continue
            header_row = cells
            header_line_idx = i
            break

    if header_row is None:
        logger.warning("Menu Markdown %s has no pipe-table header", md_path)
        return []

    header_idx = _find_header_indexes(header_row)
    if "url" not in header_idx:
        logger.warning(
            "Menu Markdown %s has no URL column (looked for: %s)",
            md_path, ", ".join(_URL_KEYWORDS),
        )
        return []

    programs = []
    for line_no, line in enumerate(lines[header_line_idx + 1:], start=header_line_idx + 2):
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        # Skip separator rows
        if all(c.replace("-", "").replace(":", "") == "" for c in cells):
            continue
        entry = _row_to_entry(cells, header_idx, line_no)
        if entry is not None:
            programs.append(entry)

    logger.info(
        "Loaded %d menu programs from Markdown: %s",
        len(programs), md_path,
    )
    return programs
