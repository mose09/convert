import logging
import os
from collections import defaultdict
from datetime import datetime

logger = logging.getLogger(__name__)


def audit_schema(schema: dict, terms_dict_path: str = None) -> dict:
    """Audit the entire schema for naming standard violations."""
    from .naming_validator import NamingValidator

    validator = NamingValidator(terms_dict_path=terms_dict_path)

    table_violations = []
    column_violations = []
    pattern_counts = defaultdict(int)
    severity_counts = defaultdict(int)

    total_tables = 0
    total_columns = 0
    invalid_tables = 0
    invalid_columns = 0

    for table in schema.get("tables", []):
        total_tables += 1
        table_name = table["name"]

        # Validate table name
        t_result = validator.validate_name(table_name, kind="table")
        if not t_result["valid"]:
            invalid_tables += 1
            for issue in t_result["issues"]:
                pattern_counts[issue["rule"]] += 1
                severity_counts[issue["severity"]] += 1

            table_violations.append({
                "table": table_name,
                "comment": table.get("comment") or "",
                "issues": t_result["issues"],
                "suggestions": t_result.get("suggestions", {}),
            })

        # Validate each column
        for col in table["columns"]:
            total_columns += 1
            col_name = col["column_name"]
            c_result = validator.validate_name(col_name, kind="column")
            if not c_result["valid"]:
                invalid_columns += 1
                for issue in c_result["issues"]:
                    pattern_counts[issue["rule"]] += 1
                    severity_counts[issue["severity"]] += 1

                column_violations.append({
                    "table": table_name,
                    "column": col_name,
                    "comment": col.get("comment") or "",
                    "data_type": col.get("data_type") or "",
                    "issues": c_result["issues"],
                    "suggestions": c_result.get("suggestions", {}),
                })

    logger.info("Audit complete: %d tables (%d invalid), %d columns (%d invalid)",
                total_tables, invalid_tables, total_columns, invalid_columns)

    return {
        "total_tables": total_tables,
        "invalid_tables": invalid_tables,
        "total_columns": total_columns,
        "invalid_columns": invalid_columns,
        "table_violations": table_violations,
        "column_violations": column_violations,
        "pattern_counts": dict(pattern_counts),
        "severity_counts": dict(severity_counts),
    }


def save_audit_markdown(audit: dict, output_dir: str) -> str:
    """Save audit report as Markdown."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"audit_standards_{timestamp}.md")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# Naming Standards Audit Report\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # Summary
        f.write("## Summary\n\n")
        f.write(f"- Total tables: {audit['total_tables']}\n")
        f.write(f"- Invalid tables: {audit['invalid_tables']} ({_pct(audit['invalid_tables'], audit['total_tables'])}%)\n")
        f.write(f"- Total columns: {audit['total_columns']}\n")
        f.write(f"- Invalid columns: {audit['invalid_columns']} ({_pct(audit['invalid_columns'], audit['total_columns'])}%)\n\n")

        # Severity breakdown
        f.write("## Severity Breakdown\n\n")
        f.write("| Severity | Count |\n")
        f.write("|----------|-------|\n")
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
            f.write(f"| {sev} | {audit['severity_counts'].get(sev, 0)} |\n")
        f.write("\n")

        # Pattern breakdown
        f.write("## Violation Patterns\n\n")
        f.write("| Rule | Count |\n")
        f.write("|------|-------|\n")
        sorted_patterns = sorted(audit["pattern_counts"].items(), key=lambda x: -x[1])
        for rule, cnt in sorted_patterns:
            f.write(f"| {rule} | {cnt} |\n")
        f.write("\n")

        # Invalid tables detail
        if audit["table_violations"]:
            f.write(f"## Invalid Tables ({len(audit['table_violations'])})\n\n")
            f.write("| Table | Comment | Severity | Rule | Message |\n")
            f.write("|-------|---------|----------|------|---------|\n")
            for v in audit["table_violations"]:
                for issue in v["issues"]:
                    f.write(f"| {v['table']} | {v['comment']} | {issue['severity']} "
                            f"| {issue['rule']} | {issue['message']} |\n")
            f.write("\n")

        # Invalid columns (top 100)
        if audit["column_violations"]:
            f.write(f"## Invalid Columns ({len(audit['column_violations'])})\n\n")
            f.write("상위 100건만 표시 (전체는 Excel 참고)\n\n")
            f.write("| Table | Column | Type | Comment | Severity | Rule | Message |\n")
            f.write("|-------|--------|------|---------|----------|------|---------|\n")
            for v in audit["column_violations"][:100]:
                for issue in v["issues"]:
                    f.write(f"| {v['table']} | {v['column']} | {v['data_type']} "
                            f"| {v['comment']} | {issue['severity']} | {issue['rule']} "
                            f"| {issue['message']} |\n")
            if len(audit["column_violations"]) > 100:
                f.write(f"| ... | +{len(audit['column_violations']) - 100} more | | | | | |\n")
            f.write("\n")

    logger.info("Audit markdown saved: %s", filepath)
    return filepath


def save_audit_excel(audit: dict, output_dir: str) -> str:
    """Save audit report as Excel with multiple sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"audit_standards_{timestamp}.xlsx")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    severity_colors = {
        "CRITICAL": "C00000",
        "HIGH": "FF6600",
        "MEDIUM": "FFC000",
        "LOW": "70AD47",
    }

    def _sev_fill(severity: str):
        color = severity_colors.get(severity, "DDDDDD")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    def _write_row(ws, row_num, values, severity=None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            if severity and col == 1:
                cell.fill = _sev_fill(severity)
                cell.font = Font(bold=True, color="FFFFFF")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, ["Category", "Count", "Percent"])
    ws.cell(row=2, column=1, value="Total Tables").border = thin_border
    ws.cell(row=2, column=2, value=audit["total_tables"]).border = thin_border
    ws.cell(row=2, column=3, value="100%").border = thin_border

    ws.cell(row=3, column=1, value="Invalid Tables").border = thin_border
    ws.cell(row=3, column=2, value=audit["invalid_tables"]).border = thin_border
    ws.cell(row=3, column=3, value=f"{_pct(audit['invalid_tables'], audit['total_tables'])}%").border = thin_border

    ws.cell(row=4, column=1, value="Total Columns").border = thin_border
    ws.cell(row=4, column=2, value=audit["total_columns"]).border = thin_border
    ws.cell(row=4, column=3, value="100%").border = thin_border

    ws.cell(row=5, column=1, value="Invalid Columns").border = thin_border
    ws.cell(row=5, column=2, value=audit["invalid_columns"]).border = thin_border
    ws.cell(row=5, column=3, value=f"{_pct(audit['invalid_columns'], audit['total_columns'])}%").border = thin_border

    # Severity section
    row = 7
    ws.cell(row=row, column=1, value="Severity").font = Font(bold=True, size=12)
    row += 1
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        _write_row(ws, row, [sev, audit["severity_counts"].get(sev, 0), ""], severity=sev)
        row += 1

    _auto_width(ws)

    # Sheet 2: Invalid tables
    ws_tables = wb.create_sheet("Invalid Tables")
    _write_header(ws_tables, ["Table", "Comment", "Severity", "Rule", "Message", "Suggestions"])
    row = 2
    for v in audit["table_violations"]:
        for issue in v["issues"]:
            sugg = ""
            if issue.get("suggestions"):
                sugg = "; ".join(
                    f"{u}→{','.join(s)}" for u, s in issue["suggestions"].items()
                )
            _write_row(ws_tables, row, [
                v["table"], v["comment"], issue["severity"],
                issue["rule"], issue["message"], sugg,
            ], severity=issue["severity"])
            row += 1
    _auto_width(ws_tables)

    # Sheet 3: Invalid columns
    ws_cols = wb.create_sheet("Invalid Columns")
    _write_header(ws_cols, ["Table", "Column", "Type", "Comment", "Severity", "Rule", "Message", "Suggestions"])
    row = 2
    for v in audit["column_violations"]:
        for issue in v["issues"]:
            sugg = ""
            if issue.get("suggestions"):
                sugg = "; ".join(
                    f"{u}→{','.join(s)}" for u, s in issue["suggestions"].items()
                )
            _write_row(ws_cols, row, [
                v["table"], v["column"], v["data_type"], v["comment"],
                issue["severity"], issue["rule"], issue["message"], sugg,
            ], severity=issue["severity"])
            row += 1
    _auto_width(ws_cols)

    # Sheet 4: Pattern summary
    ws_pattern = wb.create_sheet("Pattern Summary")
    _write_header(ws_pattern, ["Rule", "Count"])
    sorted_patterns = sorted(audit["pattern_counts"].items(), key=lambda x: -x[1])
    for i, (rule, cnt) in enumerate(sorted_patterns, 2):
        _write_row(ws_pattern, i, [rule, cnt])
    _auto_width(ws_pattern)

    wb.save(filepath)
    logger.info("Audit excel saved: %s", filepath)
    return filepath


def _pct(part: int, total: int) -> str:
    if total == 0:
        return "0.0"
    return f"{part / total * 100:.1f}"
