"""AG Grid columnDefs 의 (field, headerName) 페어 추출.

이 모듈은 React 소스 디렉토리를 재귀 walk 해서 AG Grid 의 컬럼 정의
객체 (`{ headerName: '데이터타입', field: 'DATA_TYPE' }`) 를 regex 로
모아준다. AST/LLM 의존 없음.

매칭 정책:
- `headerName: '...'` 과 `field: '...'` 가 같은 객체 안에 함께 등장
  하는 경우만 페어. 키 순서 무관.
- 같은 파일에 다수 등장 시 위치가 가장 가까운 (greedy) 페어로 매칭하고
  한 번 매칭된 field 는 재사용 X (두 헤더가 한 field 를 공유하는 사고
  방지).
- 거리 임계값 = ``_MAX_PAIR_DISTANCE`` (default 200 chars) — 같은 객체
  literal 안 쪽 페어만 잡기 위한 보수적 한도.

값에 ``${...}`` 같은 template 표현이 들어간 경우는 ``[^'"`]+`` 매칭에서
자연스럽게 제외된다 (의도된 동작).
"""

from __future__ import annotations

import os
import re
from collections import defaultdict


_HEADER_RE = re.compile(r"""headerName\s*:\s*['"`]([^'"`\n]+)['"`]""")
_FIELD_RE = re.compile(r"""field\s*:\s*['"`]([^'"`\n]+)['"`]""")

_EXTENSIONS = {".js", ".jsx", ".ts", ".tsx", ".mjs"}
_SKIP_DIRS = {
    "node_modules", "build", "dist", ".next", ".git", ".cache",
    "coverage", "__tests__", "__mocks__", "public", "storybook",
    "storybook-static", ".storybook", ".turbo", ".vercel",
}
_MAX_PAIR_DISTANCE = 200


def _iter_source_files(react_dir: str):
    for root, dirs, files in os.walk(react_dir):
        dirs[:] = [d for d in dirs if d not in _SKIP_DIRS]
        for f in files:
            ext = os.path.splitext(f)[1].lower()
            if ext in _EXTENSIONS:
                yield os.path.join(root, f)


def _pairs_in_content(content: str) -> list[tuple[str, str, int]]:
    """파일 1개에서 (field, label, line) 리스트 반환."""
    headers = [(m.start(), m.group(1)) for m in _HEADER_RE.finditer(content)]
    fields = [(m.start(), m.group(1)) for m in _FIELD_RE.finditer(content)]
    if not headers or not fields:
        return []

    used = [False] * len(fields)
    pairs: list[tuple[str, str, int]] = []
    for h_pos, label in headers:
        best_j = -1
        best_d = _MAX_PAIR_DISTANCE + 1
        for j, (f_pos, _) in enumerate(fields):
            if used[j]:
                continue
            d = abs(f_pos - h_pos)
            if d < best_d:
                best_d = d
                best_j = j
        if best_j >= 0 and best_d <= _MAX_PAIR_DISTANCE:
            used[best_j] = True
            line = content[:h_pos].count("\n") + 1
            pairs.append((fields[best_j][1], label, line))
    return pairs


def extract_grid_labels(react_dir: str) -> list[dict]:
    """``react_dir`` 재귀 walk → AG Grid (field, headerName) 페어 집계.

    Returns
    -------
    list[dict]
        ``[{"field": ..., "label": ..., "count": N,
           "file_count": M, "sample_sources": [...] }, ...]``
        count 내림차순 정렬.
    """
    if not os.path.isdir(react_dir):
        raise FileNotFoundError(f"react_dir not found: {react_dir}")

    aggregate: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"count": 0, "sources": [], "files": set()}
    )
    file_scanned = 0
    file_hit = 0

    for fp in _iter_source_files(react_dir):
        file_scanned += 1
        try:
            with open(fp, "r", encoding="utf-8", errors="ignore") as fh:
                content = fh.read()
        except Exception:
            continue
        pairs = _pairs_in_content(content)
        if not pairs:
            continue
        file_hit += 1
        rel = os.path.relpath(fp, react_dir)
        for field, label, line in pairs:
            key = (field, label)
            entry = aggregate[key]
            entry["count"] += 1
            entry["files"].add(rel)
            if len(entry["sources"]) < 5:
                src = f"{rel}:{line}"
                if src not in entry["sources"]:
                    entry["sources"].append(src)

    rows = [
        {
            "field": field,
            "label": label,
            "count": entry["count"],
            "file_count": len(entry["files"]),
            "sample_sources": entry["sources"],
        }
        for (field, label), entry in aggregate.items()
    ]
    rows.sort(key=lambda r: (-r["count"], r["field"], r["label"]))
    return rows


def save_grid_labels_excel(rows: list[dict], output_dir: str) -> str:
    """Excel 1 시트 출력. 경로 반환."""
    from datetime import datetime

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"grid_labels_{timestamp}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Labels"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["field", "label", "count", "file_count", "sample_sources"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=2):
        values = [
            r["field"],
            r["label"],
            r["count"],
            r["file_count"],
            "\n".join(r["sample_sources"]),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = border
            if col == 5:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {1: 28, 2: 30, 3: 8, 4: 10, 5: 55}
    for col_idx, w in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath
