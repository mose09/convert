import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)

SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3}
SEVERITY_COLOR = {
    "CRITICAL": "C00000",
    "HIGH": "FF6600",
    "MEDIUM": "FFC000",
    "LOW": "70AD47",
}


def save_review_markdown(review: dict, llm_reviews: list, output_dir: str) -> str:
    """Save SQL review report as Markdown."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"sql_review_{timestamp}.md")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# SQL Review Report\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # Summary
        f.write("## Summary\n\n")
        f.write(f"- Total statements: {review['total_statements']}\n")
        f.write(f"- Statements with issues: {review['statements_with_issues']}\n\n")

        f.write("### Severity Counts\n\n")
        f.write("| Severity | Count |\n")
        f.write("|----------|-------|\n")
        for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
            f.write(f"| {sev} | {review['severity_summary'].get(sev, 0)} |\n")
        f.write("\n")

        # By pattern
        f.write("## Issues by Pattern\n\n")
        sorted_patterns = sorted(
            review["by_pattern"].items(),
            key=lambda x: (SEVERITY_ORDER.get(x[1]["pattern"]["severity"], 99),
                           -len(x[1]["occurrences"]))
        )
        for p_id, data in sorted_patterns:
            p = data["pattern"]
            f.write(f"### [{p['severity']}] {p['name']} ({len(data['occurrences'])})\n\n")
            f.write(f"**Description**: {p['description']}\n\n")
            f.write(f"**Suggestion**: {p['suggestion']}\n\n")
            f.write("| Mapper | Statement ID | Type |\n")
            f.write("|--------|-------------|------|\n")
            for occ in data["occurrences"][:30]:
                f.write(f"| {occ['mapper']} | {occ['stmt_id']} | {occ['stmt_type']} |\n")
            if len(data["occurrences"]) > 30:
                f.write(f"| ... | +{len(data['occurrences']) - 30} more | |\n")
            f.write("\n")

        # LLM reviews
        if llm_reviews:
            f.write("---\n\n## LLM Review (Top Issues)\n\n")
            for r in llm_reviews:
                f.write(f"### {r['mapper']}#{r['stmt_id']} ({r['stmt_type']})\n\n")
                llm = r.get("llm_review", {})
                f.write(f"**Severity**: {llm.get('severity', '-')}\n\n")

                if llm.get("issues"):
                    f.write("**Issues**:\n")
                    for issue in llm["issues"]:
                        f.write(f"- {issue}\n")
                    f.write("\n")

                f.write("**Original SQL**:\n```sql\n")
                f.write(r["sql"])
                f.write("\n```\n\n")

                if llm.get("improved_sql"):
                    f.write("**Improved SQL**:\n```sql\n")
                    f.write(llm["improved_sql"])
                    f.write("\n```\n\n")

                if llm.get("explanation"):
                    f.write(f"**Explanation**: {llm['explanation']}\n\n")

                f.write("---\n\n")

    logger.info("SQL review markdown saved: %s", filepath)
    return filepath


def save_review_excel(review: dict, llm_reviews: list, output_dir: str) -> str:
    """Save SQL review report as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"sql_review_{timestamp}.xlsx")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    def _sev_fill(severity: str):
        color = SEVERITY_COLOR.get(severity, "DDDDDD")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _write_row(ws, row_num, values, severity=None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            if severity and col == 1:
                cell.fill = _sev_fill(severity)
                cell.font = Font(bold=True, color="FFFFFF")

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_header(ws_summary, ["Category", "Count"])
    ws_summary.cell(row=2, column=1, value="Total Statements").border = thin_border
    ws_summary.cell(row=2, column=2, value=review["total_statements"]).border = thin_border
    ws_summary.cell(row=3, column=1, value="Statements with Issues").border = thin_border
    ws_summary.cell(row=3, column=2, value=review["statements_with_issues"]).border = thin_border
    row = 5
    for sev in ("CRITICAL", "HIGH", "MEDIUM", "LOW"):
        _write_row(ws_summary, row, [sev, review["severity_summary"].get(sev, 0)], severity=sev)
        row += 1
    _auto_width(ws_summary)

    # Sheet 2: All issues
    ws_issues = wb.create_sheet("Issues")
    _write_header(ws_issues, ["Severity", "Pattern", "Mapper", "Statement ID", "Type", "Description"])

    sorted_findings = []
    for p_id, data in review["by_pattern"].items():
        for occ in data["occurrences"]:
            sorted_findings.append({
                "severity": data["pattern"]["severity"],
                "pattern_name": data["pattern"]["name"],
                "description": data["pattern"]["description"],
                **occ,
            })
    sorted_findings.sort(key=lambda x: (SEVERITY_ORDER.get(x["severity"], 99), x["mapper"]))

    for i, f in enumerate(sorted_findings, 2):
        _write_row(ws_issues, i, [
            f["severity"], f["pattern_name"], f["mapper"],
            f["stmt_id"], f["stmt_type"], f["description"],
        ], severity=f["severity"])

    _auto_width(ws_issues)

    # Sheet 3: Pattern summary
    ws_pattern = wb.create_sheet("Pattern Summary")
    _write_header(ws_pattern, ["Severity", "Pattern", "Count", "Description", "Suggestion"])

    sorted_patterns = sorted(
        review["by_pattern"].items(),
        key=lambda x: (SEVERITY_ORDER.get(x[1]["pattern"]["severity"], 99),
                       -len(x[1]["occurrences"]))
    )
    for i, (p_id, data) in enumerate(sorted_patterns, 2):
        p = data["pattern"]
        _write_row(ws_pattern, i, [
            p["severity"], p["name"], len(data["occurrences"]),
            p["description"], p["suggestion"],
        ], severity=p["severity"])

    _auto_width(ws_pattern)

    # Sheet 4: LLM Reviews (if any)
    if llm_reviews:
        ws_llm = wb.create_sheet("LLM Review")
        _write_header(ws_llm, ["Severity", "Mapper", "Statement ID", "Issues", "Original SQL", "Improved SQL", "Explanation"])
        for i, r in enumerate(llm_reviews, 2):
            llm = r.get("llm_review", {})
            issues = "\n".join(llm.get("issues", []))
            _write_row(ws_llm, i, [
                llm.get("severity", "-"),
                r["mapper"],
                r["stmt_id"],
                issues,
                r["sql"][:500],
                llm.get("improved_sql", "")[:500],
                llm.get("explanation", ""),
            ], severity=llm.get("severity"))
        _auto_width(ws_llm)

    wb.save(filepath)
    logger.info("SQL review excel saved: %s", filepath)
    return filepath
