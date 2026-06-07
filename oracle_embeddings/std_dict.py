"""표준사전 (단어사전 + 용어사전) Excel → SQLite 캐시 + 인메모리 인덱스.

AS-IS 스키마를 TO-BE 속성명으로 추천하기 위한 표준 데이터 저장소.

- **단어사전** (word): 논리명 / 물리명 / 물리의미(영문풀) / 표준여부 /
  속성분류어 / 동의어 / 설명 / 만료일자 / 출처구분
- **용어사전** (term): 논리명 / 물리명 / 구성정보 / 물리의미 / 도메인명 /
  데이터유형 / 길이 / 소수점 / 표준여부 / 개인정보구분 / 암호화여부 /
  설명 / 만료일자 / 출처구분

용어사전은 이미 `논리명 → 물리명 + 도메인 + 데이터유형` 정답표라 정확매칭(Tier1)
의 핵심이고, 단어사전은 단어조합/동의어 매칭(Tier2) 에 쓴다.

엑셀을 매 실행마다 재파싱하지 않도록 SQLite (`standard_dict.sqlite`) 에 캐시.
원본 mtime 이 바뀌면 자동 재빌드.
"""

import logging
import os
import re
import sqlite3
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────
# 헤더 자동 인식
# ─────────────────────────────────────────────────────────────────

# zero-width space/joiner/non-joiner/marks, BOM, soft-hyphen, word-joiner
_ZW_RE = re.compile("[\u200b\u200c\u200d\u200e\u200f\ufeff\u00ad\u2060]")


def _norm_header(h) -> str:
    """헤더 셀 정규화 — 유니코드 NFC + 숨은문자·괄호 부가설명·모든 공백·앞 번호 제거.

    한글이 NFD(조합형, macOS 등)로 저장돼 눈엔 같아도 바이트가 달라 매칭에
    실패하는 케이스를 NFC 정규화로 흡수한다. 예: ``1. 논리\n명(한글)`` → ``논리명``.
    """
    if h is None:
        return ""
    t = unicodedata.normalize("NFC", str(h))  # NFD 조합형 → NFC 완성형
    t = _ZW_RE.sub("", t)                      # zero-width / BOM 제거
    t = re.sub(r"\(.*?\)", "", t)              # 괄호 부가설명 제거
    t = re.sub(r"[*※★●○◦:：·]", "", t)         # 필수표시/장식 기호 제거
    t = re.sub(r"\s+", "", t)                  # 개행/탭/NBSP 포함 모든 공백 제거
    t = re.sub(r"^[0-9]+[.)\-_]?", "", t)      # 앞 번호 (1. / 2) / 3-) 제거
    return t.strip()


def _classify_header(h: str) -> str | None:
    """엑셀 헤더 셀 1개를 표준 필드 키로 분류 (없으면 None).

    정확매칭 우선 → 부분포함 폴백. 필수표시(``*``/``※``)·접두어(표준/단어/
    컬럼)·접미어가 붙은 ``표준단어논리명*`` 같은 표기도 잡는다. 단, ``물리의미``
    (eng) 가 ``물리명`` 보다 먼저 걸러지도록 순서 주의.
    """
    t = _norm_header(h)
    if not t:
        return None
    # eng 를 물리명/논리명 포함검사보다 먼저 (물리의미 가 物理 로 오분류되지 않게)
    if "물리의미" in t or "영문풀" in t or t in ("영문명", "영문"):
        return "eng"
    if t in ("논리명", "한글명", "논리명칭", "한글", "한글명칭", "속성명", "컬럼논리명") \
            or "논리명" in t or "한글명" in t:
        return "logical"
    if t in ("물리명", "컬럼명", "영문약어", "물리", "물리명칭", "컬럼물리명", "영문명약어") \
            or "물리명" in t or "컬럼명" in t or "영문약어" in t:
        return "physical"
    if "구성정보" in t or t == "구성":
        return "compose"
    if "도메인그룹" in t:
        return "domain_group"
    if "도메인" in t:
        return "domain"
    if "데이터유형" in t or "데이터타입" in t or "자료형" in t or "데이터타" in t:
        return "data_type"
    if t == "길이" or t.startswith("길이") or "데이터길이" in t:
        return "length"
    if "소수" in t:
        return "scale"
    if "분류어" in t:
        return "is_classifier"
    if "표준여부" in t or t == "표준":
        return "is_std"
    if "동의어" in t or "유의어" in t or "이음동의" in t:
        return "synonyms"
    if "개인정보" in t:
        return "privacy"
    if "암호화" in t:
        return "encrypt"
    if "설명" in t or "정의" in t or "비고" in t:
        return "desc"
    if "만료" in t:
        return "expire"
    if "출처" in t:
        return "source"
    return None


def _header_index(header_row) -> dict[str, int]:
    """헤더 행 → {필드키: 컬럼인덱스}. 같은 키 중복 시 첫 번째 우선."""
    idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _classify_header(cell)
        if key and key not in idx:
            idx[key] = i
    return idx


def _find_header_row(rows: list[tuple], required=("logical", "physical")) -> int:
    """헤더 행 위치 탐색. 상위 20행 중 `required` 키가 모두 잡히는 첫 행."""
    for r in range(min(20, len(rows))):
        idx = _header_index(rows[r])
        if all(k in idx for k in required):
            return r
    return 0


# ─────────────────────────────────────────────────────────────────
# 정규화 / 셀 추출 헬퍼
# ─────────────────────────────────────────────────────────────────

_PUNCT_RE = re.compile(r"[\s\-_./,()\[\]{}<>~!@#$%^&*+=|\\:;'\"?]+")


def norm_kor(s: str | None) -> str:
    """논리명/코멘트 매칭용 정규화 — 유니코드 NFC + 공백·기호 제거.

    사전과 스키마 코멘트의 한글 정규화 형식(NFC/NFD)이 달라도 매칭되도록 통일.
    """
    if not s:
        return ""
    t = unicodedata.normalize("NFC", str(s))
    return _PUNCT_RE.sub("", t).strip()


def _cell(row, idx: dict[str, int], key: str) -> str:
    i = idx.get(key)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    if v is None:
        return ""
    return str(v).strip()


def _is_yes(s: str) -> bool:
    return str(s).strip().upper() in ("Y", "YES", "TRUE", "1", "O", "○", "●", "예", "사용")


# 표준여부 부정 표기 (이외 값/빈값은 표준으로 간주 — 표기 다양성에 견고)
_STD_NO = {"N", "NO", "FALSE", "0", "X", "×", "비표준", "미표준", "아니오", "폐기", "삭제"}


def _is_std_value(s: str) -> bool:
    """표준여부 — 명시적 부정(N/X/비표준/×/...) 이 아니면 표준으로 간주.

    실제 사전마다 Y/N, O/X, ○/×, '표준'/'비표준' 등 표기가 제각각이라
    'Y 정확히 일치' 대신 '부정 표기 아니면 표준' 으로 견고하게 처리.
    """
    return str(s).strip().upper() not in _STD_NO


def _is_expired(expire: str) -> bool:
    """만료일자 가 오늘 이전이면 True (만료). 비어있으면 유효."""
    s = (expire or "").strip()
    if not s:
        return False
    digits = re.sub(r"[^0-9]", "", s)
    if len(digits) < 8:
        return False
    try:
        d = datetime.strptime(digits[:8], "%Y%m%d").date()
    except ValueError:
        return False
    return d < datetime.now().date()


def _split_synonyms(s: str) -> list[str]:
    if not s:
        return []
    parts = re.split(r"[,/;|·\n]+", s)
    return [p.strip() for p in parts if p.strip()]


def compose_data_type(data_type: str, length: str, scale: str) -> str:
    """데이터유형 + 길이 + 소수점 → ``VARCHAR2(50)`` / ``NUMBER(15,2)`` / ``DATE``."""
    dt = (data_type or "").strip().upper()
    if not dt:
        return ""
    ln = re.sub(r"[^0-9]", "", str(length or ""))
    sc = re.sub(r"[^0-9]", "", str(scale or ""))
    # 이미 길이가 타입에 포함돼 있으면 그대로
    if "(" in dt:
        return dt
    no_len = {"DATE", "TIMESTAMP", "CLOB", "BLOB", "LONG", "ROWID"}
    if dt in no_len or not ln:
        return dt
    if sc and sc != "0":
        return f"{dt}({ln},{sc})"
    return f"{dt}({ln})"


# ─────────────────────────────────────────────────────────────────
# Excel 로드
# ─────────────────────────────────────────────────────────────────

def _all_sheets(path: str) -> list[tuple[str, list[tuple]]]:
    """워크북의 모든 시트를 (이름, rows) 로 반환.

    `read_only=False` (일반 모드) 로 읽는다 — POI/사내도구/DRM 으로 생성된
    엑셀은 XML `<dimension>` 태그가 깨져(`A1:A1`) read_only 에서 1셀만 읽히는
    경우가 있어, dimension 에 의존하지 않는 일반 모드로 전체 셀을 읽는다.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise ImportError("openpyxl 가 필요합니다: pip install openpyxl") from e
    wb = load_workbook(path, data_only=True, read_only=False)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [row for row in ws.iter_rows(values_only=True)]
        out.append((name, rows))
    wb.close()
    return out


def _pick_sheet(path: str, sheet: str | None, required=("logical", "physical")):
    """데이터 시트 자동 선택.

    Returns ``(sheet_name, rows, header_idx, header_row)``. `required` 키가
    모두 잡히고 **행 수가 가장 많은** 시트를 고른다 (표지/거의-빈 시트가 앞에
    있어도 실제 데이터 시트 선택). `sheet` 지정 시 해당 시트만.
    """
    sheets = _all_sheets(path)
    if sheet:
        sheets = [(n, r) for (n, r) in sheets if n == sheet] or sheets
    primary = required[0]
    best_full = None      # 모든 required 충족 (rows수, name, rows, idx, hr)
    best_partial = None   # 첫 required 만 충족
    for name, rows in sheets:
        if not rows:
            continue
        hr = _find_header_row(rows, required)
        idx = _header_index(rows[hr])
        if all(k in idx for k in required):
            if best_full is None or len(rows) > best_full[0]:
                best_full = (len(rows), name, rows, idx, hr)
        elif primary in idx and best_partial is None:
            best_partial = (len(rows), name, rows, idx, hr)
    chosen = best_full or best_partial
    if chosen:
        _, name, rows, idx, hr = chosen
        return name, rows, idx, hr
    # 헤더 인식 실패 — 첫 비어있지 않은 시트 (진단용)
    for name, rows in sheets:
        if rows:
            hr = _find_header_row(rows, required)
            return name, rows, _header_index(rows[hr]), hr
    return "", [], {}, 0


def diagnose_xlsx(path: str) -> str:
    """0건일 때 자가진단 — 시트명 + 헤더행 셀별 repr·분류 결과.

    repr 로 출력해 숨은문자/None/실제 글자가 그대로 드러나게 한다.
    """
    lines = [f"  진단: {os.path.basename(path)}"]
    sheets = _all_sheets(path)
    lines.append(f"    시트({len(sheets)}): {[n for n, _ in sheets]}")
    for name, rows in sheets:
        if not rows:
            lines.append(f"    - [{name}] 빈 시트")
            continue
        nonnull = sum(1 for r in rows[:5] for c in r if c is not None)
        hr = _find_header_row(rows)
        idx = _header_index(rows[hr])
        ok = "logical" in idx and "physical" in idx
        mark = "✓논리+물리 인식" if ok else "✗논리/물리 미인식"
        lines.append(
            f"    - [{name}] {len(rows)}행, 헤더추정 {hr + 1}행: {mark} "
            f"(상위5행 비어있지않은셀 {nonnull}개)")
        if nonnull == 0:
            lines.append("        ⚠ 셀이 전부 None — DRM/보안 또는 캐시값 없음 의심")
            continue
        for ci, c in enumerate(rows[hr][:14]):
            if c is None:
                continue
            key = _classify_header(c)
            lines.append(f"        col{ci}: {c!r} → {key or '✗미분류'}")
    return "\n".join(lines)


def _load_word_rows(path: str, sheet: str | None) -> list[dict]:
    _, rows, idx, hr = _pick_sheet(path, sheet)
    if not rows or "logical" not in idx:
        return []
    out = []
    for row in rows[hr + 1:]:
        logical = _cell(row, idx, "logical")
        physical = _cell(row, idx, "physical")
        if not logical and not physical:
            continue
        out.append({
            "logical": logical,
            "physical": physical.upper(),
            "eng": _cell(row, idx, "eng"),
            "is_std": _is_std_value(_cell(row, idx, "is_std")) if "is_std" in idx else True,
            "is_classifier": _is_yes(_cell(row, idx, "is_classifier")),
            "synonyms": _cell(row, idx, "synonyms"),
            "desc": _cell(row, idx, "desc"),
            "expire": _cell(row, idx, "expire"),
            "source": _cell(row, idx, "source"),
        })
    return out


def _load_term_rows(path: str, sheet: str | None) -> list[dict]:
    _, rows, idx, hr = _pick_sheet(path, sheet)
    if not rows or "logical" not in idx:
        return []
    out = []
    for row in rows[hr + 1:]:
        logical = _cell(row, idx, "logical")
        physical = _cell(row, idx, "physical")
        if not logical and not physical:
            continue
        out.append({
            "logical": logical,
            "physical": physical.upper(),
            "compose": _cell(row, idx, "compose"),
            "eng": _cell(row, idx, "eng"),
            "domain": _cell(row, idx, "domain"),
            "data_type": _cell(row, idx, "data_type"),
            "length": _cell(row, idx, "length"),
            "scale": _cell(row, idx, "scale"),
            "is_std": _is_std_value(_cell(row, idx, "is_std")) if "is_std" in idx else True,
            "privacy": _cell(row, idx, "privacy"),
            "encrypt": _cell(row, idx, "encrypt"),
            "desc": _cell(row, idx, "desc"),
            "expire": _cell(row, idx, "expire"),
            "source": _cell(row, idx, "source"),
        })
    return out


def _load_domain_rows(path: str, sheet: str | None) -> list[dict]:
    """도메인사전 — 도메인그룹명/도메인명/데이터유형/길이/소수점/개인정보/암호화/...

    동일 도메인명이 그룹별로 여러 개 존재할 수 있어 중복을 그대로 보존한다.
    """
    _, rows, idx, hr = _pick_sheet(path, sheet, required=("domain", "data_type"))
    if not rows or "domain" not in idx:
        return []
    out = []
    for row in rows[hr + 1:]:
        name = _cell(row, idx, "domain")
        if not name:
            continue
        out.append({
            "group": _cell(row, idx, "domain_group"),
            "name": name,
            "data_type": _cell(row, idx, "data_type"),
            "length": _cell(row, idx, "length"),
            "scale": _cell(row, idx, "scale"),
            "privacy": _cell(row, idx, "privacy"),
            "encrypt": _cell(row, idx, "encrypt"),
            "desc": _cell(row, idx, "desc"),
            "expire": _cell(row, idx, "expire"),
            "source": _cell(row, idx, "source"),
        })
    return out


# ─────────────────────────────────────────────────────────────────
# SQLite 빌드 / 재빌드 판정
# ─────────────────────────────────────────────────────────────────

_WORD_COLS = ["logical", "physical", "eng", "is_std", "is_classifier",
              "synonyms", "desc", "expire", "source"]
_TERM_COLS = ["logical", "physical", "compose", "eng", "domain", "data_type",
              "length", "scale", "is_std", "privacy", "encrypt", "desc",
              "expire", "source"]


def _mtime(path: str | None) -> str:
    if path and os.path.exists(path):
        return str(int(os.path.getmtime(path)))
    return ""


def needs_rebuild(db_path: str, word_xlsx: str | None, term_xlsx: str | None,
                  domain_xlsx: str | None = None) -> bool:
    """SQLite 가 없거나 원본 Excel mtime 이 바뀌었으면 재빌드 필요."""
    if not os.path.exists(db_path):
        return True
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.cursor()
        cur.execute("SELECT key, value FROM meta")
        meta = dict(cur.fetchall())
        # domain 테이블이 없는 구버전 캐시면 재빌드
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='domain'")
        has_domain_tbl = cur.fetchone() is not None
        conn.close()
    except sqlite3.Error:
        return True
    if word_xlsx and meta.get("word_mtime", "") != _mtime(word_xlsx):
        return True
    if term_xlsx and meta.get("term_mtime", "") != _mtime(term_xlsx):
        return True
    if domain_xlsx and (not has_domain_tbl
                        or meta.get("domain_mtime", "") != _mtime(domain_xlsx)):
        return True
    return False


def build_std_dict(db_path: str, word_xlsx: str | None = None,
                   term_xlsx: str | None = None, word_sheet: str | None = None,
                   term_sheet: str | None = None, domain_xlsx: str | None = None,
                   domain_sheet: str | None = None) -> dict:
    """단어/용어/도메인 사전 Excel → SQLite 캐시 생성. 통계 dict 반환."""
    os.makedirs(os.path.dirname(os.path.abspath(db_path)), exist_ok=True)

    words = _load_word_rows(word_xlsx, word_sheet) if word_xlsx else []
    terms = _load_term_rows(term_xlsx, term_sheet) if term_xlsx else []
    domains = _load_domain_rows(domain_xlsx, domain_sheet) if domain_xlsx else []

    # 만료 항목 집계 (적재는 하되 표준 인덱스에서 제외하도록 플래그만)
    word_expired = sum(1 for w in words if _is_expired(w["expire"]))
    term_expired = sum(1 for t in terms if _is_expired(t["expire"]))
    domain_expired = sum(1 for x in domains if _is_expired(x["expire"]))

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS word")
    cur.execute("DROP TABLE IF EXISTS term")
    cur.execute("DROP TABLE IF EXISTS domain")
    cur.execute("DROP TABLE IF EXISTS meta")
    cur.execute(
        "CREATE TABLE word (logical TEXT, logical_norm TEXT, physical TEXT, "
        "eng TEXT, is_std INT, is_classifier INT, synonyms TEXT, desc TEXT, "
        "expired INT, source TEXT)"
    )
    cur.execute(
        "CREATE TABLE term (logical TEXT, logical_norm TEXT, physical TEXT, "
        "compose TEXT, eng TEXT, domain TEXT, data_type TEXT, length TEXT, "
        "scale TEXT, is_std INT, privacy TEXT, encrypt TEXT, desc TEXT, "
        "expired INT, source TEXT)"
    )
    cur.execute(
        "CREATE TABLE domain (grp TEXT, name TEXT, name_norm TEXT, "
        "data_type TEXT, length TEXT, scale TEXT, full_type TEXT, "
        "privacy TEXT, encrypt TEXT, desc TEXT, expired INT, source TEXT)"
    )
    cur.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")

    cur.executemany(
        "INSERT INTO word VALUES (?,?,?,?,?,?,?,?,?,?)",
        [(w["logical"], norm_kor(w["logical"]), w["physical"], w["eng"],
          int(w["is_std"]), int(w["is_classifier"]), w["synonyms"], w["desc"],
          int(_is_expired(w["expire"])), w["source"]) for w in words],
    )
    cur.executemany(
        "INSERT INTO term VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(t["logical"], norm_kor(t["logical"]), t["physical"], t["compose"],
          t["eng"], t["domain"], t["data_type"], t["length"], t["scale"],
          int(t["is_std"]), t["privacy"], t["encrypt"], t["desc"],
          int(_is_expired(t["expire"])), t["source"]) for t in terms],
    )
    cur.executemany(
        "INSERT INTO domain VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        [(x["group"], x["name"], norm_kor(x["name"]), x["data_type"],
          x["length"], x["scale"],
          compose_data_type(x["data_type"], x["length"], x["scale"]),
          x["privacy"], x["encrypt"], x["desc"],
          int(_is_expired(x["expire"])), x["source"]) for x in domains],
    )
    cur.execute("CREATE INDEX ix_word_norm ON word(logical_norm)")
    cur.execute("CREATE INDEX ix_word_phys ON word(physical)")
    cur.execute("CREATE INDEX ix_term_norm ON term(logical_norm)")
    cur.execute("CREATE INDEX ix_term_phys ON term(physical)")
    cur.execute("CREATE INDEX ix_domain_norm ON domain(name_norm)")

    meta = {
        "word_mtime": _mtime(word_xlsx),
        "term_mtime": _mtime(term_xlsx),
        "domain_mtime": _mtime(domain_xlsx),
        "word_xlsx": word_xlsx or "",
        "term_xlsx": term_xlsx or "",
        "domain_xlsx": domain_xlsx or "",
        "built_at": datetime.now().isoformat(timespec="seconds"),
    }
    cur.executemany("INSERT INTO meta VALUES (?,?)", list(meta.items()))
    conn.commit()
    conn.close()

    stats = {
        "words": len(words), "word_expired": word_expired,
        "terms": len(terms), "term_expired": term_expired,
        "domains": len(domains), "domain_expired": domain_expired,
    }
    logger.info("표준사전 빌드: 단어 %d (만료 %d) / 용어 %d (만료 %d) / "
                "도메인 %d (만료 %d) → %s",
                stats["words"], word_expired, stats["terms"], term_expired,
                stats["domains"], domain_expired, db_path)
    return stats


# ─────────────────────────────────────────────────────────────────
# 인메모리 인덱스 로드
# ─────────────────────────────────────────────────────────────────

@dataclass
class WordRow:
    logical: str
    physical: str
    eng: str
    is_classifier: bool
    synonyms: list[str]
    desc: str


@dataclass
class TermRow:
    logical: str
    physical: str
    domain: str
    data_type: str  # 길이·소수점 합성된 최종 타입
    eng: str
    desc: str
    privacy: str
    encrypt: str


@dataclass
class DomainEntry:
    group: str
    name: str
    data_type: str   # 길이·소수점 합성된 최종 타입 (full_type)
    privacy: str
    encrypt: str
    desc: str


@dataclass
class StdDict:
    # Tier1: 용어사전 정확매칭
    term_by_logical: dict[str, TermRow] = field(default_factory=dict)  # norm논리명 → term
    term_by_physical: dict[str, TermRow] = field(default_factory=dict)  # 물리명 → term
    # Tier2: 단어사전 조합
    word_by_logical: dict[str, WordRow] = field(default_factory=dict)  # 논리명 → word
    syn_to_logical: dict[str, str] = field(default_factory=dict)  # norm(동의어/논리명) → 논리명
    word_keys_sorted: list[str] = field(default_factory=list)  # norm 논리명/동의어 (길이 desc)
    abbr_to_logical: dict[str, str] = field(default_factory=dict)  # 물리명 → 논리명
    logical_to_abbr: dict[str, str] = field(default_factory=dict)  # 논리명 → 물리명 (표준여부 무관, 전 단어)
    classifier_type: dict[str, tuple[str, str]] = field(default_factory=dict)  # 분류어논리명 → (domain, data_type)
    # 도메인사전: 동일 도메인명 다중 엔트리 보존
    domain_by_name: dict[str, list[DomainEntry]] = field(default_factory=dict)  # norm도메인명 → [entry...]
    domain_type: dict[str, str] = field(default_factory=dict)  # norm도메인명 → 대표 데이터유형 (최빈/유일)
    counts: dict = field(default_factory=dict)

    def has_terms(self) -> bool:
        return bool(self.term_by_logical or self.term_by_physical)

    def has_words(self) -> bool:
        return bool(self.word_by_logical)

    def has_domains(self) -> bool:
        return bool(self.domain_by_name)

    def resolve_domain_type(self, domain_name: str) -> tuple[str, bool]:
        """도메인명 → (대표 데이터유형, 단일여부). 다중이면 최빈값+단일=False."""
        entries = self.domain_by_name.get(norm_kor(domain_name), [])
        if not entries:
            return "", True
        types = [e.data_type for e in entries if e.data_type]
        if not types:
            return "", True
        uniq = set(types)
        if len(uniq) == 1:
            return types[0], True
        return Counter(types).most_common(1)[0][0], False


def load_std_dict(db_path: str) -> StdDict:
    """SQLite → 매칭용 인메모리 인덱스. 만료/비표준 항목은 정확매칭에서 제외."""
    sd = StdDict()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 용어사전 — 표준이고 만료 안 된 것만 정확매칭 인덱스에 적재
    for r in cur.execute("SELECT * FROM term"):
        if r["expired"]:
            continue
        tr = TermRow(
            logical=r["logical"], physical=r["physical"], domain=r["domain"],
            data_type=compose_data_type(r["data_type"], r["length"], r["scale"]),
            eng=r["eng"], desc=r["desc"], privacy=r["privacy"], encrypt=r["encrypt"],
        )
        nl = r["logical_norm"]
        if nl and nl not in sd.term_by_logical:
            sd.term_by_logical[nl] = tr
        if r["physical"] and r["physical"] not in sd.term_by_physical:
            sd.term_by_physical[r["physical"]] = tr

    # 분류어별 대표 (도메인, 데이터유형) — 용어사전 물리명 마지막 토큰 기준 최빈값
    classifier_samples: dict[str, Counter] = defaultdict(Counter)

    # 단어사전
    word_keys: set[str] = set()
    _abbr_std: dict[str, str] = {}   # 논리명 → 물리명 (표준 행)
    _abbr_any: dict[str, str] = {}   # 논리명 → 물리명 (표준여부 무관)
    for r in cur.execute("SELECT * FROM word"):
        if r["expired"]:
            continue
        syns = _split_synonyms(r["synonyms"])
        wr = WordRow(
            logical=r["logical"], physical=r["physical"], eng=r["eng"],
            is_classifier=bool(r["is_classifier"]), synonyms=syns, desc=r["desc"],
        )
        if r["is_std"] and r["logical"] and r["logical"] not in sd.word_by_logical:
            sd.word_by_logical[r["logical"]] = wr
        if r["physical"] and r["physical"] not in sd.abbr_to_logical:
            sd.abbr_to_logical[r["physical"]] = r["logical"]
        # 논리명 → 물리명: 표준 행 우선, 없으면 비표준으로 폴백 (시설→FACI(표준)
        # 가 FAC(비표준)에 가려지지 않게 + 확인처럼 비표준만 있으면 그거라도 사용)
        if r["logical"] and r["physical"]:
            if r["logical"] not in _abbr_any:
                _abbr_any[r["logical"]] = r["physical"]
            if r["is_std"] and r["logical"] not in _abbr_std:
                _abbr_std[r["logical"]] = r["physical"]
        # 동의어/논리명 → 논리명 매핑 (표준 단어로 귀결)
        for key in [r["logical"], *syns]:
            nk = norm_kor(key)
            if nk and nk not in sd.syn_to_logical:
                sd.syn_to_logical[nk] = r["logical"]
                word_keys.add(nk)

    # 표준 행 우선으로 논리명→물리명 확정 (비표준은 표준 없을 때만 폴백)
    sd.logical_to_abbr = {**_abbr_any, **_abbr_std}

    # 분류어 데이터유형 추론: 용어사전 물리명 끝 토큰 == 분류어 물리명 인 경우 집계
    classifier_phys = {
        w.physical: w.logical for w in sd.word_by_logical.values()
        if w.is_classifier and w.physical
    }
    for tr in sd.term_by_physical.values():
        if not tr.data_type:
            continue
        last = tr.physical.split("_")[-1] if tr.physical else ""
        logical = classifier_phys.get(last)
        if logical:
            classifier_samples[logical][(tr.domain, tr.data_type)] += 1
    for logical, ctr in classifier_samples.items():
        (domain, dtype), _ = ctr.most_common(1)[0]
        sd.classifier_type[logical] = (domain, dtype)

    # 도메인사전 (동일 도메인명 다중 엔트리 보존)
    try:
        for r in cur.execute("SELECT * FROM domain"):
            if r["expired"]:
                continue
            sd.domain_by_name.setdefault(r["name_norm"], []).append(DomainEntry(
                group=r["grp"], name=r["name"], data_type=r["full_type"],
                privacy=r["privacy"], encrypt=r["encrypt"], desc=r["desc"]))
    except sqlite3.OperationalError:
        pass  # 구버전 캐시(도메인 테이블 없음)
    for nk, entries in sd.domain_by_name.items():
        types = [e.data_type for e in entries if e.data_type]
        if types:
            sd.domain_type[nk] = Counter(types).most_common(1)[0][0]

    sd.word_keys_sorted = sorted(word_keys, key=len, reverse=True)
    sd.counts = {
        "terms": len(sd.term_by_logical),
        "words": len(sd.word_by_logical),
        "synonyms": len(sd.syn_to_logical),
        "classifiers": len(sd.classifier_type),
        "domains": len(sd.domain_by_name),
        "domain_rows": sum(len(v) for v in sd.domain_by_name.values()),
    }
    conn.close()
    logger.info("표준사전 로드: 용어 %d / 단어 %d / 동의어 %d / 분류어타입 %d / "
                "도메인 %d(행 %d)",
                sd.counts["terms"], sd.counts["words"], sd.counts["synonyms"],
                sd.counts["classifiers"], sd.counts["domains"],
                sd.counts["domain_rows"])
    return sd


def ensure_std_dict(db_path: str, word_xlsx: str | None, term_xlsx: str | None,
                    force: bool = False, word_sheet: str | None = None,
                    term_sheet: str | None = None, domain_xlsx: str | None = None,
                    domain_sheet: str | None = None) -> StdDict:
    """필요 시 빌드 후 로드. (mtime 변경/강제 시 재빌드)"""
    if force or needs_rebuild(db_path, word_xlsx, term_xlsx, domain_xlsx):
        if not (word_xlsx or term_xlsx or domain_xlsx):
            raise FileNotFoundError(
                f"표준사전 SQLite 가 없고 사전 인자(--word-dict/--term-dict/"
                f"--domain-dict) 도 없습니다: {db_path}"
            )
        build_std_dict(db_path, word_xlsx, term_xlsx, word_sheet, term_sheet,
                       domain_xlsx, domain_sheet)
    return load_std_dict(db_path)
