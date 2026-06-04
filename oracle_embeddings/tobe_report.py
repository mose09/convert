"""TO-BE 속성명 추천 리포트 — Markdown + Excel 산출."""

import logging
import os
from collections import Counter
from datetime import datetime

from .tobe_recommender import ColumnRec, RecommendStats

logger = logging.getLogger(__name__)

_HEADERS = ["테이블", "AS-IS 컬럼", "코멘트", "기준", "분해 단어",
            "TO-BE 물리명", "도메인", "데이터유형", "신뢰도", "단계", "비고"]


def _basis_kor(b: str) -> str:
    return {"comment": "코멘트", "column": "물리명"}.get(b, b)


def _tokens_str(rec: ColumnRec) -> str:
    out = []
    for t in rec.tokens:
        if t.matched:
            out.append(f"{t.frag}>{t.abbr}" if t.abbr else t.frag)
        else:
            out.append(f"[{t.frag}?]")
    return " + ".join(out)


def _unmatched_counter(recs: list[ColumnRec]) -> Counter:
    c: Counter = Counter()
    for r in recs:
        for f in r.unmatched_frags:
            c[f] += 1
    return c


def _row(rec: ColumnRec) -> list:
    return [
        rec.table, rec.column, rec.comment or "", _basis_kor(rec.basis),
        _tokens_str(rec), rec.tobe_name, rec.domain, rec.data_type,
        rec.confidence, rec.tier, rec.note,
    ]


def save_recommend_markdown(recs: list[ColumnRec], stats: RecommendStats,
                            output_dir: str) -> str:
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"tobe_recommend_{ts}.md")

    lines = [
        "# TO-BE 속성명 추천 리포트", "",
        f"- 대상 컬럼: {stats.total}",
        f"- 정확매칭(용어): {stats.tier1}",
        f"- 이미표준: {stats.already_std}",
        f"- 단어조합: {stats.tier2}",
        f"- RAG/LLM 보조: {stats.tier_llm}",
        f"- 미매칭: {stats.unmatched}",
        f"- 저신뢰(<0.7): {stats.low_conf} / 길이초과: {stats.too_long}",
        f"- 생성: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", "",
        "## 추천 결과", "",
        "| " + " | ".join(_HEADERS) + " |",
        "| " + " | ".join(["---"] * len(_HEADERS)) + " |",
    ]
    for r in recs:
        cells = [str(c).replace("|", "\\|") for c in _row(r)]
        lines.append("| " + " | ".join(cells) + " |")

    unmatched = _unmatched_counter(recs)
    if unmatched:
        lines += ["", f"## 미매칭 단어 ({len(unmatched)})", "",
                  "| 단어 | 빈도 |", "| --- | --- |"]
        for word, cnt in unmatched.most_common():
            lines.append(f"| {word} | {cnt} |")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    logger.info("추천 Markdown 저장: %s", path)
    return path


def _style_header(ws, ncols: int):
    from openpyxl.styles import Font, PatternFill
    fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font
        cell.fill = fill


def _autowidth(ws):
    for col in ws.columns:
        first = col[0]
        if not hasattr(first, "column_letter"):
            continue
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[first.column_letter].width = min(width + 4, 60)


def save_recommend_excel(recs: list[ColumnRec], stats: RecommendStats,
                         output_dir: str) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"tobe_recommend_{ts}.xlsx")

    wb = Workbook()

    # 시트 1: 추천결과
    ws = wb.active
    ws.title = "추천결과"
    ws.append(_HEADERS)
    low_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    miss_fill = PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid")
    for r in recs:
        ws.append(_row(r))
        row_cells = ws[ws.max_row]
        if r.tier == "미매칭":
            for c in row_cells:
                c.fill = miss_fill
        elif 0 < r.confidence < 0.7:
            for c in row_cells:
                c.fill = low_fill
    _style_header(ws, len(_HEADERS))
    ws.freeze_panes = "A2"
    _autowidth(ws)

    # 시트 2: 미매칭 단어
    ws2 = wb.create_sheet("미매칭단어")
    ws2.append(["단어", "빈도"])
    for word, cnt in _unmatched_counter(recs).most_common():
        ws2.append([word, cnt])
    _style_header(ws2, 2)
    _autowidth(ws2)

    # 시트 3: 요약
    ws3 = wb.create_sheet("요약")
    ws3.append(["항목", "값"])
    summary = [
        ("대상 컬럼", stats.total),
        ("정확매칭(용어)", stats.tier1),
        ("이미표준", stats.already_std),
        ("단어조합", stats.tier2),
        ("RAG/LLM 보조", stats.tier_llm),
        ("미매칭", stats.unmatched),
        ("저신뢰(<0.7)", stats.low_conf),
        ("길이초과", stats.too_long),
        ("생성", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for k, v in summary:
        ws3.append([k, v])
    _style_header(ws3, 2)
    ws3["A1"].font = Font(bold=True, color="FFFFFF")
    _autowidth(ws3)

    wb.save(path)
    logger.info("추천 Excel 저장: %s", path)
    return path
