"""Markdown + Excel output for the AS-IS legacy analyzer.

Consumes the result dict produced by ``legacy_analyzer.analyze_legacy`` and
emits everything under a dedicated ``legacy_analysis/`` subfolder of the
configured output directory so that AS-IS artifacts do not pollute the
shared ``output/`` root used by other commands:

* ``<output>/legacy_analysis/as_is_analysis_<backend>_<ts>.md``
  sections: header, summary, menu hierarchy, program detail table,
  unmatched controllers, orphan menus.
* ``<output>/legacy_analysis/as_is_analysis_<backend>_<ts>.xlsx``
  7 sheets: Summary, Programs, Menu Hierarchy, Unmatched Controllers,
  Orphan Menu Entries, RFC Calls, Tables Cross-Reference.

``<backend>`` is the basename of the ``backend_dir`` argument (e.g.
``/path/to/backend/gipms-api-common`` → ``gipms-api-common``) so that
running the analyzer on multiple services of a monorepo produces
distinct, easily-identifiable output files.
"""

import logging
import os
import re
from datetime import datetime

logger = logging.getLogger(__name__)

LEGACY_SUBDIR = "legacy_analysis"


# Matches the trailing ``(CRUD)`` CRUD letters that ``_format_table_crud``
# appends to each table name in the Programs sheet. Used in
# Tables Cross-Reference aggregation so the bare table name is used as
# the dict key — otherwise ``CMN_BTN_ROLE(R)`` and ``CMN_BTN_ROLE(CR)``
# would look like two different tables.
_CRUD_SUFFIX_RE = re.compile(r"\s*\(\s*[CRUD]+\s*\)\s*$")


def _bare_table_name(cell: str) -> str:
    """Return the table name with the trailing ``(CRUD)`` stripped."""
    return _CRUD_SUFFIX_RE.sub("", cell).strip()


def _legacy_subdir_with_date() -> str:
    """Return ``legacy_analysis/<YYYYMMDD>`` per shared output convention.

    공통 출력 규약 (CLAUDE.md "출력 / 입력 경로 규약" 참고): 영역 폴더
    하위에 일자 폴더를 둔다. ``LEGACY_SUBDIR`` 는 하위 호환을 위해 남겨둔다.
    """
    from datetime import datetime as _dt
    return os.path.join(LEGACY_SUBDIR, _dt.now().strftime("%Y%m%d"))


def _backend_slug(backend_dir: str) -> str:
    """Derive a filename-safe slug from the backend directory path.

    Only the leaf directory name is used so running on
    ``/home/me/projects/monorepo/backend/gipms-api-common`` produces
    ``gipms-api-common``. Characters that could cause issues on
    Windows / macOS filesystems are replaced with ``_``.
    """
    if not backend_dir:
        return ""
    base = os.path.basename(os.path.normpath(backend_dir))
    if not base:
        return ""
    # Allow alphanumerics plus a small set of safe punctuation.
    slug = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    return slug.strip("_.")


def _legacy_output_dir(output_dir: str) -> str:
    """Return ``<output>/legacy_analysis/<YYYYMMDD>`` and create it.

    All analyze-legacy artifacts (Markdown + Excel) are written under
    the area subfolder + 일자 폴더 (공통 출력 규약). ``.biz_cache`` 는
    ``legacy_biz_extractor._cache_dir`` 가 영역 루트 (`output/
    legacy_analysis/.biz_cache`) 에 직접 만들므로 일자 폴더 영향 없음.
    """
    target = os.path.join(output_dir or ".", _legacy_subdir_with_date())
    os.makedirs(target, exist_ok=True)
    return target


def _build_filename(output_dir: str, result: dict, ts: str, ext: str) -> str:
    """Return ``<output>/legacy_analysis/as_is_analysis_<slug>_<ts>.<ext>``."""
    slug = _backend_slug(result.get("backend_dir", ""))
    prefix = f"as_is_analysis_{slug}_" if slug else "as_is_analysis_"
    return os.path.join(_legacy_output_dir(output_dir), f"{prefix}{ts}.{ext}")


def _build_diagram_folder(output_dir: str, result: dict, ts: str) -> str:
    """리포트 파일명과 동일한 이름의 폴더 경로를 반환 + 생성.

    예: ``<output>/legacy_analysis/as_is_analysis_myapp_20260424_123456/``

    이 폴더 안에 endpoint 별 ``.md`` (Mermaid 코드블럭 포함) 가 건별
    저장됨. 리포트 파일 (md / xlsx) 과 폴더가 같은 prefix 라 사용자가
    쉽게 짝지어 찾을 수 있음.
    """
    slug = _backend_slug(result.get("backend_dir", ""))
    prefix = f"as_is_analysis_{slug}_" if slug else "as_is_analysis_"
    folder = os.path.join(_legacy_output_dir(output_dir), f"{prefix}{ts}")
    os.makedirs(folder, exist_ok=True)
    return folder


def _slugify_for_filename(text: str, fallback: str = "endpoint") -> str:
    """URL / 프로그램명을 파일명으로 안전한 slug 로 변환.

    Windows/macOS/Linux 공통 금지 문자 (``<>:"/\\|?*``) + 공백/슬래시 를
    ``_`` 로. 길이 80 자 상한.
    """
    if not text:
        return fallback
    slug = re.sub(r'[<>:"/\\|?*\s]+', "_", text).strip("_.")
    slug = re.sub(r"_+", "_", slug)
    if not slug:
        return fallback
    return slug[:80]


def save_sequence_diagrams_folder(result: dict, output_dir: str,
                                   group_by: str = "main_menu") -> str:
    """Mermaid ``.md`` 파일을 폴더에 저장 — group_by 기준으로 묶음.

    ``group_by`` 옵션:
      - ``"main_menu"`` (default): 같은 main_menu 의 모든 endpoint 를 한
        파일에 묶음. 업무 단위 (대분류) 별 시퀀스 묶음.
      - ``"menu_path"``: ``main_menu/sub_menu/tab`` 까지 같은 row 끼리
        묶음 (더 세분화).
      - ``"sub_menu"``: sub_menu 단위.
      - ``"controller_class"``: Java Controller 클래스 단위.
      - ``"none"``: endpoint 별 한 파일씩 (legacy 동작).

    파일명: ``<idx>_<group_slug>.md``. 같은 그룹의 endpoint 들은 한 파일
    안에 ``## 1. <program>``, ``## 2. <program>`` ... 순으로 나열.

    sequence_diagram 이 있는 row 가 하나도 없으면 폴더 자체를 생성하지
    않고 빈 문자열 반환 (회귀 없음).
    """
    rows = result.get("rows") or []
    diagram_rows = [r for r in rows if (r.get("sequence_diagram") or "").strip()]
    if not diagram_rows:
        return ""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = _build_diagram_folder(output_dir, result, ts)

    if group_by == "none":
        # legacy: endpoint 별 한 파일씩
        seen_names: dict[str, int] = {}
        for idx, r in enumerate(diagram_rows, 1):
            http = (r.get("http_method") or "").upper() or "ANY"
            program = r.get("program_name") or ""
            url = r.get("url") or ""
            slug_source = program or url or f"endpoint_{idx}"
            slug = _slugify_for_filename(slug_source, fallback=f"endpoint_{idx}")
            base = f"{idx:03d}_{http}_{slug}"
            fname = base
            if base in seen_names:
                seen_names[base] += 1
                fname = f"{base}_{seen_names[base]}"
            else:
                seen_names[base] = 1
            filepath = os.path.join(folder, f"{fname}.md")
            _write_single_diagram_md(filepath, r)
        logger.info("Sequence diagrams folder saved: %s (%d files, per-endpoint)",
                    folder, len(diagram_rows))
        print(f"  Sequence diagrams: {len(diagram_rows)} files (per-endpoint) → {folder}")
        return folder

    # grouped 모드: row 들을 group_by field 로 묶음
    groups: dict[str, list[dict]] = {}
    for r in diagram_rows:
        key = _resolve_group_key(r, group_by)
        groups.setdefault(key, []).append(r)
    # 그룹 키 정렬 — 미분류 ("_") 는 항상 마지막으로
    sorted_keys = sorted(groups.keys(),
                         key=lambda k: (k == "_미분류_", k))
    for idx, key in enumerate(sorted_keys, 1):
        slug = _slugify_for_filename(key, fallback=f"group_{idx}")
        fname = f"{idx:03d}_{slug}.md"
        filepath = os.path.join(folder, fname)
        _write_grouped_diagram_md(filepath, key, groups[key])
    logger.info("Sequence diagrams folder saved: %s (%d files, grouped by %s)",
                folder, len(groups), group_by)
    print(f"  Sequence diagrams: {len(groups)} files "
          f"(grouped by {group_by}, {len(diagram_rows)} endpoints) → {folder}")
    return folder


def _resolve_group_key(row: dict, group_by: str) -> str:
    """row 에서 group_by field 의 값을 추출. 비어있으면 ``"_미분류_"``."""
    if group_by == "menu_path":
        parts = [row.get("main_menu", ""), row.get("sub_menu", ""),
                 row.get("tab", "")]
        joined = " / ".join(p for p in parts if p)
        return joined or "_미분류_"
    val = (row.get(group_by) or "").strip()
    return val or "_미분류_"


def _write_grouped_diagram_md(filepath: str, group_name: str,
                               rows: list[dict]) -> None:
    """그룹 .md 작성 — 같은 group_by 의 모든 endpoint sequence 를 한 파일에."""
    lines: list[str] = []
    lines.append(f"# {group_name}")
    lines.append("")
    lines.append(f"_총 {len(rows)} endpoint_")
    lines.append("")
    for idx, r in enumerate(rows, 1):
        program = r.get("program_name") or ""
        http = r.get("http_method") or ""
        url = r.get("url") or ""
        title_bits = []
        if program:
            title_bits.append(program)
        if http or url:
            title_bits.append(f"{http} {url}".strip())
        title = " — ".join(b for b in title_bits if b) or f"Endpoint {idx}"
        lines.append(f"## {idx}. {title}")
        lines.append("")
        meta_pairs = [
            ("Sub menu", r.get("sub_menu", "")),
            ("Tab", r.get("tab", "")),
            ("Program", program),
            ("HTTP", http),
            ("URL", url),
            ("Controller", r.get("controller_class", "")),
            ("File", r.get("file_name", "")),
            ("Service", r.get("service", "")),
            ("Service method", r.get("service_method", "")),
            ("XML", r.get("query_xml", "")),
            ("XML method", r.get("sql_ids", "")),
            ("Tables", r.get("related_tables", "")),
            ("Columns", r.get("related_columns", "")),
            ("Procedures", r.get("procedures", "")),
            ("RFC", r.get("rfc", "")),
        ]
        for label, value in meta_pairs:
            if not value:
                continue
            inline = str(value).replace("\n", "<br>")
            lines.append(f"- **{label}**: {inline}")
        lines.append("")
        lines.append("```mermaid")
        lines.append(r.get("sequence_diagram", "") or "")
        lines.append("```")
        lines.append("")
    with open(filepath, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def _write_single_diagram_md(filepath: str, row: dict) -> None:
    """한 endpoint 의 정보 + Mermaid 코드블럭을 ``.md`` 로 저장."""
    title_bits = []
    program = row.get("program_name") or ""
    http = row.get("http_method") or ""
    url = row.get("url") or ""
    if program:
        title_bits.append(program)
    if http or url:
        title_bits.append(f"{http} {url}".strip())
    title = " — ".join(b for b in title_bits if b) or "Endpoint"

    lines: list[str] = []
    lines.append(f"# {title}")
    lines.append("")
    lines.append("## Endpoint")
    lines.append("")
    meta_pairs = [
        ("Main menu", row.get("main_menu", "")),
        ("Sub menu", row.get("sub_menu", "")),
        ("Tab", row.get("tab", "")),
        ("Program", program),
        ("HTTP", http),
        ("URL", url),
        ("Controller", row.get("controller_class", "")),
        ("File", row.get("file_name", "")),
        ("Service", row.get("service", "")),
        ("Service method", row.get("service_method", "")),
        ("XML", row.get("query_xml", "")),
        ("XML method", row.get("sql_ids", "")),
        ("Tables", row.get("related_tables", "")),
        ("Columns", row.get("related_columns", "")),
        ("Procedures", row.get("procedures", "")),
        ("RFC", row.get("rfc", "")),
    ]
    for label, value in meta_pairs:
        if not value:
            continue
        # 여러 줄 (,\n / ;\n) 인 항목은 `<br>` 으로 합쳐서 한 셀처럼
        inline = str(value).replace("\n", "<br>")
        lines.append(f"- **{label}**: {inline}")
    lines.append("")
    lines.append("## Sequence Diagram")
    lines.append("")
    lines.append("```mermaid")
    lines.append(row.get("sequence_diagram", "") or "")
    lines.append("```")
    lines.append("")
    with open(filepath, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines))


def _md_escape(text) -> str:
    """Escape pipes/newlines so Markdown tables don't break."""
    if text is None:
        return ""
    return str(text).replace("|", "\\|").replace("\n", "<br>").replace("\r", "")


def _group_by_menu(rows: list[dict]) -> dict:
    """Group rows by (main, sub, tab) for the menu hierarchy section."""
    groups = {}
    for r in rows:
        if not r.get("matched"):
            continue
        key = (r.get("main_menu", ""), r.get("sub_menu", ""), r.get("tab", ""))
        groups.setdefault(key, []).append(r)
    return groups


def save_legacy_markdown(result: dict, output_dir: str, menu_only: bool = False) -> str:
    """Render the analysis result as a Markdown document."""
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = _build_filename(output_dir, result, ts, "md")

    # result["rows"] is already menu-ordered (matched rows + menu-only
    # placeholders for un-matched menu entries). Un-matched endpoints
    # are in result["unmatched_controllers"]. No more per-report split.
    rows = result.get("rows", [])
    unmatched = result.get("unmatched_controllers", [])
    orphans = result.get("orphan_menus", [])
    stats = result.get("stats", {})

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# AS-IS Legacy Source Analysis\n\n")
        f.write(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"- Backend dir: `{result.get('backend_dir', '')}`\n")
        framework = result.get("backend_framework", "unknown")
        f.write(f"- Backend framework: **{framework}**\n")
        if result.get("frontend_dir"):
            f.write(f"- Frontend dir: `{result.get('frontend_dir', '')}`\n")
        f.write("\n")

        f.write("## Summary\n\n")
        f.write("| Category | Value |\n|---|---|\n")
        f.write(f"| Backend framework | {stats.get('backend_framework', 'unknown')} |\n")
        for label, key in [
            ("Controllers scanned", "controllers"),
            ("Services scanned", "services"),
            ("Java Mapper classes", "mappers"),
            ("MyBatis XML files", "mapper_xml_files"),
            ("MyBatis XML namespaces", "mapper_xml_namespaces"),
            ("Endpoints total", "endpoints"),
            ("Matched to menu", "matched"),
            ("Unmatched controllers", "unmatched"),
            ("Orphan menu entries", "orphan_menus"),
            ("Endpoints with React file", "with_react"),
            ("Endpoints with RFC", "with_rfc"),
        ]:
            f.write(f"| {label} | {stats.get(key, 0)} |\n")
        f.write("\n")

        # Menu hierarchy
        groups = _group_by_menu(rows)
        if groups:
            f.write("## Programs by Menu Hierarchy\n\n")
            last_main = None
            last_sub = None
            for (main, sub, tab), entries in sorted(groups.items()):
                if main != last_main:
                    f.write(f"### {main or '(no main)'}\n\n")
                    last_main = main
                    last_sub = None
                if sub != last_sub:
                    f.write(f"#### {sub or '(no sub)'}\n\n")
                    last_sub = sub
                if tab:
                    f.write(f"- **{tab}**\n")
                for e in entries:
                    f.write(f"  - `{e['http_method']}` `{e['url']}` → {e['program_name']} "
                            f"({e['controller_class']})\n")
            f.write("\n")

        # Program detail table — column layout depends on menu presence
        cols = _SINGLE_COLUMNS_WITH_MENU if _has_menu_data(rows) else _SINGLE_COLUMNS_NO_MENU
        f.write("## Program Detail\n\n")
        f.write("| " + " | ".join(label for label, _ in cols) + " |\n")
        f.write("|" + "|".join(["---"] * len(cols)) + "|\n")
        for idx, r in enumerate(rows, 1):
            f.write(
                "| " + " | ".join(
                    str(idx) if key == "__row_no__"
                    else _md_escape(r.get(key, ""))
                    for _, key in cols
                ) + " |\n"
            )
        f.write("\n")

        # Mermaid Sequence Diagrams (Phase A) — row 에 sequence_diagram
        # 있을 때만 섹션 생성. GitHub / VSCode 는 ```mermaid 코드블럭을
        # 자동 렌더. 그 외 환경은 Mermaid Live (https://mermaid.live) 에
        # 복붙해서 확인.
        diagram_rows = [r for r in rows if (r.get("sequence_diagram") or "").strip()]
        if diagram_rows:
            f.write(f"## Sequence Diagrams ({len(diagram_rows)})\n\n")
            f.write("endpoint 당 controller → service → XML → DB / RFC 호출 체인 "
                    "(source offset 순서, LLM 없이 static 추출).\n\n")
            for r in diagram_rows:
                title = (f"{r.get('http_method','')} {r.get('url','')}".strip()
                         or r.get("program_name", "") or "endpoint")
                f.write(f"### {title}\n\n")
                f.write("```mermaid\n")
                f.write(r["sequence_diagram"])
                f.write("\n```\n\n")

        if unmatched:
            f.write(f"## Unmatched Controllers ({len(unmatched)})\n\n")
            f.write("메뉴 테이블에 매칭되지 않은 컨트롤러 엔드포인트. 내부 API 이거나 메뉴 누락일 수 있습니다.\n\n")
            f.write("| HTTP | URL | Controller | Method | File |\n|---|---|---|---|---|\n")
            for u in unmatched:
                f.write(
                    f"| {u['http_method']} | {_md_escape(u['url'])} "
                    f"| {_md_escape(u['controller_class'])} "
                    f"| {_md_escape(u.get('program_name', ''))} "
                    f"| {_md_escape(u.get('file_name', ''))} |\n"
                )
            f.write("\n")

        if orphans:
            f.write(f"## Orphan Menu Entries ({len(orphans)})\n\n")
            f.write("메뉴 테이블에는 있는데 대응하는 컨트롤러를 찾지 못한 URL. 미구현 or 삭제된 기능.\n\n")
            f.write("| Program ID | Main | Sub | Tab | Program | URL |\n|---|---|---|---|---|---|\n")
            for o in orphans:
                f.write(
                    f"| {_md_escape(o.get('program_id', ''))} "
                    f"| {_md_escape(o.get('main_menu', ''))} "
                    f"| {_md_escape(o.get('sub_menu', ''))} "
                    f"| {_md_escape(o.get('tab', ''))} "
                    f"| {_md_escape(o.get('program_name', ''))} "
                    f"| {_md_escape(o.get('url', ''))} |\n"
                )
            f.write("\n")

    logger.info("Legacy markdown saved: %s", filepath)
    return filepath


def _write_biz_logic_sheet(wb, biz_map: dict, rows: list[dict]) -> None:
    """Common helper — both single + batch modes emit the same layout.

    biz_map: ``{fqcn#method: BizResult}``. rows 는 reverse-index 를 위해 씀
    (어떤 Program 이 이 메서드를 쓰는지).
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from .legacy_biz_extractor import biz_detail_sheet_rows

    ws = wb.create_sheet("Business Logic")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="0F3460", end_color="0F3460", fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = [
        "Service#Method", "Validations", "Biz Rules", "State Changes",
        "Calculations", "External Calls", "Summary", "Source", "Programs",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sheet_rows = biz_detail_sheet_rows(biz_map, rows)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    fallback_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid",
    )
    for i, r in enumerate(sheet_rows, 2):
        values = [
            r["key"], r["validations"], r["biz_rules"], r["state_changes"],
            r["calculations"], r["external_calls"], r["summary"],
            r["source"], r["programs"],
        ]
        fill = fallback_fill if r["source"] == "fallback" else None
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = wrap_align
            if fill is not None:
                cell.fill = fill

    # Column width — 넓어지기 쉬워서 max 는 60 으로 캡.
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                longest = max(len(line) for line in str(cell.value).split("\n"))
                max_len = max(max_len, longest)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = "A2"


def _write_frontend_biz_sheet(wb, fe_biz_map: dict) -> None:
    """Phase B: Frontend Logic 시트 (React handler 단위, 8컬럼).

    columns: Screen | Button | Handler | URL | Field Validations |
             Pre-checks | Conditional Calls | State Reads | Summary | Source
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from .legacy_biz_extractor import frontend_biz_sheet_rows

    ws = wb.create_sheet("Frontend Logic")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="0F3460", end_color="0F3460", fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = [
        "Screen", "Button", "Handler", "URL",
        "Field Validations", "Pre-checks", "Conditional Calls",
        "State Reads", "Summary", "Source",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sheet_rows = frontend_biz_sheet_rows(fe_biz_map)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    fallback_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid",
    )
    for i, r in enumerate(sheet_rows, 2):
        values = [
            r["screen"], r["button"], r["handler"], r["url"],
            r["field_validations"], r["pre_checks"], r["conditional_calls"],
            r["state_reads"], r["summary"], r["source"],
        ]
        fill = fallback_fill if r["source"] == "fallback" else None
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = wrap_align
            if fill is not None:
                cell.fill = fill

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                longest = max(len(line) for line in str(cell.value).split("\n"))
                max_len = max(max_len, longest)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = "A2"


def _write_program_spec_sheet(wb, spec_map: dict, rows: list[dict]) -> None:
    """Phase II: Program Specification 시트 (endpoint 단위).

    columns: Main / Sub / Tab / Program / HTTP / URL / Trigger label /
             Trigger type / Input fields / Validations / Business flow /
             Read targets / Write targets / Purpose / Source

    spec_map 이 비어있으면 시트 생성 자체를 skip (회귀 없음). rows 는
    controller/menu context 를 endpoint_spec 에 join 하기 위해 필요.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from .legacy_biz_extractor import program_spec_sheet_rows

    ws = wb.create_sheet("Program Specification")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="0F3460", end_color="0F3460", fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = [
        "Main", "Sub", "Tab", "Program", "HTTP", "URL",
        "Trigger label", "Trigger type",
        "Input fields", "Validations", "Business flow",
        "Read targets", "Write targets",
        "Purpose", "Source",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    sheet_rows = program_spec_sheet_rows(spec_map, rows)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    # Rows where the LLM call failed (fallback-only) are tinted so the
    # operator can spot them quickly.
    fallback_fill = PatternFill(
        start_color="FFF2CC", end_color="FFF2CC", fill_type="solid",
    )
    for i, r in enumerate(sheet_rows, 2):
        values = [
            r["main_menu"], r["sub_menu"], r["tab"], r["program_name"],
            r["http_method"], r["url"],
            r["trigger_label"], r["trigger_type"],
            r["input_fields"], r["validations"], r["business_flow"],
            r["read_targets"], r["write_targets"],
            r["purpose_ko"], r["spec_source"],
        ]
        fill = fallback_fill if r["spec_source"] == "fallback" else None
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = wrap_align
            if fill is not None:
                cell.fill = fill

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                longest = max(len(line) for line in str(cell.value).split("\n"))
                max_len = max(max_len, longest)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)
    ws.freeze_panes = "A2"


def _write_sequence_diagram_sheet(wb, rows: list[dict]) -> None:
    """Phase A: Mermaid Sequence Diagrams 시트.

    각 endpoint row 가 ``sequence_diagram`` 필드 (Mermaid 텍스트) 를 가질
    때만 한 행 emit. columns: Main / Sub / Program / HTTP / URL /
    Mermaid. 사용자가 Mermaid 텍스트를 <https://mermaid.live> 또는
    VSCode 의 Mermaid Preview 확장에 복붙해서 다이어그램 확인.

    비어있는 row 는 skip — opt-in 플래그 없이 돌린 경우 아무것도 안
    그려지는 회귀 없는 동작.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    diagram_rows = [r for r in rows if (r.get("sequence_diagram") or "").strip()]
    if not diagram_rows:
        return

    ws = wb.create_sheet("Sequence Diagrams")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(
        start_color="0F3460", end_color="0F3460", fill_type="solid",
    )
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = ["Main", "Sub", "Program", "HTTP", "URL", "Mermaid"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    wrap_align = Alignment(vertical="top", wrap_text=True)
    for i, r in enumerate(diagram_rows, 2):
        vals = [
            r.get("main_menu", "") or "",
            r.get("sub_menu", "") or "",
            r.get("program_name", "") or "",
            r.get("http_method", "") or "",
            r.get("url", "") or "",
            r.get("sequence_diagram", "") or "",
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = thin_border
            cell.alignment = wrap_align

    # Mermaid 열은 긴 텍스트 — 넉넉히
    widths = [20, 20, 30, 8, 40, 80]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = w
    ws.freeze_panes = "A2"


def save_legacy_excel(result: dict, output_dir: str, menu_only: bool = False) -> str:
    """Render the analysis result as a multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = _build_filename(output_dir, result, ts, "xlsx")

    rows = result.get("rows", [])
    unmatched = result.get("unmatched_controllers", [])
    orphans = result.get("orphan_menus", [])
    stats = result.get("stats", {})

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    # Wrap text on data cells so Tables / RFC / Service / Query XML / SQL
    # ids — which now pack multiple items as ``,\n`` or ``;\n`` — render
    # with one item per visible line inside a single Excel cell.
    data_align = Alignment(vertical="top", wrap_text=True)

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def _write_row(ws, row_num, values, fill=None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            cell.alignment = data_align
            if fill is not None:
                cell.fill = fill

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, ["Category", "Value"])
    summary_rows = [
        ("Backend framework", stats.get("backend_framework", "unknown")),
        ("Controllers scanned", stats.get("controllers", 0)),
        ("Services scanned", stats.get("services", 0)),
        ("Java Mapper classes", stats.get("mappers", 0)),
        ("MyBatis XML files", stats.get("mapper_xml_files", 0)),
        ("MyBatis XML namespaces", stats.get("mapper_xml_namespaces", 0)),
        ("Endpoints total", stats.get("endpoints", 0)),
        ("Matched to menu", stats.get("matched", 0)),
        ("Unmatched controllers", stats.get("unmatched", 0)),
        ("Orphan menu entries", stats.get("orphan_menus", 0)),
        ("Endpoints with React file", stats.get("with_react", 0)),
        ("Endpoints with RFC", stats.get("with_rfc", 0)),
    ]
    for i, (k, v) in enumerate(summary_rows, 2):
        _write_row(ws, i, [k, v])
    _auto_width(ws)

    # Sheet 2: Programs (main deliverable)
    ws = wb.create_sheet("Programs")
    cols = _SINGLE_COLUMNS_WITH_MENU if _has_menu_data(rows) else _SINGLE_COLUMNS_NO_MENU
    has_row_no_col = any(k == "__row_no__" for _, k in cols)
    if has_row_no_col:
        headers = [label for label, _ in cols]
    else:
        headers = ["No"] + [label for label, _ in cols]
    _write_header(ws, headers)
    for i, r in enumerate(rows, 2):
        fill = None
        if not r.get("matched"):
            fill = yellow_fill
        elif not r.get("query_xml") and not r.get("related_tables"):
            fill = gray_fill
        if has_row_no_col:
            values = [i - 1 if k == "__row_no__" else r.get(k, "") for _, k in cols]
        else:
            values = [i - 1] + [r.get(k, "") for _, k in cols]
        _write_row(ws, i, values, fill=fill)
    ws.freeze_panes = "A2"
    _auto_width(ws)

    # Sheet 3: Menu Hierarchy
    ws = wb.create_sheet("Menu Hierarchy")
    _write_header(ws, ["Program ID", "Main", "Sub", "Tab", "Program", "URL",
                        "Matched", "# Endpoints"])
    hier_seen = set()
    row_num = 2
    for r in rows:
        if not r.get("matched"):
            continue
        key = (r.get("program_id", ""), r.get("url", ""))
        if key in hier_seen:
            continue
        hier_seen.add(key)
        count = sum(1 for x in rows if x.get("program_id") == r.get("program_id"))
        _write_row(ws, row_num, [
            r.get("program_id", ""), r.get("main_menu", ""), r.get("sub_menu", ""),
            r.get("tab", ""), r.get("program_name", ""), r.get("url", ""),
            "Y", count,
        ])
        row_num += 1
    for o in orphans:
        _write_row(ws, row_num, [
            o.get("program_id", ""), o.get("main_menu", ""), o.get("sub_menu", ""),
            o.get("tab", ""), o.get("program_name", ""), o.get("url", ""),
            "N", 0,
        ], fill=yellow_fill)
        row_num += 1
    _auto_width(ws)

    # Sheet 4: Unmatched Controllers
    ws = wb.create_sheet("Unmatched Controllers")
    _write_header(ws, ["HTTP", "URL", "Controller", "Method", "File"])
    for i, u in enumerate(unmatched, 2):
        _write_row(ws, i, [
            u.get("http_method", ""), u.get("url", ""),
            u.get("controller_class", ""), u.get("program_name", ""),
            u.get("file_name", ""),
        ])
    _auto_width(ws)

    # Sheet 5: Orphan Menu Entries
    ws = wb.create_sheet("Orphan Menu Entries")
    _write_header(ws, ["Program ID", "Main", "Sub", "Tab", "Program", "URL"])
    for i, o in enumerate(orphans, 2):
        _write_row(ws, i, [
            o.get("program_id", ""), o.get("main_menu", ""),
            o.get("sub_menu", ""), o.get("tab", ""),
            o.get("program_name", ""), o.get("url", ""),
        ])
    _auto_width(ws)

    # Sheet 6: RFC Calls (cross-reference)
    ws = wb.create_sheet("RFC Calls")
    _write_header(ws, ["RFC", "Program", "Controller", "URL", "File"])
    rfc_row = 2
    for r in rows:
        if not r.get("rfc"):
            continue
        for name in [s.strip() for s in r["rfc"].split(",") if s.strip()]:
            _write_row(ws, rfc_row, [
                name, r.get("program_name", ""),
                r.get("controller_class", ""), r.get("url", ""),
                r.get("file_name", ""),
            ])
            rfc_row += 1
    _auto_width(ws)

    # Sheet 7: Tables Cross-Reference
    ws = wb.create_sheet("Tables Cross-Reference")
    _write_header(ws, ["Table", "# Programs", "Programs"])
    table_to_programs = {}
    for r in rows:
        if not r.get("related_tables"):
            continue
        # Programs sheet stores tables as ``TABLE_A(R),\nTABLE_B(CRU)``.
        # Strip CRUD suffix so the Cross-Reference sheet keys by the
        # bare table name and aggregates across CRUD variations.
        for raw in [s.strip() for s in r["related_tables"].split(",") if s.strip()]:
            t = _bare_table_name(raw)
            if t:
                table_to_programs.setdefault(t, []).append(r.get("program_name", ""))
    for i, (table, progs) in enumerate(sorted(table_to_programs.items()), 2):
        progs_sorted = sorted(set(progs))
        preview = ", ".join(progs_sorted[:10])
        if len(progs_sorted) > 10:
            preview += f", … (+{len(progs_sorted) - 10})"
        _write_row(ws, i, [table, len(progs_sorted), preview])
    _auto_width(ws)

    # Sheet 8: Business Logic (Phase A — opt-in via --extract-biz-logic).
    # biz_map 이 비어있으면 시트 자체를 만들지 않아 기존 리포트와 동일.
    biz_map = result.get("biz_map") or {}
    if biz_map:
        _write_biz_logic_sheet(wb, biz_map, rows)
    fe_biz_map = result.get("fe_biz_map") or {}
    if fe_biz_map:
        _write_frontend_biz_sheet(wb, fe_biz_map)
    spec_map = result.get("endpoint_spec_map") or {}
    if spec_map:
        _write_program_spec_sheet(wb, spec_map, rows)
    # Mermaid Sequence Diagrams (Phase A) — row 에 sequence_diagram 필드가
    # 있는 경우에만 시트 생성. 없으면 skip → 회귀 없음.
    _write_sequence_diagram_sheet(wb, rows)

    wb.save(filepath)
    logger.info("Legacy excel saved: %s", filepath)
    return filepath


# ---------------------------------------------------------------------------
# Batch reports — multi-project monorepo output
# ---------------------------------------------------------------------------

# Column definitions: menu-first vs no-menu layouts.
#
# 메뉴가 있을 때는 **메뉴 → 프론트 → 백엔드 → SQL/RFC** 순서로 좌→우 를
# 읽어 내려가도록 배치한다. 맨 앞 "No" 는 렌더 시점에 자동 번호 부여.
_BATCH_COLUMNS_WITH_MENU = [
    ("No",                "__row_no__"),
    ("메뉴1뎁스",         "main_menu"),
    ("메뉴2뎁스",         "sub_menu"),
    ("메뉴3뎁스",         "tab"),
    ("Menu path",         "menu_path"),
    ("Menu URL",          "menu_url"),
    ("Frontend project",  "frontend_project"),
    ("Frontend screen",   "presentation_layer"),
    ("Trigger",           "frontend_trigger"),
    ("Frontend Validation", "frontend_validation_summary"),
    ("Backend project",   "backend_project"),
    ("Backend framework", "backend_framework"),
    ("Program",           "program_name"),
    ("HTTP",              "http_method"),
    ("Controller URL",    "url"),
    ("Controller file",   "file_name"),
    ("Controller",        "controller_class"),
    ("Service",           "service_class"),
    ("Service method",    "service_methods"),
    ("Business Logic",    "biz_summary"),
    ("XML",               "query_xml"),
    ("XML method",        "sql_ids"),
    ("Table",             "related_tables"),
    ("Columns",           "related_columns"),
    ("Procedure",         "procedures"),
    ("RFC",               "rfc"),
]

_BATCH_COLUMNS_NO_MENU = [
    ("Backend project",   "backend_project"),
    ("Backend framework", "backend_framework"),
    ("Frontend project",  "frontend_project"),
    ("Frontend screen",   "presentation_layer"),
    ("Trigger",           "frontend_trigger"),
    ("Frontend Validation", "frontend_validation_summary"),
    ("File",              "file_name"),
    ("Controller",        "controller_class"),
    ("URL",               "url"),
    ("HTTP",              "http_method"),
    ("Program",           "program_name"),
    ("Service",           "service_class"),
    ("Service method",    "service_methods"),
    ("Business Logic",    "biz_summary"),
    ("XML",               "query_xml"),
    ("XML method",        "sql_ids"),
    ("Table",             "related_tables"),
    ("Columns",           "related_columns"),
    ("Procedure",         "procedures"),
    ("RFC",               "rfc"),
]

_SINGLE_COLUMNS_WITH_MENU = [
    ("No",                "__row_no__"),
    ("메뉴1뎁스",         "main_menu"),
    ("메뉴2뎁스",         "sub_menu"),
    ("메뉴3뎁스",         "tab"),
    ("Menu path",         "menu_path"),
    ("Menu URL",          "menu_url"),
    ("Frontend project",  "frontend_project"),
    ("Frontend screen",   "presentation_layer"),
    ("Trigger",           "frontend_trigger"),
    ("Frontend Validation", "frontend_validation_summary"),
    ("Program",           "program_name"),
    ("HTTP",              "http_method"),
    ("Controller URL",    "url"),
    ("Controller file",   "file_name"),
    ("Controller",        "controller_class"),
    ("Service",           "service_class"),
    ("Service method",    "service_methods"),
    ("Business Logic",    "biz_summary"),
    ("XML",               "query_xml"),
    ("XML method",        "sql_ids"),
    ("Tables",            "related_tables"),
    ("Columns",           "related_columns"),
    ("Procedure",         "procedures"),
    ("RFC",               "rfc"),
]

_SINGLE_COLUMNS_NO_MENU = [
    ("Program",           "program_name"),
    ("HTTP",              "http_method"),
    ("URL",               "url"),
    ("File",              "file_name"),
    ("Frontend project",  "frontend_project"),
    ("React",             "presentation_layer"),
    ("Trigger",           "frontend_trigger"),
    ("Frontend Validation", "frontend_validation_summary"),
    ("Controller",        "controller_class"),
    ("Service",           "service_class"),
    ("Service method",    "service_methods"),
    ("Business Logic",    "biz_summary"),
    ("XML",               "query_xml"),
    ("XML method",        "sql_ids"),
    ("Tables",            "related_tables"),
    ("Columns",           "related_columns"),
    ("Procedure",         "procedures"),
    ("RFC",               "rfc"),
]


def _has_menu_data(rows: list[dict]) -> bool:
    """Return True if any row has non-empty menu info."""
    return any(r.get("main_menu") or r.get("menu_path") for r in rows)


def _build_batch_filename(output_dir: str, ts: str, ext: str) -> str:
    """``<output>/legacy_analysis/as_is_analysis_batch_<ts>.<ext>``."""
    return os.path.join(_legacy_output_dir(output_dir), f"as_is_analysis_batch_{ts}.{ext}")


def save_legacy_batch_markdown(result: dict, output_dir: str, menu_only: bool = False) -> str:
    """Render a batch (multi-project) analysis result as Markdown."""
    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = _build_batch_filename(output_dir, ts, "md")

    rows = result.get("rows", [])
    stats = result.get("stats", {})
    per_project = result.get("per_project_stats", {}) or {}
    project_frameworks = result.get("project_frameworks", {}) or {}

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# AS-IS Legacy Source Analysis (Batch)\n\n")
        f.write(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"- Backends root: `{result.get('backends_root', '')}`\n")
        if result.get("frontend_dir"):
            f.write(f"- Frontend dir: `{result.get('frontend_dir', '')}`\n")
        f.write(f"- Projects analyzed: {stats.get('projects', 0)}\n")
        f.write("\n")

        f.write("## Summary\n\n")
        f.write("| Category | Value |\n|---|---|\n")
        for label, key in [
            ("Projects", "projects"),
            ("Controllers scanned", "controllers"),
            ("Services scanned", "services"),
            ("Java Mapper classes", "mappers"),
            ("MyBatis XML files", "mapper_xml_files"),
            ("MyBatis XML namespaces", "mapper_xml_namespaces"),
            ("Endpoints total", "endpoints"),
            ("Matched to menu", "matched"),
            ("Unmatched controllers", "unmatched"),
            ("Endpoints with React file", "with_react"),
            ("Endpoints with RFC", "with_rfc"),
            ("Resolved via method-scope", "resolved_method_scope"),
            ("Resolved via class-scope fallback", "resolved_class_scope"),
        ]:
            f.write(f"| {label} | {stats.get(key, 0)} |\n")
        f.write("\n")

        if per_project:
            f.write("## Per-project breakdown\n\n")
            f.write("| Project | Framework | Controllers | Endpoints | Method-scope | Fallback | With RFC |\n")
            f.write("|---|---|---|---|---|---|---|\n")
            for name in sorted(per_project.keys()):
                ps = per_project[name]
                f.write(
                    f"| {name} "
                    f"| {project_frameworks.get(name, '') or ps.get('backend_framework', '')} "
                    f"| {ps.get('controllers', 0)} "
                    f"| {ps.get('endpoints', 0)} "
                    f"| {ps.get('resolved_method_scope', 0)} "
                    f"| {ps.get('resolved_class_scope', 0)} "
                    f"| {ps.get('with_rfc', 0)} |\n"
                )
            f.write("\n")

        batch_cols = _BATCH_COLUMNS_WITH_MENU if _has_menu_data(rows) else _BATCH_COLUMNS_NO_MENU
        f.write("## Program Detail\n\n")
        f.write("| " + " | ".join(label for label, _ in batch_cols) + " |\n")
        f.write("|" + "|".join(["---"] * len(batch_cols)) + "|\n")
        for idx, r in enumerate(rows, 1):
            f.write(
                "| " + " | ".join(
                    str(idx) if key == "__row_no__"
                    else _md_escape(r.get(key, ""))
                    for _, key in batch_cols
                ) + " |\n"
            )
        f.write("\n")

        # Mermaid Sequence Diagrams (Phase A) — single-mode 와 동일 구조
        diagram_rows = [r for r in rows if (r.get("sequence_diagram") or "").strip()]
        if diagram_rows:
            f.write(f"## Sequence Diagrams ({len(diagram_rows)})\n\n")
            f.write("endpoint 당 controller → service → XML → DB / RFC 호출 체인 "
                    "(source offset 순서, LLM 없이 static 추출).\n\n")
            for r in diagram_rows:
                title = (f"{r.get('http_method','')} {r.get('url','')}".strip()
                         or r.get("program_name", "") or "endpoint")
                project = r.get("backend_project", "")
                if project:
                    title = f"[{project}] {title}"
                f.write(f"### {title}\n\n")
                f.write("```mermaid\n")
                f.write(r["sequence_diagram"])
                f.write("\n```\n\n")

        unmatched = result.get("unmatched_controllers", [])
        if unmatched:
            f.write(f"## Unmatched Controllers ({len(unmatched)})\n\n")
            f.write("| Project | HTTP | URL | Controller | Method | File |\n")
            f.write("|---|---|---|---|---|---|\n")
            for u in unmatched:
                f.write(
                    f"| {_md_escape(u.get('backend_project', ''))} "
                    f"| {u.get('http_method', '')} | {_md_escape(u.get('url', ''))} "
                    f"| {_md_escape(u.get('controller_class', ''))} "
                    f"| {_md_escape(u.get('program_name', ''))} "
                    f"| {_md_escape(u.get('file_name', ''))} |\n"
                )
            f.write("\n")

    logger.info("Legacy batch markdown saved: %s", filepath)
    return filepath


def save_legacy_batch_excel(result: dict, output_dir: str, menu_only: bool = False) -> str:
    """Render a batch (multi-project) analysis result as a multi-sheet workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    os.makedirs(output_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = _build_batch_filename(output_dir, ts, "xlsx")

    rows = result.get("rows", [])
    unmatched = result.get("unmatched_controllers", [])
    stats = result.get("stats", {})
    per_project = result.get("per_project_stats", {}) or {}
    project_frameworks = result.get("project_frameworks", {}) or {}

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    # Wrap text on data cells so Tables / RFC / Service / Query XML / SQL
    # ids — which now pack multiple items as ``,\n`` or ``;\n`` — render
    # with one item per visible line inside a single Excel cell.
    data_align = Alignment(vertical="top", wrap_text=True)

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def _write_row(ws, row_num, values, fill=None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            cell.alignment = data_align
            if fill is not None:
                cell.fill = fill

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # Sheet 1: Summary (aggregate + per-project breakdown)
    ws = wb.active
    ws.title = "Summary"
    _write_header(ws, ["Category", "Value"])
    summary_rows = [
        ("Projects", stats.get("projects", 0)),
        ("Controllers scanned", stats.get("controllers", 0)),
        ("Services scanned", stats.get("services", 0)),
        ("Java Mapper classes", stats.get("mappers", 0)),
        ("MyBatis XML files", stats.get("mapper_xml_files", 0)),
        ("MyBatis XML namespaces", stats.get("mapper_xml_namespaces", 0)),
        ("Endpoints total", stats.get("endpoints", 0)),
        ("Matched to menu", stats.get("matched", 0)),
        ("Unmatched controllers", stats.get("unmatched", 0)),
        ("Endpoints with React file", stats.get("with_react", 0)),
        ("Endpoints with RFC", stats.get("with_rfc", 0)),
        ("Resolved via method-scope", stats.get("resolved_method_scope", 0)),
        ("Resolved via class-scope fallback", stats.get("resolved_class_scope", 0)),
    ]
    for i, (k, v) in enumerate(summary_rows, 2):
        _write_row(ws, i, [k, v])
    _auto_width(ws)

    # Sheet 2: Per-project breakdown
    ws = wb.create_sheet("Per Project")
    _write_header(ws, ["Project", "Framework", "Controllers", "Services",
                        "Mappers", "Endpoints", "Method-scope",
                        "Fallback", "With RFC"])
    for i, name in enumerate(sorted(per_project.keys()), 2):
        ps = per_project[name]
        _write_row(ws, i, [
            name,
            project_frameworks.get(name, "") or ps.get("backend_framework", ""),
            ps.get("controllers", 0),
            ps.get("services", 0),
            ps.get("mappers", 0),
            ps.get("endpoints", 0),
            ps.get("resolved_method_scope", 0),
            ps.get("resolved_class_scope", 0),
            ps.get("with_rfc", 0),
        ])
    _auto_width(ws)

    # Sheet 3: Programs (the requested column order)
    ws = wb.create_sheet("Programs")
    batch_cols = _BATCH_COLUMNS_WITH_MENU if _has_menu_data(rows) else _BATCH_COLUMNS_NO_MENU
    has_row_no_col = any(k == "__row_no__" for _, k in batch_cols)
    if has_row_no_col:
        headers = [label for label, _ in batch_cols]
    else:
        headers = ["No"] + [label for label, _ in batch_cols]
    _write_header(ws, headers)
    for i, r in enumerate(rows, 2):
        fill = None
        if not r.get("matched"):
            fill = yellow_fill
        elif not r.get("query_xml") and not r.get("related_tables"):
            fill = gray_fill
        if has_row_no_col:
            values = [i - 1 if k == "__row_no__" else r.get(k, "") for _, k in batch_cols]
        else:
            values = [i - 1] + [r.get(k, "") for _, k in batch_cols]
        _write_row(ws, i, values, fill=fill)
    ws.freeze_panes = "A2"
    _auto_width(ws)

    # Sheet 4: Unmatched Controllers
    ws = wb.create_sheet("Unmatched Controllers")
    _write_header(ws, ["Backend project", "HTTP", "URL", "Controller", "Method", "File"])
    for i, u in enumerate(unmatched, 2):
        _write_row(ws, i, [
            u.get("backend_project", ""),
            u.get("http_method", ""),
            u.get("url", ""),
            u.get("controller_class", ""),
            u.get("program_name", ""),
            u.get("file_name", ""),
        ])
    _auto_width(ws)

    # Sheet 5: RFC Calls (cross-reference across projects)
    ws = wb.create_sheet("RFC Calls")
    _write_header(ws, ["RFC", "Backend project", "Program", "Controller", "URL", "File"])
    rfc_row = 2
    for r in rows:
        if not r.get("rfc"):
            continue
        for name in [s.strip() for s in r["rfc"].split(",") if s.strip()]:
            _write_row(ws, rfc_row, [
                name,
                r.get("backend_project", ""),
                r.get("program_name", ""),
                r.get("controller_class", ""),
                r.get("url", ""),
                r.get("file_name", ""),
            ])
            rfc_row += 1
    _auto_width(ws)

    # Sheet 6: Tables Cross-Reference (across projects)
    ws = wb.create_sheet("Tables Cross-Reference")
    _write_header(ws, ["Table", "# Programs", "# Projects", "Projects", "Programs"])
    table_to_programs: dict[str, list[str]] = {}
    table_to_projects: dict[str, list[str]] = {}
    for r in rows:
        if not r.get("related_tables"):
            continue
        program = r.get("program_name", "")
        project = r.get("backend_project", "")
        # Strip ``(CRUD)`` suffix so batch cross-ref keys by bare name
        # (matches single-mode aggregation).
        for raw in [s.strip() for s in r["related_tables"].split(",") if s.strip()]:
            t = _bare_table_name(raw)
            if not t:
                continue
            table_to_programs.setdefault(t, []).append(program)
            if project:
                table_to_projects.setdefault(t, []).append(project)
    for i, (table, progs) in enumerate(sorted(table_to_programs.items()), 2):
        progs_sorted = sorted(set(progs))
        projs_sorted = sorted(set(table_to_projects.get(table, [])))
        prog_preview = ", ".join(progs_sorted[:10])
        if len(progs_sorted) > 10:
            prog_preview += f", … (+{len(progs_sorted) - 10})"
        proj_preview = ", ".join(projs_sorted)
        _write_row(ws, i, [
            table, len(progs_sorted), len(projs_sorted),
            proj_preview, prog_preview,
        ])
    _auto_width(ws)

    # Sheet 7: Business Logic (Phase A — opt-in).
    biz_map = result.get("biz_map") or {}
    if biz_map:
        _write_biz_logic_sheet(wb, biz_map, rows)
    fe_biz_map = result.get("fe_biz_map") or {}
    if fe_biz_map:
        _write_frontend_biz_sheet(wb, fe_biz_map)
    spec_map = result.get("endpoint_spec_map") or {}
    if spec_map:
        _write_program_spec_sheet(wb, spec_map, rows)
    # Mermaid Sequence Diagrams (Phase A) — row 에 sequence_diagram 필드가
    # 있는 경우에만 시트 생성. 없으면 skip → 회귀 없음.
    _write_sequence_diagram_sheet(wb, rows)

    wb.save(filepath)
    logger.info("Legacy batch excel saved: %s", filepath)
    return filepath
