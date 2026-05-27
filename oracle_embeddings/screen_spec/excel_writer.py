"""ScreenSpec 리스트 → 마스터 xlsx (시트=영역, 1열=화면명).

openpyxl 사용. python-pptx 가 lxml 을 끌어와서 openpyxl 도 같이 들어오는
일반적인 경우 OK. 별도 설치되어 있어야 한다.

시트 구성 (모든 시트의 1열은 '화면명'):
 - 개요
 - 검색조건
 - 그리드컬럼
 - 탭
 - 이벤트
 - 검증규칙
 - 이벤트플로우

같은 ScreenSpec → 같은 xlsx (서식 포함 byte-level identical 가까움).
"""
from __future__ import annotations

from pathlib import Path

from .models import ScreenSpec


_HEADERS = {
    "개요": ["화면명", "entry 파일", "closure 파일 수", "closure tokens",
            "truncated", "검색 필드 수", "그리드 컬럼 수", "탭 수",
            "버튼 수", "검증 규칙 수", "API 호출 수 (factual)",
            "팝업 ref 수"],
    "검색조건": ["화면명", "No", "라벨", "타입", "길이",
                "필수", "기본값", "유효성 규칙 및 비고", "UI 타입",
                "동작", "필드명", "소스파일"],
    "그리드컬럼": ["화면명", "NO", "필드명(영문)", "필드설명", "타입",
                  "필수여부", "속성", "UI타입", "설명", "동작",
                  "너비", "정렬", "소스파일"],
    "탭": ["화면명", "순번", "탭명", "컨텐츠 컴포넌트", "소스파일"],
    "이벤트": ["화면명", "트리거(라벨)", "종류", "핸들러", "API 호출",
              "화면 호출", "비고", "소스파일"],
    "검증규칙": ["화면명", "필드", "규칙", "상세", "메시지", "출처",
                "소스파일"],
    "이벤트플로우": ["화면명", "이벤트", "step#", "동작", "상세", "조건"],
}


def _row_for_overview(s: ScreenSpec) -> list:
    return [
        s.screen_id,
        s.entry_file,
        s.closure_file_count,
        s.closure_tokens,
        "Y" if s.closure_truncated else "N",
        len(s.form_fields),
        len(s.grid_columns),
        len(s.tabs),
        len(s.buttons),
        len(s.validations),
        len(s.api_calls_factual),
        len(s.popup_refs_factual),
    ]


def _rows_for_form_fields(s: ScreenSpec):
    for ff in s.form_fields:
        # 기본값: placeholder 우선 (UI 가시값) → defaultValue fallback.
        # placeholder 가 "Select 하세요" 같은 빈 안내문이라도 default 보다
        # 사용자 화면에 보이는 값이므로 우선.
        display_default = ff.placeholder or ff.default or ""
        # 유효성 규칙 = LLM 판단 칸 (validation_rule) > 인라인 prop 요약 (validation)
        validation_display = ff.validation_rule or ff.validation or ""
        yield [
            s.screen_id, ff.order, ff.label,
            ff.input_data_type,         # 타입 (keyboard input 만)
            ff.max_length,              # 길이 (keyboard input 만)
            "필수" if ff.required else "선택",
            display_default,
            validation_display,
            ff.ui_type or "",
            ff.action or "",
            ff.name,
            ff.source_file,
        ]


def _rows_for_grid_columns(s: ScreenSpec):
    from .extractors import _compose_attribute
    for c in s.grid_columns:
        attribute = _compose_attribute(c.visible, c.editable)
        yield [s.screen_id, c.order, c.data_key, c.header, c.data_type,
               "필수" if c.required else "",
               attribute,
               c.ui_type or "Text Field(Basic)",
               c.description, c.action,
               c.width, "Y" if c.sortable else "N", c.source_file]


def _rows_for_tabs(s: ScreenSpec):
    for t in s.tabs:
        yield [s.screen_id, t.order, t.label, t.panel_component, t.source_file]


def _rows_for_buttons(s: ScreenSpec):
    for b in s.buttons:
        yield [s.screen_id, b.trigger_label, b.trigger_kind, b.handler_name,
               "\n".join(b.api_calls) if b.api_calls else "",
               "\n".join(b.screen_calls) if b.screen_calls else "",
               b.notes, b.source_file]


def _rows_for_validations(s: ScreenSpec):
    for v in s.validations:
        yield [s.screen_id, v.field, v.rule, v.detail, v.message,
               v.source, v.source_file]


def _rows_for_flows(s: ScreenSpec):
    for b in s.buttons:
        for step in b.flow:
            yield [s.screen_id, b.trigger_label or b.handler_name,
                   step.step, step.action, step.detail, step.condition]


def write_master_xlsx(specs: list[ScreenSpec], output_path: Path) -> None:
    """ScreenSpec 리스트 → 1개 마스터 xlsx (7시트)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as e:
        raise SystemExit(
            "openpyxl 미설치. `pip install openpyxl` 필요."
        ) from e

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # default 시트 제거
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="top", wrap_text=True)

    def _add_sheet(title: str, headers: list[str], rows: list[list]):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for r in rows:
            ws.append(r)
        # 너비 자동 조정 (대략)
        for col_idx, h in enumerate(headers, start=1):
            # 헤더 길이 기반 + 데이터 max 길이
            max_len = len(str(h))
            for r in rows:
                if col_idx - 1 < len(r):
                    s = str(r[col_idx - 1] or "")
                    # 줄바꿈 있으면 가장 긴 줄로
                    line_max = max((len(line) for line in s.split("\n")),
                                   default=0)
                    if line_max > max_len:
                        max_len = line_max
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = \
                min(60, max(8, max_len + 2))
        # 데이터 행 wrap
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for c in row_cells:
                c.alignment = wrap_align
        # 1행 freeze
        ws.freeze_panes = "A2"

    # 시트별 데이터 집계
    overview_rows = [_row_for_overview(s) for s in specs]
    form_rows = [r for s in specs for r in _rows_for_form_fields(s)]
    grid_rows = [r for s in specs for r in _rows_for_grid_columns(s)]
    tab_rows = [r for s in specs for r in _rows_for_tabs(s)]
    btn_rows = [r for s in specs for r in _rows_for_buttons(s)]
    val_rows = [r for s in specs for r in _rows_for_validations(s)]
    flow_rows = [r for s in specs for r in _rows_for_flows(s)]

    _add_sheet("개요",       _HEADERS["개요"],       overview_rows)
    _add_sheet("검색조건",   _HEADERS["검색조건"],   form_rows)
    _add_sheet("그리드컬럼", _HEADERS["그리드컬럼"], grid_rows)
    _add_sheet("탭",         _HEADERS["탭"],         tab_rows)
    _add_sheet("이벤트",     _HEADERS["이벤트"],     btn_rows)
    _add_sheet("검증규칙",   _HEADERS["검증규칙"],   val_rows)
    _add_sheet("이벤트플로우", _HEADERS["이벤트플로우"], flow_rows)

    try:
        wb.save(str(output_path))
    except PermissionError as e:
        raise SystemExit(
            f"xlsx 저장 실패 — {output_path}\n"
            f"  원인: 파일이 다른 프로그램(Excel 등) 에 열려있어 잠긴 상태.\n"
            f"  해결: 1) 해당 Excel 창 닫고 재실행, 또는\n"
            f"        2) --output <다른경로.xlsx> 로 다른 파일명 지정.\n"
            f"  상세: {e}"
        ) from e
