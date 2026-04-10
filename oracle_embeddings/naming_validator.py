import logging
import os
import re
from difflib import get_close_matches

logger = logging.getLogger(__name__)

# Oracle 제약
MAX_TABLE_NAME_LEN = 30  # Oracle 11g 기준 (12c+부터는 128)
MAX_COLUMN_NAME_LEN = 30

# 허용 접두어/접미어
TABLE_PREFIXES = {"TB", "TBL", "T", "VW", "V"}
PK_SUFFIXES = {"ID", "NO", "CD", "SEQ"}


class NamingValidator:
    """네이밍룰 검증 엔진."""

    def __init__(self, terms_dict_path: str = None):
        """
        terms_dict_path: 용어사전 .md 파일 경로 (선택)
                         없으면 기본 약어 사전만 사용
        """
        self.standard_words = set()  # 표준 영문 Full
        self.standard_abbreviations = set()  # 표준 약어
        self.word_to_abbr = {}  # Full → abbr 매핑
        self.abbr_to_word = {}  # abbr → Full 매핑

        if terms_dict_path and os.path.exists(terms_dict_path):
            self._load_terms_dict(terms_dict_path)
        else:
            self._load_default_dict()

        logger.info("Loaded %d standard words, %d abbreviations",
                    len(self.standard_words), len(self.standard_abbreviations))

    def _load_terms_dict(self, path: str):
        """용어사전 .md 파일에서 표준 단어 로드."""
        with open(path, "r", encoding="utf-8") as f:
            content = f.read()
        content = content.replace("\r\n", "\n")

        # 테이블 행: | Word | Abbreviation | English Full | Korean | DB | FE | Total |
        pattern = r'^\|\s*(\w+)\s*\|\s*(\w+)?\s*\|\s*(\w+)?\s*\|\s*(\S+)?\s*\|'
        for match in re.finditer(pattern, content, re.MULTILINE):
            word = match.group(1)
            if word in ("Word", "--------", "-----"):
                continue
            abbr = match.group(2) or ""
            english = match.group(3) or ""

            self.standard_words.add(word.upper())
            if abbr:
                self.standard_abbreviations.add(abbr.upper())
                if english:
                    self.word_to_abbr[english.upper()] = abbr.upper()
                    self.abbr_to_word[abbr.upper()] = english.upper()

    def _load_default_dict(self):
        """기본 약어 사전 (용어사전 없을 때 fallback)."""
        default_abbrs = {
            "CUST": "CUSTOMER", "ORD": "ORDER", "PROD": "PRODUCT",
            "EMP": "EMPLOYEE", "DEPT": "DEPARTMENT", "USR": "USER",
            "DT": "DATE", "DTM": "DATETIME", "NM": "NAME", "NO": "NUMBER",
            "CD": "CODE", "ID": "IDENTIFIER", "ST": "STATUS", "TYPE": "TYPE",
            "AMT": "AMOUNT", "QTY": "QUANTITY", "PRC": "PRICE",
            "YN": "YESNO", "SEQ": "SEQUENCE", "CNT": "COUNT",
            "REG": "REGISTER", "MOD": "MODIFY", "DEL": "DELETE",
            "ADDR": "ADDRESS", "TEL": "TELEPHONE", "EMAIL": "EMAIL",
        }
        for abbr, full in default_abbrs.items():
            self.standard_abbreviations.add(abbr)
            self.standard_words.add(full)
            self.abbr_to_word[abbr] = full
            self.word_to_abbr[full] = abbr

    def validate_name(self, name: str, kind: str = "table") -> dict:
        """이름 검증.

        kind: 'table' or 'column'
        """
        issues = []
        name_upper = name.upper()

        # 1. 길이 체크
        max_len = MAX_TABLE_NAME_LEN if kind == "table" else MAX_COLUMN_NAME_LEN
        if len(name) > max_len:
            issues.append({
                "severity": "CRITICAL",
                "rule": "LENGTH",
                "message": f"길이 초과 ({len(name)}/{max_len}자)",
            })

        # 2. 대소문자 체크
        if any(c.islower() for c in name):
            issues.append({
                "severity": "HIGH",
                "rule": "CASE",
                "message": "대문자 + 언더스코어 형식 권장 (SNAKE_CASE)",
            })

        # 3. 특수문자 체크
        if re.search(r'[^a-zA-Z0-9_]', name):
            issues.append({
                "severity": "CRITICAL",
                "rule": "SPECIAL_CHAR",
                "message": "특수문자 사용 불가 (영문/숫자/_만 허용)",
            })

        # 4. 첫 글자 체크
        if name and name[0].isdigit():
            issues.append({
                "severity": "CRITICAL",
                "rule": "FIRST_CHAR",
                "message": "첫 글자는 숫자 불가",
            })

        # 3. 접두어 체크 (테이블만)
        if kind == "table":
            first_part = name_upper.split("_")[0] if "_" in name_upper else name_upper
            if first_part not in TABLE_PREFIXES:
                issues.append({
                    "severity": "LOW",
                    "rule": "PREFIX",
                    "message": f"테이블 접두어 권장: {', '.join(sorted(TABLE_PREFIXES))}",
                })

        # 4. 단어별 약어 검증
        parts = name_upper.split("_")
        unknown_parts = []
        suggestions = {}

        for part in parts:
            if not part:
                continue
            # 테이블 접두어는 스킵
            if kind == "table" and part in TABLE_PREFIXES:
                continue
            # 이미 표준 약어 또는 표준 단어
            if part in self.standard_abbreviations or part in self.standard_words:
                continue

            unknown_parts.append(part)
            # 유사 약어 추천
            close = get_close_matches(
                part,
                list(self.standard_abbreviations) + list(self.standard_words),
                n=3, cutoff=0.6,
            )
            if close:
                suggestions[part] = close

        if unknown_parts:
            issues.append({
                "severity": "MEDIUM",
                "rule": "UNKNOWN_ABBREVIATION",
                "message": f"표준 사전에 없는 약어: {', '.join(unknown_parts)}",
                "unknown_parts": unknown_parts,
                "suggestions": suggestions,
            })

        return {
            "name": name,
            "kind": kind,
            "valid": len(issues) == 0,
            "issues": issues,
            "parts": parts,
            "unknown_parts": unknown_parts,
            "suggestions": suggestions,
        }

    def validate_ddl(self, ddl_text: str) -> list[dict]:
        """DDL 텍스트에서 CREATE TABLE 구문 추출하여 검증."""
        results = []

        # CREATE TABLE 추출
        table_pattern = r'CREATE\s+TABLE\s+["\']?(\w+)["\']?\s*\('
        for match in re.finditer(table_pattern, ddl_text, re.IGNORECASE):
            table_name = match.group(1)
            table_result = self.validate_name(table_name, kind="table")
            table_result["columns"] = []

            # 해당 테이블의 컬럼 추출
            start = match.end()
            # 매칭 괄호 찾기
            depth = 1
            end = start
            while end < len(ddl_text) and depth > 0:
                if ddl_text[end] == '(':
                    depth += 1
                elif ddl_text[end] == ')':
                    depth -= 1
                end += 1

            body = ddl_text[start:end-1]
            # 컬럼 추출: COLUMN_NAME DATA_TYPE ...
            col_pattern = r'^\s*["\']?(\w+)["\']?\s+(?:VARCHAR2|NUMBER|DATE|CHAR|CLOB|BLOB|NVARCHAR2|NCHAR|FLOAT|TIMESTAMP)'
            for col_match in re.finditer(col_pattern, body, re.MULTILINE | re.IGNORECASE):
                col_name = col_match.group(1)
                if col_name.upper() in ("PRIMARY", "FOREIGN", "CONSTRAINT", "UNIQUE", "CHECK"):
                    continue
                col_result = self.validate_name(col_name, kind="column")
                table_result["columns"].append(col_result)

            results.append(table_result)

        return results


def format_result_console(result: dict) -> str:
    """콘솔 출력용 포맷팅."""
    lines = []
    name = result["name"]
    kind = result["kind"]

    if result["valid"]:
        lines.append(f"  OK {name} ({kind})")
    else:
        lines.append(f"  FAIL {name} ({kind})")
        for issue in result["issues"]:
            lines.append(f"    [{issue['severity']}] {issue['rule']}: {issue['message']}")
            if issue.get("suggestions"):
                for unknown, suggs in issue["suggestions"].items():
                    lines.append(f"      → {unknown} 추천: {', '.join(suggs)}")

    return "\n".join(lines)


def save_validation_report(results: list[dict], output_dir: str) -> tuple:
    """검증 결과를 Markdown + Excel로 저장."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    md_path = os.path.join(output_dir, f"naming_validation_{timestamp}.md")
    xlsx_path = os.path.join(output_dir, f"naming_validation_{timestamp}.xlsx")

    valid_count = sum(1 for r in results if r["valid"])
    invalid_count = len(results) - valid_count

    # Markdown
    with open(md_path, "w", encoding="utf-8") as f:
        f.write("# Naming Validation Report\n\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"- Total: {len(results)}\n")
        f.write(f"- Valid: {valid_count}\n")
        f.write(f"- Invalid: {invalid_count}\n\n")

        f.write("## Invalid Names\n\n")
        f.write("| Name | Kind | Severity | Rule | Message | Suggestions |\n")
        f.write("|------|------|----------|------|---------|-------------|\n")
        for r in results:
            if r["valid"]:
                continue
            for issue in r["issues"]:
                sugg = ""
                if issue.get("suggestions"):
                    sugg = "; ".join(
                        f"{u}→{','.join(s)}" for u, s in issue["suggestions"].items()
                    )
                f.write(f"| {r['name']} | {r['kind']} | {issue['severity']} "
                        f"| {issue['rule']} | {issue['message']} | {sugg} |\n")
        f.write("\n")

    # Excel
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws = wb.active
    ws.title = "Validation"
    headers = ["Name", "Kind", "Valid", "Severity", "Rule", "Message", "Suggestions"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for r in results:
        if not r["issues"]:
            ws.cell(row=row, column=1, value=r["name"]).border = thin_border
            ws.cell(row=row, column=2, value=r["kind"]).border = thin_border
            ws.cell(row=row, column=3, value="OK").border = thin_border
            row += 1
        else:
            for issue in r["issues"]:
                sugg = ""
                if issue.get("suggestions"):
                    sugg = "; ".join(
                        f"{u}→{','.join(s)}" for u, s in issue["suggestions"].items()
                    )
                values = [
                    r["name"], r["kind"], "FAIL",
                    issue["severity"], issue["rule"], issue["message"], sugg,
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    cell.border = thin_border
                row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(xlsx_path)

    return md_path, xlsx_path
