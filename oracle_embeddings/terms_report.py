import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)


def _md_escape(text: str) -> str:
    """Escape characters that would break a Markdown table cell."""
    if not text:
        return ""
    return str(text).replace("|", "\\|").replace("\n", "<br>").replace("\r", "")


def save_terms_markdown(words: list[dict], output_dir: str) -> str:
    """Save terminology dictionary as Markdown."""
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"terms_dictionary_{timestamp}.md")

    enriched = [w for w in words if w.get("korean")]
    unknown = [w for w in words if not w.get("korean")]

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("# Terminology Dictionary\n\n")
        f.write(f"- Total words: {len(words)}\n")
        f.write(f"- Enriched (LLM): {len(enriched)}\n")
        f.write(f"- Unknown: {len(unknown)}\n")
        f.write(f"- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

        # Enriched terms
        f.write("## Terminology\n\n")
        f.write("| Word | Abbreviation | English Full | Korean | Definition | DB | FE | Total |\n")
        f.write("|------|-------------|-------------|--------|------------|----|----|-------|\n")
        for w in enriched:
            f.write(f"| {w['word']} | {w.get('abbreviation', '')} "
                    f"| {w.get('english_full', '')} | {w.get('korean', '')} "
                    f"| {_md_escape(w.get('definition', ''))} "
                    f"| {w['db_count']} | {w['fe_count']} | {w['total_count']} |\n")
        f.write("\n")

        # DB only terms
        db_only = [w for w in enriched if w["db_count"] > 0 and w["fe_count"] == 0]
        fe_only = [w for w in enriched if w["fe_count"] > 0 and w["db_count"] == 0]
        both = [w for w in enriched if w["db_count"] > 0 and w["fe_count"] > 0]

        if both:
            f.write(f"## DB + FE 공통 단어 ({len(both)})\n\n")
            f.write("DB와 프론트엔드 양쪽에서 사용되는 단어입니다. 표준화 우선 대상입니다.\n\n")
            f.write("| Word | Abbreviation | English Full | Korean | Definition | DB | FE |\n")
            f.write("|------|-------------|-------------|--------|------------|----|----|  \n")
            for w in both:
                f.write(f"| {w['word']} | {w.get('abbreviation', '')} "
                        f"| {w.get('english_full', '')} | {w.get('korean', '')} "
                        f"| {_md_escape(w.get('definition', ''))} "
                        f"| {w['db_count']} | {w['fe_count']} |\n")
            f.write("\n")

        # Unknown terms
        if unknown:
            f.write(f"## Unknown Words ({len(unknown)})\n\n")
            f.write("LLM이 해석하지 못한 단어입니다.\n\n")
            f.write("| Word | DB | FE | Sample Sources |\n")
            f.write("|------|----|----|----------------|\n")
            for w in unknown:
                sources = ", ".join(w.get("sample_sources", [])[:3])
                f.write(f"| {w['word']} | {w['db_count']} | {w['fe_count']} | {sources} |\n")
            f.write("\n")

    logger.info("Terms markdown saved: %s", filepath)
    return filepath


def save_terms_excel(words: list[dict], output_dir: str) -> str:
    """Save terminology dictionary as Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(output_dir, f"terms_dictionary_{timestamp}.xlsx")

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def _write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

    def _write_row(ws, row_num, values, highlight=False):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=v)
            cell.border = thin_border
            if highlight:
                cell.fill = yellow_fill

    def _auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    enriched = [w for w in words if w.get("korean")]
    unknown = [w for w in words if not w.get("korean")]

    # Sheet 1: 전체 용어사전
    ws_all = wb.active
    ws_all.title = "용어사전"
    _write_header(ws_all, ["Word", "Abbreviation", "English Full", "Korean", "Definition",
                            "DB Count", "FE Count", "Total", "Sources"])

    for i, w in enumerate(enriched, 2):
        is_both = w["db_count"] > 0 and w["fe_count"] > 0
        _write_row(ws_all, i, [
            w["word"], w.get("abbreviation", ""), w.get("english_full", ""),
            w.get("korean", ""), w.get("definition", ""),
            w["db_count"], w["fe_count"], w["total_count"],
            ", ".join(w.get("sample_sources", [])[:3]),
        ], highlight=is_both)

    _auto_width(ws_all)

    # Sheet 2: DB+FE 공통
    ws_both = wb.create_sheet("DB+FE공통")
    _write_header(ws_both, ["Word", "Abbreviation", "English Full", "Korean", "Definition",
                             "DB Count", "FE Count"])
    both = [w for w in enriched if w["db_count"] > 0 and w["fe_count"] > 0]
    for i, w in enumerate(both, 2):
        _write_row(ws_both, i, [
            w["word"], w.get("abbreviation", ""), w.get("english_full", ""),
            w.get("korean", ""), w.get("definition", ""),
            w["db_count"], w["fe_count"],
        ])
    _auto_width(ws_both)

    # Sheet 3: DB Only
    ws_db = wb.create_sheet("DB전용")
    _write_header(ws_db, ["Word", "Abbreviation", "English Full", "Korean", "Definition", "DB Count"])
    db_only = [w for w in enriched if w["db_count"] > 0 and w["fe_count"] == 0]
    for i, w in enumerate(db_only, 2):
        _write_row(ws_db, i, [
            w["word"], w.get("abbreviation", ""), w.get("english_full", ""),
            w.get("korean", ""), w.get("definition", ""), w["db_count"],
        ])
    _auto_width(ws_db)

    # Sheet 4: FE Only
    ws_fe = wb.create_sheet("FE전용")
    _write_header(ws_fe, ["Word", "Abbreviation", "English Full", "Korean", "Definition", "FE Count"])
    fe_only = [w for w in enriched if w["fe_count"] > 0 and w["db_count"] == 0]
    for i, w in enumerate(fe_only, 2):
        _write_row(ws_fe, i, [
            w["word"], w.get("abbreviation", ""), w.get("english_full", ""),
            w.get("korean", ""), w.get("definition", ""), w["fe_count"],
        ])
    _auto_width(ws_fe)

    # Sheet 5: Unknown
    ws_unknown = wb.create_sheet("미식별")
    _write_header(ws_unknown, ["Word", "DB Count", "FE Count", "Sample Sources"])
    for i, w in enumerate(unknown, 2):
        _write_row(ws_unknown, i, [
            w["word"], w["db_count"], w["fe_count"],
            ", ".join(w.get("sample_sources", [])[:3]),
        ])
    _auto_width(ws_unknown)

    wb.save(filepath)
    logger.info("Terms excel saved: %s", filepath)
    return filepath
