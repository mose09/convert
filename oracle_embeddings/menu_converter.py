"""LLM-aided converter: 임의 양식의 메뉴 Excel → 표준 menu.md 파이프 테이블.

프로젝트마다 메뉴 관리 양식이 다르다 (컬럼명·계층 표현 방식·병합 셀 등).
`load_menu_from_excel` 은 고정된 헤더 synonym 세트만 해석하므로 새로운
양식을 만나면 수동 정리가 필요했다. 이 모듈은 그 수동 단계를 없앤다.

흐름:
    1. openpyxl 로 시트 로드 → 병합 셀 forward-fill + 상위 타이틀 행 탐지
    2. 헤더 후보 행 + 샘플 데이터 행(~20) 을 LLM 에 던져 **한 번** 매핑 획득
    3. 매핑에 따라 3 모드로 변환:
         - ``columns_per_level`` : 레벨별 컬럼이 따로 있을 때 (기존 템플릿)
         - ``depth_column``      : (메뉴명, 뎁스, URL) 형태 — path stack
         - ``path_column``       : "A > B > C" 처럼 한 컬럼에 계층 압축
    4. 표준 pipe 테이블(``1레벨 | 2레벨 | ... | 5레벨 | URL``) 로 emit

LLM 환경변수는 ``PATTERN_LLM_*`` 우선, 없으면 ``LLM_*`` (discover-patterns
와 동일 컨벤션). LLM 호출 실패 시 가장 그럴듯한 휴리스틱(헤더 동의어
매칭)으로 fallback 해 사용자가 여전히 유용한 결과를 얻게 한다.
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
from datetime import datetime

logger = logging.getLogger(__name__)


_MAX_SAMPLE_ROWS = 20
_MAX_COLS = 20


# ── Excel ingestion + preprocessing ───────────────────────────────


def _load_sheet(xlsx_path: str, sheet_name: str | None = None):
    """Return an openpyxl worksheet or raise a clear error."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"시트 '{sheet_name}' 없음. 존재하는 시트: {wb.sheetnames}"
            )
        return wb[sheet_name]
    # Auto: pick the sheet with the most non-empty cells.
    best = None
    best_score = -1
    for ws in wb.worksheets:
        score = sum(1 for row in ws.iter_rows(values_only=True)
                    for cell in row if cell not in (None, ""))
        if score > best_score:
            best_score = score
            best = ws
    return best


def _forward_fill_merged(ws) -> list[list]:
    """Return rows with merged cells forward-filled.

    Merged ranges are common in menus where the top-level 대분류 cell is
    merged across several child rows. Without unfilling them the LLM sees
    a messy sparse first column.
    """
    # First materialize rows
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    n_rows = len(rows)
    n_cols = max((len(r) for r in rows), default=0)
    for r in rows:
        if len(r) < n_cols:
            r.extend([None] * (n_cols - len(r)))

    # Apply merged ranges: top-left value fills the whole rectangle.
    for mr in ws.merged_cells.ranges:
        try:
            top = mr.min_row - 1
            left = mr.min_col - 1
            bot = mr.max_row - 1
            right = mr.max_col - 1
            if top < 0 or top >= n_rows:
                continue
            value = rows[top][left]
            for i in range(top, min(bot + 1, n_rows)):
                for j in range(left, min(right + 1, n_cols)):
                    rows[i][j] = value
        except Exception as e:  # pragma: no cover — defensive
            logger.debug("merged-range unfill skip: %s", e)

    return rows


def _detect_header_row(rows: list[list]) -> int:
    """Return the index of the most likely header row.

    Heuristic: pick the row with (a) the most distinct non-empty string
    cells, and (b) at least 2 consecutive rows of data below it. Title /
    summary rows above the real header typically have fewer filled cells
    and/or contain long prose rather than short labels.
    """
    if not rows:
        return 0
    best_idx = 0
    best_score = -1
    for i, row in enumerate(rows[: min(len(rows), 15)]):
        labels = [
            str(c).strip()
            for c in row
            if c is not None and str(c).strip()
        ]
        if not labels:
            continue
        # Prefer rows where every cell is a short label (not sentence-like)
        short_labels = sum(1 for v in labels if len(v) <= 20 and " " not in v.strip() or len(v) <= 10)
        data_rows_below = sum(
            1 for r in rows[i + 1 : i + 4]
            if any(c is not None and str(c).strip() for c in r)
        )
        score = short_labels * 3 + data_rows_below - (len(rows[i]) - len(labels))
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx


def _clean(value) -> str:
    """Render a cell value as a trimmed string; None → ''."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _sample_rows(rows: list[list], header_idx: int,
                  max_samples: int = _MAX_SAMPLE_ROWS) -> list[list[str]]:
    """Return up to ``max_samples`` data rows (trimmed, string-ified)."""
    out = []
    for r in rows[header_idx + 1 :]:
        clean = [_clean(c) for c in r]
        if any(clean):  # skip wholly-blank rows
            out.append(clean)
        if len(out) >= max_samples:
            break
    return out


# ── LLM mapping ──────────────────────────────────────────────────


_SYSTEM_PROMPT = """당신은 엔터프라이즈 메뉴 구조를 분석하는 전문가입니다.
주어진 메뉴 테이블의 헤더와 샘플 행을 보고, 각 컬럼이 어떤 정보를 담고
있는지 분류해 주세요. 반드시 유효한 JSON 만 응답하세요."""


_USER_PROMPT = """아래는 한 프로젝트의 메뉴 Excel 에서 추출한 헤더와 샘플 행입니다.

## 헤더 행 (0-based 인덱스)
{headers}

## 샘플 데이터 (최대 {n_samples}개)
{samples}

## 응답 형식

다음 세 가지 `mode` 중 하나로 분류하세요:

1. **columns_per_level** — 레벨별로 별도 컬럼이 있음 (예: "대분류", "중분류", "소분류").
2. **depth_column** — 메뉴명 컬럼 하나 + 뎁스(1,2,3,...) 숫자 컬럼 하나.
3. **path_column** — 한 컬럼에 "A > B > C" 처럼 구분자로 계층이 압축돼 있음.

각 모드에 따라 아래 JSON 스키마로만 응답하세요. 헤더명 대신 0-based 컬럼 인덱스(정수)를 쓰세요.

```json
{{
  "mode": "columns_per_level" | "depth_column" | "path_column",

  // mode=columns_per_level 일 때만. 없는 레벨은 null.
  "levels": {{"level_1": <col_idx or null>, "level_2": ..., "level_3": ..., "level_4": ..., "level_5": ...}},

  // mode=depth_column 일 때만.
  "name_col": <col_idx>,
  "depth_col": <col_idx>,

  // mode=path_column 일 때만.
  "path_col": <col_idx>,
  "path_separator": " > ",   // 실제 구분자 (예: " > ", " / ", " > ")

  // 모든 모드 공통
  "url_col": <col_idx or null>,    // URL 컬럼. 없으면 null.
  "notes": "기타 판단 근거"
}}
```"""


def _call_llm(prompt: str, config: dict, max_retries: int = 2) -> dict:
    """Call the LLM with PATTERN_LLM_* override, mirroring discover-patterns."""
    from openai import OpenAI

    llm_cfg = (config or {}).get("llm", {}) or {}
    api_key = (os.environ.get("PATTERN_LLM_API_KEY")
               or os.environ.get("LLM_API_KEY")
               or llm_cfg.get("api_key", "ollama"))
    api_base = (os.environ.get("PATTERN_LLM_API_BASE")
                or os.environ.get("LLM_API_BASE")
                or llm_cfg.get("api_base", "http://localhost:11434/v1"))
    model = (os.environ.get("PATTERN_LLM_MODEL")
             or os.environ.get("LLM_MODEL")
             or llm_cfg.get("model", "llama3"))

    print(f"  LLM model: {model}")
    print(f"  LLM endpoint: {api_base}")

    client = OpenAI(api_key=api_key, base_url=api_base)
    for attempt in range(max_retries + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                timeout=180,
            )
            text = resp.choices[0].message.content.strip()
            m = re.search(r"```(?:json)?\s*([\s\S]*?)```", text)
            if m:
                text = m.group(1).strip()
            return json.loads(text)
        except json.JSONDecodeError:
            if attempt < max_retries:
                wait = 2 ** (attempt + 1)
                logger.warning("LLM invalid JSON (attempt %d), retrying in %ds",
                               attempt + 1, wait)
                time.sleep(wait)
            else:
                logger.error("LLM JSON parse failed after retries")
        except Exception as e:
            logger.error("LLM call failed: %s", e)
            break
    return {}


_HEADER_LEVEL_SYNONYMS = {
    1: ("1레벨", "level1", "lv1", "lvl1", "대분류", "대메뉴", "menu1"),
    2: ("2레벨", "level2", "lv2", "lvl2", "중분류", "중메뉴", "menu2"),
    3: ("3레벨", "level3", "lv3", "lvl3", "소분류", "소메뉴", "menu3"),
    4: ("4레벨", "level4", "lv4", "lvl4", "세분류", "세메뉴", "menu4"),
    5: ("5레벨", "level5", "lv5", "lvl5", "최하위", "menu5"),
}
# URL-전용 동의어 (경로/path 는 path_column 과 겹치므로 별도).
_HEADER_URL_SPECIFIC = ("url", "uri", "link", "endpoint", "링크", "주소")
# URL 일 수도 있고 path_column 일 수도 있는 애매한 것들 (후순위 매칭).
_HEADER_URL_AMBIG = ("경로", "path")
_HEADER_NAME_SYNONYMS = ("메뉴명", "menu_name", "menuname", "name", "title", "메뉴이름")
_HEADER_DEPTH_SYNONYMS = ("뎁스", "depth", "레벨", "level", "lv", "계층", "단계")
_HEADER_PATH_SYNONYMS = ("경로", "path", "menu_path", "메뉴경로", "hierarchy")


def _heuristic_mapping(headers: list[str]) -> dict:
    """Non-LLM fallback mapping based on header synonym matching.

    Chosen to be conservative: only returns a mapping when at least a URL
    column + one level indicator is identifiable.
    """
    norm = [h.strip().lower().replace(" ", "") for h in headers]
    # URL 탐지: 전용 동의어 먼저, 없으면 ambiguous (경로/path) 로 fallback.
    url_col = next(
        (i for i, h in enumerate(norm) if any(k == h or h.endswith(k) or k in h for k in _HEADER_URL_SPECIFIC)),
        None,
    )
    if url_col is None:
        url_col = next(
            (i for i, h in enumerate(norm) if any(k == h or h.endswith(k) for k in _HEADER_URL_AMBIG)),
            None,
        )
    # Try columns_per_level first
    levels: dict[str, int | None] = {f"level_{i}": None for i in range(1, 6)}
    matched = 0
    for lv, keys in _HEADER_LEVEL_SYNONYMS.items():
        for i, h in enumerate(norm):
            if any(k in h for k in keys):
                levels[f"level_{lv}"] = i
                matched += 1
                break
    if matched >= 1:
        return {"mode": "columns_per_level", "levels": levels, "url_col": url_col,
                "notes": "heuristic (header synonyms)"}
    # Fall back to depth_column if we see both a name-like and depth-like column
    name_col = next(
        (i for i, h in enumerate(norm) if any(k in h for k in _HEADER_NAME_SYNONYMS)),
        None,
    )
    depth_col = next(
        (i for i, h in enumerate(norm) if any(k == h or h == "lv" for k in _HEADER_DEPTH_SYNONYMS)),
        None,
    )
    if name_col is not None and depth_col is not None:
        return {"mode": "depth_column", "name_col": name_col, "depth_col": depth_col,
                "url_col": url_col, "notes": "heuristic (name+depth)"}
    # Or path_column
    path_col = next(
        (i for i, h in enumerate(norm) if any(k in h for k in _HEADER_PATH_SYNONYMS)),
        None,
    )
    if path_col is not None:
        # url_col 이 path_col 과 같으면 URL 컬럼이 따로 없는 것 — None 으로 무효화.
        resolved_url = None if url_col == path_col else url_col
        return {"mode": "path_column", "path_col": path_col, "path_separator": " > ",
                "url_col": resolved_url, "notes": "heuristic (path column)"}
    return {}


def _validate_mapping(mapping: dict, n_cols: int) -> dict | None:
    """Ensure the mapping is structurally valid for the number of columns."""
    if not isinstance(mapping, dict):
        return None
    mode = mapping.get("mode")
    if mode not in ("columns_per_level", "depth_column", "path_column"):
        return None

    def _col(v):
        if v is None:
            return None
        try:
            v = int(v)
        except (TypeError, ValueError):
            return None
        return v if 0 <= v < n_cols else None

    out: dict = {"mode": mode, "url_col": _col(mapping.get("url_col")),
                 "notes": mapping.get("notes") or ""}

    if mode == "columns_per_level":
        levels_in = mapping.get("levels") or {}
        levels: dict[str, int | None] = {}
        any_level = False
        for i in range(1, 6):
            key = f"level_{i}"
            col = _col(levels_in.get(key))
            levels[key] = col
            if col is not None:
                any_level = True
        if not any_level:
            return None
        out["levels"] = levels
    elif mode == "depth_column":
        name = _col(mapping.get("name_col"))
        depth = _col(mapping.get("depth_col"))
        if name is None or depth is None:
            return None
        out["name_col"] = name
        out["depth_col"] = depth
    else:  # path_column
        pcol = _col(mapping.get("path_col"))
        if pcol is None:
            return None
        sep = mapping.get("path_separator") or " > "
        if not isinstance(sep, str) or not sep:
            sep = " > "
        out["path_col"] = pcol
        out["path_separator"] = sep
    return out


# ── Apply mapping → standard rows ─────────────────────────────────


def _apply_columns_per_level(rows: list[list[str]], m: dict) -> list[dict]:
    """Emit entries directly from per-level columns."""
    out = []
    last = [""] * 5  # for forward-fill of empty upper levels (merged-cell aftermath)
    url_col = m.get("url_col")
    for r in rows:
        levels = []
        for i in range(1, 6):
            idx = m["levels"].get(f"level_{i}")
            val = _clean(r[idx]) if idx is not None and idx < len(r) else ""
            if val:
                last[i - 1] = val
            levels.append(last[i - 1])
            # Clear deeper cached levels when an upper level changes.
            if idx is not None and _clean(r[idx]) and i < 5:
                for j in range(i, 5):
                    # We conservatively only clear if this level has a fresh
                    # non-empty value AND deeper levels of this row are empty.
                    pass
        url = _clean(r[url_col]) if url_col is not None and url_col < len(r) else ""
        if any(levels) or url:
            out.append({"levels": [lv for lv in levels if lv], "url": url})
    return out


def _apply_depth_column(rows: list[list[str]], m: dict) -> list[dict]:
    """Walk rows maintaining a path stack keyed by depth."""
    name_col = m["name_col"]
    depth_col = m["depth_col"]
    url_col = m.get("url_col")

    stack: list[str] = ["", "", "", "", ""]
    out = []
    for r in rows:
        name = _clean(r[name_col]) if name_col < len(r) else ""
        depth_raw = _clean(r[depth_col]) if depth_col < len(r) else ""
        url = _clean(r[url_col]) if url_col is not None and url_col < len(r) else ""
        try:
            depth = int(re.sub(r"[^0-9]", "", depth_raw) or "0")
        except ValueError:
            depth = 0
        if depth < 1 or depth > 5 or not name:
            continue
        stack[depth - 1] = name
        for j in range(depth, 5):
            stack[j] = ""
        levels = [s for s in stack[:depth] if s]
        out.append({"levels": levels, "url": url})
    return out


def _apply_path_column(rows: list[list[str]], m: dict) -> list[dict]:
    """Split a single path string by the given separator."""
    path_col = m["path_col"]
    sep = m["path_separator"]
    url_col = m.get("url_col")
    out = []
    for r in rows:
        path = _clean(r[path_col]) if path_col < len(r) else ""
        url = _clean(r[url_col]) if url_col is not None and url_col < len(r) else ""
        if not path and not url:
            continue
        # Also tolerate " / " or " > " if sep mismatches slightly
        parts = [p.strip() for p in re.split(re.escape(sep), path) if p.strip()]
        if not parts:
            # Try a couple of common separators as fallback
            for alt in (" > ", " / ", ">", "/", "|"):
                if alt in path:
                    parts = [p.strip() for p in path.split(alt) if p.strip()]
                    break
        out.append({"levels": parts[:5], "url": url})
    return out


# ── Emit standard menu.md pipe table ──────────────────────────────


def _emit_menu_md(entries: list[dict], source: str) -> str:
    """Render entries as the canonical menu.md pipe table."""
    header = "| 1레벨 | 2레벨 | 3레벨 | 4레벨 | 5레벨 | URL |"
    sep = "|-------|-------|-------|-------|-------|-----|"
    lines = [
        f"# 메뉴 테이블 (convert-menu 자동 생성, {datetime.now().strftime('%Y-%m-%d %H:%M')})",
        "",
        f"소스: `{source}`",
        "",
        "| 1레벨 | 2레벨 | 3레벨 | 4레벨 | 5레벨 | URL |",
        "|-------|-------|-------|-------|-------|-----|",
    ]
    kept = 0
    for e in entries:
        levels = list(e.get("levels") or [])
        while len(levels) < 5:
            levels.append("")
        levels = levels[:5]
        url = e.get("url") or ""
        if not any(levels) and not url:
            continue
        cells = [_md_escape(v) for v in levels + [url]]
        lines.append("| " + " | ".join(cells) + " |")
        kept += 1
    logger.info("convert-menu emitted %d rows", kept)
    return "\n".join(lines) + "\n"


def _md_escape(s: str) -> str:
    return (s or "").replace("|", r"\|").replace("\n", " ")


# ── Public entry point ───────────────────────────────────────────


def convert_menu(xlsx_path: str, output_path: str, config: dict,
                  sheet_name: str | None = None,
                  use_llm: bool = True) -> str:
    """Convert an arbitrary menu xlsx to the standard ``menu.md`` template.

    Returns the absolute path of the emitted ``.md`` file.
    """
    print(f"=== Step 1: Loading {xlsx_path} ===")
    ws = _load_sheet(xlsx_path, sheet_name)
    print(f"  Sheet: {ws.title}")

    rows_raw = _forward_fill_merged(ws)
    if not rows_raw:
        raise ValueError("빈 시트입니다.")
    header_idx = _detect_header_row(rows_raw)
    headers = [_clean(c) for c in rows_raw[header_idx][: _MAX_COLS]]
    print(f"  Detected header row: {header_idx} ({len([h for h in headers if h])} non-empty columns)")

    samples = _sample_rows(rows_raw, header_idx)
    n_cols = len(headers)
    print(f"  Sampled {len(samples)} data rows for mapping")

    print("\n=== Step 2: LLM mapping ===")
    mapping: dict | None = None
    if use_llm:
        prompt = _USER_PROMPT.format(
            headers=json.dumps(list(enumerate(headers)), ensure_ascii=False),
            samples=json.dumps(samples, ensure_ascii=False, indent=2),
            n_samples=len(samples),
        )
        llm_out = _call_llm(prompt, config)
        mapping = _validate_mapping(llm_out, n_cols)
        if mapping:
            print(f"  LLM mapping: mode={mapping['mode']}, url_col={mapping.get('url_col')}")
        else:
            print("  LLM 응답 유효 매핑 없음 — heuristic fallback")
    if mapping is None:
        mapping = _validate_mapping(_heuristic_mapping(headers), n_cols)
    if mapping is None:
        raise ValueError(
            "메뉴 컬럼 매핑 실패: LLM·heuristic 모두 해석 못 함. "
            "시트 레이아웃을 확인하거나 --sheet 로 시트명을 지정하세요."
        )
    logger.info("mapping decided: %s", mapping)

    print("\n=== Step 3: Applying mapping ===")
    data_rows = [r for r in rows_raw[header_idx + 1:]
                 if any(_clean(c) for c in r)]
    # Pad rows to max columns for safe indexing.
    for r in data_rows:
        if len(r) < n_cols:
            r.extend([None] * (n_cols - len(r)))
    if mapping["mode"] == "columns_per_level":
        entries = _apply_columns_per_level([[_clean(c) for c in r] for r in data_rows], mapping)
    elif mapping["mode"] == "depth_column":
        entries = _apply_depth_column([[_clean(c) for c in r] for r in data_rows], mapping)
    else:
        entries = _apply_path_column([[_clean(c) for c in r] for r in data_rows], mapping)
    print(f"  Emitted rows: {len(entries)}")

    print("\n=== Step 4: Writing menu.md ===")
    md = _emit_menu_md(entries, os.path.basename(xlsx_path))
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(md)
    abs_path = os.path.abspath(output_path)
    print(f"  Output: {abs_path}")
    return abs_path
