"""recommend-names 결정적 코어 회귀 테스트.

LLM/임베딩 없이 동작하는 Tier1·2 만 검증한다.
실행: ``python tests/test_recommend_names.py`` 또는
      ``python -m unittest tests.test_recommend_names``
"""

import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from oracle_embeddings.std_dict import (  # noqa: E402
    _is_std_value,
    build_std_dict,
    compose_data_type,
    ensure_std_dict,
    load_std_dict,
    norm_kor,
)
from oracle_embeddings.tobe_recommender import recommend_column  # noqa: E402


def _make_word_dict(path):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["논리명", "물리명", "물리의미(영문풀네임)", "표준여부(Y,N)",
               "속성분류어(Y,N)", "동의어", "설명", "만료일자", "출처구분"])
    for r in [
        ("고객", "CUST", "CUSTOMER", "Y", "N", "손님,거래처", "", "", "표준"),
        ("주문", "ORD", "ORDER", "Y", "N", "오더", "", "", "표준"),
        ("금액", "AMT", "AMOUNT", "Y", "Y", "", "", "", "표준"),
        ("번호", "NO", "NUMBER", "Y", "Y", "", "", "", "표준"),
        ("등록", "REG", "REGISTER", "Y", "N", "", "", "20200101", "폐기"),  # 만료
    ]:
        ws.append(r)
    wb.save(path)


def _make_term_dict(path):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.append(["논리명", "물리명", "구성정보", "물리의미", "도메인명", "데이터유형",
               "길이", "소수점", "표준여부(Y,N)", "개인정보구분", "암호화여부",
               "설명", "만료일자", "출처구분"])
    for r in [
        ("고객번호", "CUST_NO", "고객+번호", "", "번호", "VARCHAR2", "20", "",
         "Y", "Y", "N", "", "", "표준"),
        ("주문금액", "ORD_AMT", "주문+금액", "", "금액", "NUMBER", "15", "2",
         "Y", "N", "N", "", "", "표준"),
    ]:
        ws.append(r)
    wb.save(path)


class RecommendCoreTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.mkdtemp()
        wpath = os.path.join(cls.tmp, "word.xlsx")
        tpath = os.path.join(cls.tmp, "term.xlsx")
        _make_word_dict(wpath)
        _make_term_dict(tpath)
        cls.sd = ensure_std_dict(os.path.join(cls.tmp, "dict.sqlite"), wpath, tpath)

    def test_norm_and_type(self):
        self.assertEqual(norm_kor(" 고객 번호 "), "고객번호")
        self.assertEqual(compose_data_type("VARCHAR2", "20", ""), "VARCHAR2(20)")
        self.assertEqual(compose_data_type("NUMBER", "15", "2"), "NUMBER(15,2)")
        self.assertEqual(compose_data_type("DATE", "", ""), "DATE")

    def test_term_exact_match(self):
        rec = recommend_column("T", "CUST_NO", "고객번호", self.sd)
        self.assertEqual(rec.tier, "정확매칭(용어)")
        self.assertEqual(rec.tobe_name, "CUST_NO")
        self.assertEqual(rec.data_type, "VARCHAR2(20)")
        self.assertEqual(rec.confidence, 1.0)

    def test_word_composition(self):
        rec = recommend_column("T", "ORD_NO", "주문번호", self.sd)
        self.assertEqual(rec.tobe_name, "ORD_NO")
        self.assertTrue(rec.tier.startswith("단어조합"))
        self.assertEqual(rec.data_type, "VARCHAR2(20)")  # 번호 분류어 추론

    def test_synonym_match(self):
        # '손님' 은 '고객' 의 동의어
        rec = recommend_column("T", "GUEST_AMT", "손님금액", self.sd)
        self.assertEqual(rec.tobe_name, "CUST_AMT")

    def test_expired_word_excluded(self):
        # '등록' 단어는 만료 → 미매칭 단편으로 남고 표준 약어로 조합 안 됨
        rec = recommend_column("T", "REG_NO", "등록번호", self.sd)
        self.assertIn("등록", rec.unmatched_frags)
        self.assertNotIn("REG", rec.tobe_name.split("_"))

    def test_unmatched_no_comment(self):
        rec = recommend_column("T", "XYZ123", None, self.sd)
        self.assertEqual(rec.tier, "미매칭")
        self.assertEqual(rec.confidence, 0.0)


class SheetDetectionTest(unittest.TestCase):
    """표지 시트 + 개행/번호 헤더 자동 인식 회귀."""

    def test_cover_sheet_skipped_and_header_normalized(self):
        from openpyxl import Workbook

        from oracle_embeddings.std_dict import _classify_header, _pick_sheet
        tmp = tempfile.mkdtemp()
        wb = Workbook()
        cover = wb.active
        cover.title = "표지"
        cover.append(["표준 단어사전", "v1"])
        ws = wb.create_sheet("단어목록")
        ws.append(["1. 논리\n명", "물리명 ", "표준여부(Y,N)"])
        ws.append(["고객", "CUST", "Y"])
        path = os.path.join(tmp, "w.xlsx")
        wb.save(path)

        name, _, idx, _ = _pick_sheet(path, None)
        self.assertEqual(name, "단어목록")
        self.assertIn("logical", idx)
        self.assertIn("physical", idx)
        self.assertEqual(_classify_header("1. 논리\n명"), "logical")
        self.assertEqual(_classify_header("물리명 "), "physical")

    def test_picks_data_sheet_over_near_empty(self):
        # 거의 빈 "Sheet 1"(1셀) 보다 실제 데이터 시트를 선택해야 함
        from openpyxl import Workbook

        from oracle_embeddings.std_dict import _pick_sheet
        tmp = tempfile.mkdtemp()
        wb = Workbook()
        s1 = wb.active
        s1.title = "Sheet 1"
        s1["A1"] = "논리명"  # 거의 빈 시트
        s2 = wb.create_sheet("단어사전")
        s2.append(["논리명", "물리명", "표준여부"])
        for i in range(30):
            s2.append([f"단어{i}", f"W{i}", "Y"])
        path = os.path.join(tmp, "multi.xlsx")
        wb.save(path)

        name, rows, idx, _ = _pick_sheet(path, None)
        self.assertEqual(name, "단어사전")
        self.assertEqual(len(rows), 31)
        self.assertIn("physical", idx)

    def test_nfd_unicode_header_and_data(self):
        # macOS 등에서 생성된 NFD(조합형) 한글도 인식돼야 함
        import unicodedata

        from oracle_embeddings.std_dict import _classify_header, norm_kor
        self.assertEqual(_classify_header(unicodedata.normalize("NFD", "논리명")),
                         "logical")
        self.assertEqual(_classify_header(unicodedata.normalize("NFD", "물리명")),
                         "physical")
        self.assertEqual(norm_kor(unicodedata.normalize("NFD", "고객번호")),
                         norm_kor("고객번호"))

    def test_decorated_and_prefixed_headers(self):
        from oracle_embeddings.std_dict import _classify_header
        for h, expect in [
            ("논리명*", "logical"),
            ("표준단어 논리명", "logical"),
            ("컬럼한글명", "logical"),
            ("물리명 ※", "physical"),
            ("표준단어 물리명", "physical"),
            ("물리의미(영문풀네임)", "eng"),  # 물리명 으로 오분류되면 안 됨
        ]:
            self.assertEqual(_classify_header(h), expect, h)


class WordAbbrAndStdTest(unittest.TestCase):
    """표준여부 표기 다양성 + 중복 논리명 약어 해결 회귀."""

    def test_is_std_value_lenient(self):
        for v in ("Y", "표준", "○", "사용", "", "예", "O"):
            self.assertTrue(_is_std_value(v), v)
        for v in ("N", "X", "×", "비표준", "폐기"):
            self.assertFalse(_is_std_value(v), v)

    def _build(self, rows):
        from openpyxl import Workbook
        tmp = tempfile.mkdtemp()
        wb = Workbook(); ws = wb.active
        ws.append(["논리명", "물리명", "표준여부", "속성분류어"])
        for r in rows:
            ws.append(r)
        wp = os.path.join(tmp, "w.xlsx"); wb.save(wp)
        db = os.path.join(tmp, "sd.sqlite")
        build_std_dict(db, word_xlsx=wp)
        return load_std_dict(db)

    def test_duplicate_logical_first_empty_physical(self):
        # 같은 '확인' 이 두 번: 첫 행 물리명 비었고 둘째 행 CHK → CHK 써야 함
        sd = self._build([["확인", "", "Y", "N"], ["확인", "CHK", "Y", "N"],
                          ["여부", "YN", "Y", "Y"]])
        self.assertEqual(sd.resolve_word("확인"), ("확인", "CHK", "논리명"))
        rec = recommend_column("T", "X1", "확인여부", sd)
        self.assertEqual(rec.tobe_name, "CHK_YN")

    def test_nonstd_text_still_resolves_abbr(self):
        # 표준여부 가 '표준' 텍스트여도 약어 사용
        sd = self._build([["처리", "PROC", "표준", "N"]])
        rec = recommend_column("T", "X", "처리", sd)
        self.assertEqual(rec.tobe_name, "PROC")

    def test_unmatched_fragment_wrapped(self):
        sd = self._build([["번호", "NO", "Y", "Y"]])
        rec = recommend_column("T", "X", "미지단어번호", sd)
        self.assertIn("«미지단어»", rec.tobe_name)
        self.assertTrue(rec.tobe_name.endswith("_NO"))

    def test_comment_marker_stripped(self):
        sd = self._build([["고객", "CUST", "Y", "N"]])
        rec = recommend_column("T", "X", "고객(LLM추천)", sd)
        self.assertEqual(rec.tobe_name, "CUST")


class DomainDictTest(unittest.TestCase):
    def test_domain_multi_entry_and_type(self):
        from openpyxl import Workbook
        tmp = tempfile.mkdtemp()
        wb = Workbook(); ws = wb.active
        ws.append(["도메인그룹명", "도메인명", "데이터유형", "길이", "소수점",
                   "개인정보구분", "암호화여부", "설명", "만료일자", "출처구분"])
        ws.append(["금액", "금액", "NUMBER", "15", "2", "N", "N", "", "", ""])
        ws.append(["외화금액", "금액", "NUMBER", "18", "3", "N", "N", "", "", ""])
        ws.append(["율", "율", "NUMBER", "5", "2", "N", "N", "", "", ""])
        dp = os.path.join(tmp, "dom.xlsx"); wb.save(dp)
        db = os.path.join(tmp, "sd.sqlite")
        st = build_std_dict(db, domain_xlsx=dp)
        self.assertEqual(st["domains"], 3)
        sd = load_std_dict(db)
        self.assertTrue(sd.has_domains())
        # 동일 도메인명 '금액' 2개 보존
        self.assertEqual(len(sd.domain_by_name[norm_kor("금액")]), 2)
        dtype, single = sd.resolve_domain_type("금액")
        self.assertEqual(dtype, "NUMBER(15,2)")  # 최빈/첫값
        self.assertFalse(single)                 # 다중이라 단일 아님


class PriorityTest(unittest.TestCase):
    """6단계 우선순위: 표준Y[논리명>동의어>물리의미] > 표준N[...]."""

    def _build(self, rows):
        from openpyxl import Workbook
        tmp = tempfile.mkdtemp()
        wb = Workbook(); ws = wb.active
        ws.append(["논리명", "물리명", "물리의미", "표준여부", "속성분류어",
                   "동의어", "설명", "만료일자", "출처구분"])
        for r in rows:
            ws.append(r)
        wp = os.path.join(tmp, "w.xlsx"); wb.save(wp)
        db = os.path.join(tmp, "sd.sqlite")
        build_std_dict(db, word_xlsx=wp)
        return load_std_dict(db)

    def test_primary_logical_beats_synonym(self):
        # 시설: 자기 논리명(FACI) vs 공장의 동의어(FAC) → 논리명 우선
        sd = self._build([
            ["시설", "FACI", "FACILITY", "Y", "N", "", "", "", ""],
            ["공장", "FAC", "FACTORY", "Y", "N", "시설", "", "", ""],
        ])
        self.assertEqual(sd.resolve_word("시설"), ("시설", "FACI", "논리명"))

    def test_physical_meaning_matches_english(self):
        # 영문 풀네임 코멘트 → 물리의미 매칭
        sd = self._build([["설명", "DESC", "DESCRIPTION", "Y", "N", "", "", "", ""]])
        self.assertEqual(sd.resolve_word("Description"), ("설명", "DESC", "물리의미"))
        rec = recommend_column("T", "X", "Detail Description", sd)
        self.assertTrue(rec.tobe_name.endswith("DESC"))

    def test_standard_tier_beats_nonstandard(self):
        # 같은 키 '확인': 표준N 논리명(CHK) vs 표준Y 동의어(소유=점검,INSP)
        sd = self._build([
            ["확인", "CHK", "", "N", "N", "", "", "", ""],
            ["점검", "INSP", "", "Y", "N", "확인", "", "", ""],
        ])
        # 표준Y 동의어가 표준N 논리명을 이긴다
        self.assertEqual(sd.resolve_word("확인"), ("점검", "INSP", "동의어"))

    def test_synonym_comma_separated(self):
        sd = self._build([["고객", "CUST", "", "Y", "N", "손님, 거래처", "", "", ""]])
        self.assertEqual(sd.resolve_word("거래처")[1], "CUST")
        self.assertEqual(sd.resolve_word("손님")[1], "CUST")


if __name__ == "__main__":
    unittest.main(verbosity=2)
