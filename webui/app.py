"""로컬 실행용 Streamlit 화면 — 비개발자가 CLI 없이 마이그레이션 / 용어검색.

실행:
    streamlit run webui/app.py
    (또는 루트의 run_webui.bat 더블클릭)

두 화면:
  1) SQL 마이그레이션 — AS-IS mapper XML + 매핑 YAML → 변환 XML + 리포트 zip
  2) 용어 검색 — build-dict 로 적재한 표준사전 SQLite 를 LIKE 검색

기존 CLI/함수를 그대로 재사용한다:
  - 마이그레이션은 ``python main.py migrate-sql`` 를 subprocess 로 호출
    (동작 100% 동일 보장). 출력은 임시 폴더로 격리 후 zip.
  - 용어 검색은 표준사전 SQLite 를 직접 조회.
폐쇄망/오프라인 전제 — 외부 네트워크 호출 없음.
"""
from __future__ import annotations

import glob
import io
import os
import sqlite3
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path

import streamlit as st

try:
    import yaml
except Exception:  # pragma: no cover
    yaml = None

ROOT = Path(__file__).resolve().parent.parent  # 프로젝트 루트
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

st.set_page_config(page_title="SQL 마이그레이션 & 용어검색",
                   page_icon="🛠️", layout="wide")


# ---------------------------------------------------------------------------
# 공통 헬퍼
# ---------------------------------------------------------------------------
def _default_dict_db() -> Path:
    """config.yaml 의 vectordb.db_path 기준 표준사전 SQLite 경로."""
    db_dir = "./vectordb"
    cfg = ROOT / "config.yaml"
    if yaml and cfg.is_file():
        try:
            data = yaml.safe_load(cfg.read_text(encoding="utf-8")) or {}
            db_dir = (data.get("vectordb") or {}).get("db_path", db_dir)
        except Exception:
            pass
    return (ROOT / db_dir / "standard_dict.sqlite").resolve()


def _write_temp_config(out_dir: Path) -> Path:
    """기존 config.yaml 복사 + storage.output_dir 만 임시 폴더로 덮어써서
    마이그레이션 산출물을 격리한다 (사용자 output/ 오염 방지)."""
    data = {}
    cfg = ROOT / "config.yaml"
    if yaml and cfg.is_file():
        try:
            data = yaml.safe_load(cfg.read_text(encoding="utf-8")) or {}
        except Exception:
            data = {}
    data.setdefault("storage", {})["output_dir"] = str(out_dir)
    tmp = out_dir / "_webui_config.yaml"
    tmp.write_text(yaml.safe_dump(data, allow_unicode=True) if yaml
                   else f'storage:\n  output_dir: "{out_dir}"\n',
                   encoding="utf-8")
    return tmp


# ---------------------------------------------------------------------------
# .env 읽기/쓰기 (설정 화면)
# ---------------------------------------------------------------------------
def _env_path() -> Path:
    return ROOT / ".env"


def _read_env() -> dict:
    """현재 ``.env`` (없으면 ``.env.example``) 를 KEY=VALUE dict 로."""
    path = _env_path()
    src = path if path.is_file() else (ROOT / ".env.example")
    env: dict = {}
    if src.is_file():
        for line in src.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if not s or s.startswith("#") or "=" not in s:
                continue
            k, v = s.split("=", 1)
            env[k.strip()] = v.strip()
    return env


def _write_env(updates: dict) -> None:
    """``.env`` 의 알려진 키만 갱신하고 나머지 줄/주석은 보존. 없던 키는 추가."""
    path = _env_path()
    seen: set = set()
    out_lines: list = []
    if path.is_file():
        for line in path.read_text(encoding="utf-8").splitlines():
            s = line.strip()
            if s and not s.startswith("#") and "=" in s:
                k = s.split("=", 1)[0].strip()
                if k in updates:
                    out_lines.append(f"{k}={updates[k]}")
                    seen.add(k)
                    continue
            out_lines.append(line)
    for k, v in updates.items():
        if k not in seen:
            out_lines.append(f"{k}={v}")
    path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")


# .env 필드 그룹 — (그룹명, [(키, 라벨, 비밀번호여부), ...])
_ENV_GROUPS = [
    ("Oracle AS-IS", [
        ("ORACLE_USER", "사용자", False),
        ("ORACLE_PASSWORD", "비밀번호", True),
        ("ORACLE_DSN", "DSN (host:port/service)", False),
        ("ORACLE_SCHEMA_OWNER", "스키마 owner", False),
        ("ORACLE_INSTANT_CLIENT_DIR", "Instant Client 경로 (thick)", False),
    ]),
    ("Oracle TO-BE (validate-migration Stage B)", [
        ("ORACLE_TOBE_DSN", "TO-BE DSN", False),
        ("ORACLE_TOBE_USER", "TO-BE 사용자", False),
        ("ORACLE_TOBE_PASSWORD", "TO-BE 비밀번호", True),
    ]),
    ("LLM (enrich-schema / terms / standardize 등)", [
        ("LLM_API_BASE", "API Base", False),
        ("LLM_API_KEY", "API Key", True),
        ("LLM_MODEL", "모델", False),
    ]),
    ("임베딩 (erd-rag / recommend-names RAG)", [
        ("EMBEDDING_API_BASE", "API Base", False),
        ("EMBEDDING_API_KEY", "API Key", True),
        ("EMBEDDING_MODEL", "모델", False),
    ]),
    ("패턴/코딩 LLM (discover-patterns / biz 추출 / 화면변환)", [
        ("PATTERN_LLM_API_BASE", "API Base", False),
        ("PATTERN_LLM_API_KEY", "API Key", True),
        ("PATTERN_LLM_MODEL", "모델", False),
    ]),
]


def render_settings() -> None:
    st.header("⚙️ 설정 (LLM / DB 환경)")
    st.caption("`.env` 파일을 편집합니다. 값은 이 PC 로컬에만 저장되고 외부로 "
               "전송되지 않습니다. 저장 후 실행 중인 화면/CLI 에 반영됩니다.")
    if not _env_path().is_file():
        st.info("`.env` 가 없어 `.env.example` 기본값을 표시합니다. 저장하면 "
                "`.env` 가 생성됩니다.")

    env = _read_env()
    updates: dict = {}
    with st.form("settings_form"):
        for group, fields in _ENV_GROUPS:
            st.subheader(group)
            cols = st.columns(2)
            for i, (key, label, secret) in enumerate(fields):
                updates[key] = cols[i % 2].text_input(
                    f"{label}  ·  `{key}`", value=env.get(key, ""),
                    type="password" if secret else "default",
                    key=f"env_{key}")
        saved = st.form_submit_button("💾 저장", type="primary")
    if saved:
        # 빈 값도 그대로 저장 (사용자가 지운 것 반영)
        _write_env(updates)
        st.success(f"저장됨: {_env_path()}")


# ---------------------------------------------------------------------------
# 화면: 매핑 만들기 (9컬럼 → MD)
# ---------------------------------------------------------------------------
_MAPPING_HEADERS = [
    "asis_table", "asis_column", "asis_column_type",
    "tobe_table", "tobe_table_comment", "tobe_column",
    "tobe_column_type", "tobe_column_comment", "remark",
]
_MAPPING_EXAMPLE = [
    ["CUST", "CUST_ID", "NUMBER(10)", "CUSTOMER", "고객 마스터",
     "CUSTOMER_ID", "NUMBER(10)", "고객ID", ""],
    ["CUST", "REG_DT", "VARCHAR2(8)", "CUSTOMER", "고객 마스터",
     "REGISTER_DATE", "DATE", "등록일자", "YYYYMMDD → DATE"],
    ["CUST", "OBSOLETE_FLAG", "CHAR(1)", "CUSTOMER", "고객 마스터",
     "", "", "", "TO-BE 에서 삭제"],
]


def _macro_bas_text() -> str:
    p = Path(__file__).resolve().parent / "assets" / "mapping_macro.bas"
    return p.read_text(encoding="utf-8") if p.is_file() else ""


def _rows_to_md(rows: list) -> str:
    """9컬럼 행 리스트 → convert-mapping 용 마크다운 표 텍스트."""
    def esc(v):
        return str(v if v is not None else "").replace("|", "\\|").replace(
            "\r", " ").replace("\n", " ").strip()
    out = ["| " + " | ".join(_MAPPING_HEADERS) + " |",
           "|" + "------|" * len(_MAPPING_HEADERS)]
    for r in rows:
        cells = [esc(r[i]) if i < len(r) else "" for i in range(len(_MAPPING_HEADERS))]
        if any(c for c in cells):
            out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out) + "\n"


def _build_mapping_xlsx() -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "매핑"
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    for j, h in enumerate(_MAPPING_HEADERS, start=1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hdr_fill
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 20
    for i, row in enumerate(_MAPPING_EXAMPLE, start=2):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)

    guide = wb.create_sheet("매크로_설치법")
    steps = [
        "AS-IS→TO-BE 컬럼 매핑 작성 → MD 내보내기",
        "",
        "[1] '매핑' 시트 2행부터 값을 채웁니다 (예시 3행 참고).",
        "    - tobe_column 을 비우면 해당 컬럼은 삭제(drop) 로 처리됩니다.",
        "",
        "[2] 매크로 설치 (최초 1회):",
        "    1) Alt+F11 (Visual Basic 편집기)",
        "    2) 파일 > 파일 가져오기 > mapping_macro.bas 선택",
        "       (또는 아래 소스를 새 모듈에 붙여넣기)",
        "    3) 통합문서를 'Excel 매크로 사용 통합문서(*.xlsm)' 로 저장",
        "",
        "[3] 사용: 개발도구 > 매크로 > ExportMappingMd 실행",
        "    → 통합문서와 같은 폴더에 column_mapping.md 생성 (UTF-8).",
        "    이 .md 를 마이그레이션 화면 입력으로 사용하세요.",
        "",
        "──────── 매크로 소스 (mapping_macro.bas) ────────",
    ]
    row = 1
    for s in steps:
        guide.cell(row=row, column=1, value=s)
        row += 1
    for line in _macro_bas_text().splitlines():
        guide.cell(row=row, column=1, value=line)
        row += 1
    guide.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_mapping_maker() -> None:
    st.header("📝 매핑 만들기 (AS-IS → TO-BE)")
    st.caption("9컬럼 엑셀에 매핑을 채워 convert-mapping 용 `column_mapping.md` "
               "를 만듭니다. 이후 마이그레이션 화면 입력으로 사용합니다.")

    st.subheader("① 엑셀 템플릿 + 매크로 (권장)")
    c1, c2 = st.columns(2)
    c1.download_button("⬇ 매핑 템플릿 (.xlsx)", data=_build_mapping_xlsx(),
                       file_name="mapping_template.xlsx",
                       mime="application/vnd.openxmlformats-officedocument."
                            "spreadsheetml.sheet")
    c2.download_button("⬇ VBA 매크로 (.bas)", data=_macro_bas_text(),
                       file_name="mapping_macro.bas", mime="text/plain")
    st.markdown(
        "- 템플릿의 **매핑** 시트에 값을 채우고, **매크로_설치법** 시트의 "
        "안내대로 매크로를 1회 임포트 → `.xlsm` 로 저장.\n"
        "- 이후 **개발도구 ▸ 매크로 ▸ `ExportMappingMd`** 실행하면 같은 "
        "폴더에 `column_mapping.md` 가 생성됩니다 (UTF-8, 한글 OK).")

    st.divider()
    st.subheader("② (매크로 없이) 채운 엑셀 → MD + YAML 바로 변환")
    up = st.file_uploader("작성한 매핑 엑셀 (.xlsx) — '매핑' 시트 9컬럼",
                          type=["xlsx"])
    if up is not None:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(up.getbuffer()), data_only=True)
            ws = wb["매핑"] if "매핑" in wb.sheetnames else wb.active
            rows = [[c.value for c in row]
                    for row in ws.iter_rows(min_row=2, max_col=len(_MAPPING_HEADERS))]
            md = _rows_to_md(rows)
        except Exception as e:  # noqa: BLE001
            st.error(f"엑셀 읽기 실패: {type(e).__name__}: {e}")
        else:
            n = md.count("\n") - 2
            st.success(f"{max(n, 0)} 행 변환됨.")
            st.download_button("⬇ column_mapping.md 내려받기", data=md,
                               file_name="column_mapping.md",
                               mime="text/markdown")
            with st.expander("MD 미리보기"):
                st.code(md, language="markdown")
            st.markdown("**→ 마이그레이션 입력용 YAML 로 변환**")
            _render_md_to_yaml(md, "mm_xlsx_yaml")

    st.divider()
    st.subheader("③ 기존 MD → 매핑 YAML (매크로로 만든 .md 등)")
    st.caption("VBA 매크로로 만든 `column_mapping.md` 를 마이그레이션 입력용 "
               "YAML 로 변환합니다 (convert-mapping).")
    mdup = st.file_uploader("column_mapping.md 업로드", type=["md"])
    if mdup is not None:
        # utf-8-sig: 매크로/메모장 UTF-8 저장의 BOM 제거 (없어도 안전)
        md_text = mdup.getvalue().decode("utf-8-sig")
        with st.expander("MD 미리보기"):
            st.code(md_text, language="markdown")
        _render_md_to_yaml(md_text, "mm_md_yaml")


def _md_to_yaml(md_text: str, no_llm: bool):
    """MD (9컬럼 flat) → column_mapping.yaml (convert-mapping subprocess).

    Returns ``(yaml_text | None, log)``."""
    work = Path(tempfile.mkdtemp(prefix="webui_map_"))
    mp = work / "column_mapping.md"
    mp.write_text(md_text, encoding="utf-8")
    yp = work / "column_mapping.yaml"
    cmd = [sys.executable, str(ROOT / "main.py"), "convert-mapping",
           "--mapping-md", str(mp), "--output", str(yp)]
    if no_llm:
        cmd += ["--no-llm"]
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(ROOT))
    log = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
    yaml_text = yp.read_text(encoding="utf-8") if yp.is_file() else None
    return yaml_text, log


def _render_md_to_yaml(md_text: str, state_key: str) -> None:
    """MD→YAML 변환 버튼 + 결과(다운로드/미리보기/검증) 렌더 (섹션 ②/③ 공유)."""
    no_llm = st.checkbox("LLM 사용 안 함 (heuristic 만)", value=True,
                         key=f"{state_key}_nollm",
                         help="폐쇄망/LLM 미설정이면 켜두세요. 끄면 설정의 LLM "
                              "으로 split/merge 등 복잡 케이스를 더 잘 추론.")
    if st.button("▶ MD → 매핑 YAML 변환", key=f"{state_key}_btn"):
        with st.spinner("변환 중…"):
            yaml_text, log = _md_to_yaml(md_text, no_llm)
        st.session_state[state_key] = {"yaml": yaml_text, "log": log}

    res = st.session_state.get(state_key)
    if res:
        if res["yaml"]:
            val = next((l.strip() for l in res["log"].splitlines()
                        if "Validation" in l), "")
            st.success(f"YAML 생성 완료. {val}")
            st.download_button("⬇ column_mapping.yaml 내려받기",
                               data=res["yaml"],
                               file_name="column_mapping.yaml",
                               mime="text/yaml", key=f"{state_key}_dl")
            with st.expander("YAML 미리보기"):
                st.code(res["yaml"], language="yaml")
        else:
            st.error("변환 실패 — 로그를 확인하세요.")
        with st.expander("변환 로그"):
            st.code(res["log"] or "(로그 없음)")


# ---------------------------------------------------------------------------
# 화면 1: SQL 마이그레이션
# ---------------------------------------------------------------------------
def _migrate_flags(schema_from_mapping: bool, schema_path, emit_comments: bool,
                   no_validate: bool, llm_fallback: bool) -> list:
    flags = ["--output-format", "excel,xml"]
    if schema_from_mapping:
        flags += ["--to-be-schema-from-mapping"]
    elif schema_path:
        flags += ["--to-be-schema", str(schema_path)]
    if emit_comments:
        flags += ["--emit-column-comments"]
    if no_validate:
        flags += ["--no-validate"]
    if llm_fallback:
        flags += ["--llm-fallback"]
    return flags


def _run_migrate(in_dir, map_path, out_dir, flags):
    """migrate-sql 을 subprocess 로 실행. output_dir 을 out_dir 로 격리.

    Returns ``(returncode, log, dated_dir | None, summary)``."""
    cfg = _write_temp_config(Path(out_dir))
    # ``--config`` 는 전역 인자라 subcommand 앞.
    cmd = [sys.executable, str(ROOT / "main.py"), "--config", str(cfg),
           "migrate-sql", "--mybatis-dir", str(in_dir),
           "--mapping", str(map_path)] + flags
    proc = subprocess.run(cmd, capture_output=True, text=True, cwd=str(ROOT))
    log = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
    dated = sorted(glob.glob(str(Path(out_dir) / "migration" / "*")))
    summary = next((l.strip() for l in log.splitlines()
                    if l.startswith("Converted ")), "")
    return proc.returncode, log, (Path(dated[-1]) if dated else None), summary


def render_migration() -> None:
    st.header("🛠️ SQL 마이그레이션")
    mode = st.radio("실행 방식", ["파일 업로드 (결과 zip 다운로드)",
                                  "폴더 지정 (지정 폴더로 출력)"],
                    horizontal=True)
    st.divider()
    if mode.startswith("파일 업로드"):
        _render_migration_upload()
    else:
        _render_migration_folder()


def _schema_options():
    """공통: TO-BE 스키마 라디오 + 옵션 체크박스. (from_mapping, emit, no_validate, llm)."""
    schema_mode = st.radio(
        "TO-BE 스키마",
        ["매핑 YAML 에서 자동 파생 (DB/스키마 없이)", "TO-BE 스키마 .md 지정"],
        help="스키마가 아직 없으면 매핑에서 파생. pass-through 컬럼은 Stage A "
             "검증에서 오탐될 수 있으나 변환 XML 은 정상입니다.")
    c1, c2, c3 = st.columns(3)
    emit = c1.checkbox("한글 주석 삽입", value=False, help="--emit-column-comments")
    no_validate = c2.checkbox("Stage A 검증 건너뛰기", value=True,
                              help="--no-validate")
    llm = c3.checkbox("LLM 보조 변환", value=False, help="--llm-fallback")
    return schema_mode.startswith("매핑 YAML"), emit, no_validate, llm


def _render_migration_upload() -> None:
    st.caption("mapper XML(여러 개) + 매핑 YAML 업로드 → 변환 → 결과 zip.")
    mappers = st.file_uploader("① AS-IS mapper XML (여러 개)", type=["xml"],
                               accept_multiple_files=True)
    mapping = st.file_uploader("② 컬럼 매핑 YAML", type=["yaml", "yml"])
    from_mapping, emit, no_validate, llm = _schema_options()
    schema_md = None
    if not from_mapping:
        schema_md = st.file_uploader("TO-BE 스키마 .md", type=["md"])

    ready = bool(mappers) and mapping is not None and (from_mapping or schema_md)
    if st.button("▶ 변환 실행", type="primary", disabled=not ready):
        with st.spinner("변환 중…"):
            work = Path(tempfile.mkdtemp(prefix="webui_mig_"))
            in_dir = work / "mapper"
            in_dir.mkdir()
            for f in mappers:
                (in_dir / f.name).write_bytes(f.getbuffer())
            map_path = work / "mapping.yaml"
            map_path.write_bytes(mapping.getbuffer())
            out_dir = work / "out"
            out_dir.mkdir()
            schema_path = None
            if schema_md is not None:
                schema_path = work / "to_be_schema.md"
                schema_path.write_bytes(schema_md.getbuffer())
            flags = _migrate_flags(from_mapping, schema_path, emit,
                                   no_validate, llm)
            rc, log, dated, summary = _run_migrate(in_dir, map_path, out_dir, flags)
            zbuf = io.BytesIO()
            if dated:
                with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                    for p in dated.rglob("*"):
                        if p.is_file():
                            zf.write(p, p.relative_to(dated))
        st.session_state["mig_result"] = {
            "ok": rc == 0 and dated is not None,
            "summary": summary or "변환 완료 (요약 없음 — 로그 확인)",
            "zip": zbuf.getvalue(), "log": log}
        st.rerun()

    res = st.session_state.get("mig_result")
    if res:
        if res["ok"]:
            st.success(res["summary"])
            st.download_button("⬇ 변환 결과 zip", data=res["zip"],
                               file_name="migration_result.zip",
                               mime="application/zip")
        else:
            st.error("변환 실패")
        with st.expander("실행 로그"):
            st.code(res["log"] or "(로그 없음)")


def _render_migration_folder() -> None:
    st.caption("로컬 폴더 경로를 지정합니다. 이 PC 에서 직접 접근하므로 대량 "
               "변환에 적합합니다.")
    in_dir = st.text_input("① AS-IS mapper 폴더 경로 (하위 .xml 재귀)")
    map_path = st.text_input("② 컬럼 매핑 YAML 경로")
    out_dir = st.text_input("③ 출력 폴더 경로",
                            help="여기 아래 migration/<날짜>/ 로 결과가 떨어집니다.")
    from_mapping, emit, no_validate, llm = _schema_options()
    schema_path = None
    if not from_mapping:
        schema_path = st.text_input("TO-BE 스키마 .md 경로")

    def _bad(p, is_dir):
        return not p or (not Path(p).is_dir() if is_dir else not Path(p).is_file())

    ready = not _bad(in_dir, True) and not _bad(map_path, False) and bool(out_dir) \
        and (from_mapping or not _bad(schema_path, False))
    if st.button("▶ 변환 실행", type="primary", disabled=not ready):
        with st.spinner("변환 중…"):
            Path(out_dir).mkdir(parents=True, exist_ok=True)
            flags = _migrate_flags(from_mapping, schema_path, emit,
                                   no_validate, llm)
            rc, log, dated, summary = _run_migrate(
                in_dir, map_path, out_dir, flags)
        if rc == 0 and dated is not None:
            st.success(summary or "변환 완료")
            st.info(f"결과 폴더: {dated}")
            files = sorted(str(p.relative_to(dated))
                           for p in dated.rglob("*") if p.is_file())
            st.write("생성 파일:", files)
        else:
            st.error("변환 실패 — 로그를 확인하세요.")
        with st.expander("실행 로그"):
            st.code(log or "(로그 없음)")


# ---------------------------------------------------------------------------
# 화면 2: 용어 검색
# ---------------------------------------------------------------------------
_SEARCH_SPECS = {
    "용어": ("term",
             ["logical", "physical", "eng", "domain", "data_type",
              "length", "scale", "is_std", "desc"],
             ["logical", "physical", "eng", "desc"]),
    "단어": ("word",
             ["logical", "physical", "eng", "is_std", "is_classifier",
              "synonyms", "desc"],
             ["logical", "physical", "eng", "synonyms", "desc"]),
    "도메인": ("domain",
              ["grp", "name", "data_type", "length", "scale", "full_type",
               "desc"],
              ["name", "grp", "desc"]),
}


def render_term_search() -> None:
    st.header("🔎 용어 검색")
    st.caption("build-dict 로 적재한 표준 단어/용어/도메인 사전(SQLite)을 "
               "검색합니다.")

    db_path = Path(st.text_input("표준사전 SQLite 경로",
                                 value=str(_default_dict_db())))
    if not db_path.is_file():
        st.warning(f"표준사전이 없습니다: {db_path}\n\n"
                   "먼저 `python main.py build-dict --word-dict ... "
                   "--term-dict ...` 로 적재하세요.")
        return

    c1, c2 = st.columns([3, 1])
    q = c1.text_input("검색어 (한글 논리명 / 영문 / 물리명 일부)")
    target = c2.radio("대상", list(_SEARCH_SPECS.keys()), horizontal=False)

    include_expired = st.checkbox("만료 항목 포함", value=False)

    if not q.strip():
        st.info("검색어를 입력하세요.")
        return

    table, cols, search_cols = _SEARCH_SPECS[target]
    where = " OR ".join(f"{c} LIKE ?" for c in search_cols)
    params = [f"%{q.strip()}%"] * len(search_cols)
    sql = f"SELECT {', '.join(cols)} FROM {table} WHERE ({where})"
    if not include_expired and table in ("word", "term", "domain"):
        sql += " AND COALESCE(expired, 0) = 0"
    sql += " LIMIT 500"

    try:
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        rows = [dict(r) for r in conn.execute(sql, params)]
        conn.close()
    except Exception as e:  # noqa: BLE001
        st.error(f"검색 실패: {type(e).__name__}: {e}")
        return

    st.write(f"**{len(rows)}건** (최대 500)")
    if rows:
        st.dataframe(rows, use_container_width=True, hide_index=True)
    else:
        st.info("일치하는 항목이 없습니다.")


# ---------------------------------------------------------------------------
# 화면: 표준사전 적재 (build-dict → SQLite)
# ---------------------------------------------------------------------------
def render_build_dict() -> None:
    st.header("📚 표준사전 적재 (단어/용어/도메인 → SQLite)")
    st.caption("표준 사전 Excel 을 SQLite 로 적재합니다. 적재 후 **용어 검색** "
               "화면에서 조회 가능. **기존 내용은 삭제 후 재적재**됩니다.")

    word = st.file_uploader("단어사전 (.xlsx)", type=["xlsx"])
    term = st.file_uploader("용어사전 (.xlsx)", type=["xlsx"])
    domain = st.file_uploader("도메인사전 (.xlsx, 선택)", type=["xlsx"])
    no_embed = st.checkbox(
        "임베딩 건너뛰기 (--no-embed)", value=True,
        help="용어검색은 임베딩이 불필요합니다. 임베딩 API 미설정 시 켜두세요 "
             "(recommend-names 의 RAG 를 쓸 때만 임베딩 필요).")

    db_path = _default_dict_db()
    st.caption(f"저장 위치: `{db_path}`")

    ready = word is not None or term is not None or domain is not None
    if st.button("▶ 적재 실행", type="primary", disabled=not ready):
        with st.spinner("적재 중…"):
            work = Path(tempfile.mkdtemp(prefix="webui_dict_"))
            cmd = [sys.executable, str(ROOT / "main.py"), "build-dict",
                   "--dict-db", str(db_path)]
            for label, f, flag in [("word", word, "--word-dict"),
                                   ("term", term, "--term-dict"),
                                   ("domain", domain, "--domain-dict")]:
                if f is not None:
                    p = work / f"{label}.xlsx"
                    p.write_bytes(f.getbuffer())
                    cmd += [flag, str(p)]
            if no_embed:
                cmd += ["--no-embed"]
            db_path.parent.mkdir(parents=True, exist_ok=True)
            proc = subprocess.run(cmd, capture_output=True, text=True,
                                  cwd=str(ROOT))
        if proc.returncode == 0:
            st.success("적재 완료 — '용어 검색' 화면에서 조회하세요.")
        else:
            st.error("적재 실패 — 로그를 확인하세요.")
        st.code((proc.stdout or "") + (
            "\n" + proc.stderr if proc.stderr else ""))


# ---------------------------------------------------------------------------
# 공통: 산출물 위치 찾기 + 명령 결과 표시
# ---------------------------------------------------------------------------
def _latest_output_file(area: str, exts: tuple):
    """``output/<area>/<최신날짜>/`` 에서 확장자 매칭 최신(mtime) 파일."""
    base = ROOT / "output" / area
    if not base.is_dir():
        return None
    for d in sorted((p for p in base.glob("*") if p.is_dir()), reverse=True):
        files = [p for p in d.iterdir() if p.is_file() and p.suffix in exts]
        if files:
            return max(files, key=lambda p: p.stat().st_mtime)
    return None


def _show_cmd_result(proc, area, exts, session_key, label):
    log = (proc.stdout or "") + ("\n" + proc.stderr if proc.stderr else "")
    if proc.returncode == 0:
        f = _latest_output_file(area, exts)
        if f:
            st.success(f"{label} 생성: {f}")
            if session_key:
                st.session_state[session_key] = str(f)
            st.download_button(f"⬇ {f.name}", data=f.read_bytes(),
                               file_name=f.name)
            return f
        st.warning(f"{label} 명령은 끝났으나 산출물을 못 찾았습니다 — 로그 확인.")
    else:
        st.error(f"{label} 실패 — DB/LLM 접속·설정('설정' 화면)을 확인하세요.")
    with st.expander("실행 로그", expanded=proc.returncode != 0):
        st.code(log or "(로그 없음)")
    return None


# ---------------------------------------------------------------------------
# 화면: 스키마 추출 (schema — Oracle → .md)
# ---------------------------------------------------------------------------
def render_schema_extract() -> None:
    st.header("📤 스키마 추출 (Oracle → .md)")
    st.caption("Oracle DB 에 접속해 테이블/컬럼/PK/FK/인덱스를 .md 로 추출합니다. "
               "접속 정보는 **설정** 화면(.env)에서 지정하세요.")
    fmt = st.radio("형식", ["markdown", "txt"], horizontal=True)
    c1, c2 = st.columns(2)
    owner = c1.text_input("스키마 owner (선택)",
                          help="미지정 시 .env 의 ORACLE_SCHEMA_OWNER")
    table = c2.text_input("특정 테이블만 (선택)")
    if st.button("▶ 추출 실행", type="primary"):
        with st.spinner("추출 중… (DB 접속)"):
            cmd = [sys.executable, str(ROOT / "main.py"), "schema",
                   "--format", fmt]
            if owner.strip():
                cmd += ["--owner", owner.strip()]
            if table.strip():
                cmd += ["--table", table.strip()]
            proc = subprocess.run(cmd, capture_output=True, text=True,
                                  cwd=str(ROOT))
        _show_cmd_result(proc, "schema", (".md", ".txt"),
                         "last_schema_md", "스키마")


# ---------------------------------------------------------------------------
# 화면: 코멘트 증강 (enrich-schema — LLM)
# ---------------------------------------------------------------------------
def render_enrich_schema() -> None:
    st.header("💬 스키마 코멘트 증강 (LLM)")
    st.caption("빈 테이블/컬럼 코멘트를 LLM 이 약어를 해석해 채웁니다. LLM 접속은 "
               "**설정** 화면(.env).")
    default = st.session_state.get("last_schema_md", "")
    path = st.text_input("스키마 .md 경로", value=default,
                         help="스키마 추출 화면에서 만든 파일 경로가 자동 채워집니다.")
    up = st.file_uploader("또는 스키마 .md 업로드", type=["md"])
    ready = bool(path.strip()) or up is not None
    if st.button("▶ 증강 실행", type="primary", disabled=not ready):
        with st.spinner("증강 중… (LLM 호출)"):
            if up is not None:
                work = Path(tempfile.mkdtemp(prefix="webui_enrich_"))
                schema_md = work / up.name
                schema_md.write_bytes(up.getbuffer())
            else:
                schema_md = Path(path.strip())
            cmd = [sys.executable, str(ROOT / "main.py"), "enrich-schema",
                   "--schema-md", str(schema_md)]
            proc = subprocess.run(cmd, capture_output=True, text=True,
                                  cwd=str(ROOT))
        _show_cmd_result(proc, "enrich-schema", (".md",),
                         "last_schema_md", "증강 스키마")


# ---------------------------------------------------------------------------
# 화면: ERD 추출 (query + erd-md / erd-group)
# ---------------------------------------------------------------------------
def render_erd_extract() -> None:
    st.header("🔗 ERD 추출")
    st.caption("스키마 .md (+ 선택: mapper XML 폴더의 JOIN 관계) → 인터랙티브 "
               "HTML ERD. 생성 후 **ERD 보기** 화면에서 열람.")
    default = st.session_state.get("last_schema_md", "")
    schema_path = st.text_input("스키마 .md 경로", value=default)
    schema_up = st.file_uploader("또는 스키마 .md 업로드", type=["md"])
    mapper_dir = st.text_input(
        "mapper XML 폴더 경로 (선택 — 쿼리 JOIN 관계 추가)",
        help="지정하면 query 로 JOIN 관계를 뽑아 ERD 에 반영합니다.")
    erd_type = st.radio("ERD 종류",
                        ["erd-md (단일)", "erd-group (주제영역 분할)"],
                        horizontal=True)
    related_only, tables = False, ""
    if erd_type.startswith("erd-md"):
        c1, c2 = st.columns(2)
        related_only = c1.checkbox("관계 있는 테이블만 (--related-only)")
        tables = c2.text_input("특정 테이블만 (쉼표, 선택)")

    ready = bool(schema_path.strip()) or schema_up is not None
    if st.button("▶ ERD 생성", type="primary", disabled=not ready):
        with st.spinner("생성 중…"):
            work = Path(tempfile.mkdtemp(prefix="webui_erd_"))
            if schema_up is not None:
                schema_md = work / "schema.md"
                schema_md.write_bytes(schema_up.getbuffer())
            else:
                schema_md = Path(schema_path.strip())

            query_md = None
            if mapper_dir.strip() and Path(mapper_dir.strip()).is_dir():
                subprocess.run(
                    [sys.executable, str(ROOT / "main.py"), "query",
                     mapper_dir.strip(), "--schema-md", str(schema_md)],
                    capture_output=True, text=True, cwd=str(ROOT))
                query_md = _latest_output_file("query", (".md",))

            sub = "erd-md" if erd_type.startswith("erd-md") else "erd-group"
            cmd = [sys.executable, str(ROOT / "main.py"), sub,
                   "--schema-md", str(schema_md)]
            if query_md:
                cmd += ["--query-md", str(query_md)]
            if sub == "erd-md":
                if related_only:
                    cmd += ["--related-only"]
                if tables.strip():
                    cmd += ["--tables", tables.strip()]
            proc = subprocess.run(cmd, capture_output=True, text=True,
                                  cwd=str(ROOT))
        f = _show_cmd_result(proc, "erd", (".html",), "last_erd_html", "ERD HTML")
        if f:
            st.info("→ **ERD 보기** 화면에서 열 수 있습니다.")


# ---------------------------------------------------------------------------
# 화면: ERD 보기 (생성된 HTML 임베드)
# ---------------------------------------------------------------------------
def render_erd_view() -> None:
    import streamlit.components.v1 as components
    st.header("👁️ ERD 보기")
    last = st.session_state.get("last_erd_html", "")
    src = st.radio("ERD 소스", ["최근 생성", "파일 경로", "업로드"],
                   horizontal=True)
    html, name = None, None
    if src == "최근 생성":
        if last and Path(last).is_file():
            html, name = Path(last).read_text(encoding="utf-8"), Path(last).name
        else:
            st.info("최근 생성된 ERD 가 없습니다. 'ERD 추출' 을 먼저 실행하거나 "
                    "다른 소스를 선택하세요.")
    elif src == "파일 경로":
        p = st.text_input("ERD HTML 경로", value=last)
        if p.strip() and Path(p.strip()).is_file():
            html, name = Path(p.strip()).read_text(encoding="utf-8"), Path(p.strip()).name
        elif p.strip():
            st.warning("파일을 찾을 수 없습니다.")
    else:
        up = st.file_uploader("ERD HTML 업로드", type=["html"])
        if up is not None:
            html, name = up.getvalue().decode("utf-8"), up.name

    if html:
        st.caption(f"표시 중: `{name}`")
        components.html(html, height=800, scrolling=True)
        st.caption("⚠ 폐쇄망에서는 ERD 의 D3(CDN)가 안 열려 인터랙티브가 "
                   "제한될 수 있습니다 — d3.v7.min.js 를 HTML 옆에 로컬 반입하면 "
                   "정상 동작합니다.")


# ---------------------------------------------------------------------------
# 라우팅
# ---------------------------------------------------------------------------
_PAGES = {
    "스키마 추출": render_schema_extract,
    "코멘트 증강": render_enrich_schema,
    "ERD 추출": render_erd_extract,
    "ERD 보기": render_erd_view,
    "표준사전 적재": render_build_dict,
    "용어 검색": render_term_search,
    "매핑 만들기": render_mapping_maker,
    "SQL 마이그레이션": render_migration,
    "설정": render_settings,
}


def main() -> None:
    st.sidebar.title("메뉴")
    page = st.sidebar.radio("화면", list(_PAGES.keys()),
                            label_visibility="collapsed")
    st.sidebar.markdown("---")
    st.sidebar.caption("로컬 실행 · 오프라인 · 외부 전송 없음")
    _PAGES[page]()


main()
